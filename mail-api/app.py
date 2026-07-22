from __future__ import annotations

import csv
import io
import json
import hashlib
import hmac
import os
import secrets
import smtplib
import tempfile
import threading
import time
import math
from email.message import EmailMessage
from urllib.parse import urlencode, urlparse, urlunparse, parse_qsl
from urllib.request import Request, urlopen

from dotenv import load_dotenv
from flask import Flask, g, has_request_context, jsonify, redirect, request, send_file, send_from_directory
from werkzeug.exceptions import HTTPException
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, ".env"))
DATA_DIR = os.getenv("PORTAL_DATA_DIR", BASE_DIR).strip() or BASE_DIR
FRONTEND_PUBLIC_DIR = os.getenv("PORTAL_FRONTEND_DIR", os.path.join(BASE_DIR, "public")).strip() or os.path.join(BASE_DIR, "public")
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
PG_LOCK_KEY = 2026071501
SESSION_COOKIE_NAME = "susu_session"
TRUSTED_DEVICE_COOKIE_NAME = "susu_trusted_device"
TRUSTED_DEVICE_TTL_SECONDS = 30 * 24 * 60 * 60
PORTAL_AUTHORIZATION_MINUTES = 10

OFFICIAL_EMAIL_DOMAIN = "@bawjiasecommunitybank.com"
PRESENCE_STORE_PATH = os.path.join(DATA_DIR, "presence_store.json")
PASSWORD_STORE_PATH = os.path.join(DATA_DIR, "password_store.json")
USERS_STORE_PATH = os.path.join(DATA_DIR, "users_store.json")
PENDING_VERIFICATIONS_PATH = os.path.join(DATA_DIR, "pending_verifications.json")
RESET_TOKENS_PATH = os.path.join(DATA_DIR, "reset_tokens.json")
AGENT_SETUP_TOKENS_PATH = os.path.join(DATA_DIR, "agent_setup_tokens.json")
PRIVILEGED_MFA_PATH = os.path.join(DATA_DIR, "privileged_mfa.json")
TRUSTED_DEVICES_PATH = os.path.join(DATA_DIR, "trusted_devices.json")
AUTH_RATE_LIMITS_PATH = os.path.join(DATA_DIR, "auth_rate_limits.json")
SESSIONS_STORE_PATH = os.path.join(DATA_DIR, "sessions_store.json")
PORTAL_AUTHORIZATIONS_PATH = os.path.join(DATA_DIR, "portal_authorizations.json")
NOTIFICATIONS_STORE_PATH = os.path.join(DATA_DIR, "notifications_store.json")
AUDIT_LOGS_STORE_PATH = os.path.join(DATA_DIR, "audit_logs_store.json")
PORTAL_SETTINGS_STORE_PATH = os.path.join(DATA_DIR, "portal_settings_store.json")
CUSTOMERS_STORE_PATH = os.path.join(DATA_DIR, "customers_store.json")
CUSTOMER_IMPORTS_STORE_PATH = os.path.join(DATA_DIR, "customer_imports_store.json")
COLLECTIONS_STORE_PATH = os.path.join(DATA_DIR, "collections_store.json")
DAILY_CLOSES_STORE_PATH = os.path.join(DATA_DIR, "daily_closes_store.json")
UPLOADS_DIR = os.path.join(DATA_DIR, "uploads")
PRESENCE_TTL_SECONDS = 20
ONLINE_WINDOW_SECONDS = 20
RESET_TOKEN_TTL_SECONDS = 30 * 60
VERIFICATION_TTL_SECONDS = 15 * 60
SESSION_TTL_SECONDS = 30 * 24 * 60 * 60
RATE_LIMIT_WINDOW_SECONDS = 15 * 60
RATE_LIMIT_MAX_ATTEMPTS = 8
DEFAULT_PORTAL_BRANCHES = [
    "HEAD OFFICE",
    "BAWJIASE",
    "ADEISO",
    "OFAAKOR",
    "KASOA NEW MARKET",
    "KASOA MAIN",
]
DEFAULT_PORTAL_DEPARTMENTS = [
    "SUSU",
]
REQUIRED_PORTAL_DEPARTMENTS = ["SUSU"]
SUSU_DEPARTMENTS = {"SUSU", "SUSU AGENT", "SUSU SUPERVISOR"}
DEFAULT_PORTAL_SETTINGS = {
    "bankName": "Bawjiase Community Bank PLC",
    "shortBankName": "BCB",
    "portalName": "SUSU Collection Portal",
    "emailDomain": OFFICIAL_EMAIL_DOMAIN,
    "branches": DEFAULT_PORTAL_BRANCHES,
    "departments": DEFAULT_PORTAL_DEPARTMENTS,
    "formCategories": [],
    "departmentChangeTypes": [],
    "transferLocations": [],
    "loginSubtitle": "Sign in to manage SUSU collections, customers, staff, and branch reports.",
    "loginButtonText": "Secure Login",
    "authorizedAccessText": "Authorized Access Only",
    "appMode": "test",
    "publicRegistrationEnabled": False,
    "itAccessCode": "",
    "hrAccessCode": "",
    "sessionMinutes": 30,
    "absoluteSessionHours": 12,
    "sensitiveReauthMinutes": 15,
    "verificationMinutes": 15,
    "passwordResetMinutes": 30,
    "auditRetentionDays": 2555,
    "notificationRetentionDays": 90,
    "verificationRetentionHours": 24,
    "expiredSessionRetentionDays": 7,
    "securityReviewStatus": "not-scheduled",
    "securityReviewProvider": "",
    "securityReviewReference": "",
    "securityReviewDate": "",
    "dashboardLabel": "Dashboard",
    "formsLabel": "",
    "profileLabel": "Profile",
    "activeStaffLabel": "Active Staff",
    "branchCoverageLabel": "Branch Coverage",
    "openOperationsLabel": "Open Operations",
    "resolutionRateLabel": "Resolution Rate",
}

TEST_CUSTOMER_SEED_ROWS = [
    ("TEST AMA MENSAH", "1310000100001", "0240000001", "BAWJIASE"),
    ("TEST KWAME ADJEI", "1310000100002", "0240000002", "BAWJIASE"),
    ("TEST EFUA BOATENG", "1310000100003", "0240000003", "OFAAKOR"),
    ("TEST KOJO APPIAH", "1310000100004", "0240000004", "OFAAKOR"),
    ("TEST ABENA OWUSU", "1310000100005", "0240000005", "ADEISO"),
    ("TEST YAW ASANTE", "1310000100006", "0240000006", "ADEISO"),
    ("TEST AKOSUA DARKO", "1310000100007", "0240000007", "HEAD OFFICE"),
    ("TEST KOFI SARPONG", "1310000100008", "0240000008", "KASOA MAIN"),
]


def env_secret(name: str) -> str:
    return str(os.getenv(name, "") or "").strip()


DEFAULT_INITIAL_PASSWORD = env_secret("PORTAL_DEFAULT_INITIAL_PASSWORD")
IT_ACCESS_CODE = env_secret("IT_ACCESS_CODE")
HR_ACCESS_CODE = env_secret("HR_ACCESS_CODE")
GLOBAL_MANAGER_ROLES = {"OwnerAdmin"}
OWNER_ADMIN_ROLE = "OwnerAdmin"

OWNER_ADMIN_USER = {
    "id": "owner-admin-1",
    "fullname": "Site Creator Owner",
    "phone": "0000000000",
    "email": "sitecreator@bawjiasecommunitybank.com",
    "role": OWNER_ADMIN_ROLE,
    "position": "Site Creator",
    "department": "SUSU",
    "branch": "HEAD OFFICE",
    "imageFile": None,
    "managedBranches": ["ALL"],
    "managedDepartmentsByBranch": {},
    "permissions": {
        "userManagement": True,
        "customers": True,
        "transactions": True,
        "reports": True,
        "agents": True,
        "branches": True,
        "auditLog": True,
        "backupExport": True,
    },
    "isActive": True,
    "isVerified": True,
    "lastSeen": 0,
    "registrationTime": 0,
    "isArchived": False,
}

INITIAL_USERS = [
    {
        "id": "db-user-6",
        "fullname": "Desmond Tettey Quarshie",
        "phone": "0243670230",
        "email": "dquarshie@bawjiasecommunitybank.com",
        "role": "GeneralStaff",
        "position": "Staff",
        "department": "SUSU",
        "branch": "BAWJIASE",
        "imageFile": None,
        "isActive": True,
        "isVerified": True,
        "lastSeen": 1772637593885,
        "registrationTime": 0,
        "isArchived": False,
    },
    {
        "id": "db-user-9",
        "fullname": "Jane Afua Bruku",
        "phone": "0248154869",
        "email": "jbruku@bawjiasecommunitybank.com",
        "role": "Supervisor",
        "position": "Staff",
        "department": "SUSU",
        "branch": "HEAD OFFICE",
        "imageFile": None,
        "isActive": True,
        "isVerified": True,
        "lastSeen": 1770741882598,
        "registrationTime": 0,
        "isArchived": False,
    },
    {
        "id": "db-user-5",
        "fullname": "Kwabena Asare",
        "phone": "0599779664",
        "email": "kasare@bawjiasecommunitybank.com",
        "role": "GeneralStaff",
        "position": "Staff",
        "department": "SUSU",
        "branch": "HEAD OFFICE",
        "imageFile": None,
        "isActive": True,
        "isVerified": True,
        "lastSeen": 1770990814598,
        "registrationTime": 0,
        "isArchived": False,
    },
    {
        "id": "db-user-8",
        "fullname": "Kwesi Adu Snr Yeenu-Prah",
        "phone": "0555443053",
        "email": "kyeenu-prah@bawjiasecommunitybank.com",
        "role": "GeneralStaff",
        "position": "Staff",
        "department": "SUSU",
        "branch": "HEAD OFFICE",
        "imageFile": "profile_pics/f658de3c2aa8ca6d.jpeg",
        "isActive": True,
        "isVerified": True,
        "lastSeen": 1770296150530,
        "registrationTime": 0,
        "isArchived": False,
    },
    {
        "id": "db-user-4",
        "fullname": "Ato Asiedu Mensah",
        "phone": "0247554428",
        "email": "amensah@bawjiasecommunitybank.com",
        "role": "Supervisor",
        "position": "Staff",
        "department": "SUSU",
        "branch": "HEAD OFFICE",
        "imageFile": None,
        "isActive": True,
        "isVerified": True,
        "lastSeen": 1770975614364,
        "registrationTime": 0,
        "isArchived": False,
    },
    {
        "id": "db-user-2",
        "fullname": "James Lincoln Awuah",
        "phone": "0536799490",
        "email": "lawuah@bawjiasecommunitybank.com",
        "role": "GeneralStaff",
        "position": "Staff",
        "department": "SUSU",
        "branch": "HEAD OFFICE",
        "imageFile": "profile_pics/88efb134d068db11.jpg",
        "isActive": True,
        "isVerified": True,
        "lastSeen": 1775309044811,
        "registrationTime": 0,
        "isArchived": False,
    },
    {
        "id": "db-user-3",
        "fullname": "Nathaniel Oglie Narh",
        "phone": "0246377830",
        "email": "nnarh@bawjiasecommunitybank.com",
        "role": "Supervisor",
        "position": "Staff",
        "department": "SUSU",
        "branch": "HEAD OFFICE",
        "imageFile": None,
        "isActive": True,
        "isVerified": True,
        "lastSeen": 1769519876185,
        "registrationTime": 0,
        "isArchived": False,
    },
    {
        "id": "db-user-7",
        "fullname": "GABRIEL OWUSU",
        "phone": "0246315586",
        "email": "gowusu@bawjiasecommunitybank.com",
        "role": "GeneralStaff",
        "position": "Staff",
        "department": "SUSU",
        "branch": "HEAD OFFICE",
        "imageFile": None,
        "isActive": True,
        "isVerified": True,
        "lastSeen": 1769689048721,
        "registrationTime": 0,
        "isArchived": False,
    },
]
app = Flask(__name__, static_folder=None)
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(UPLOADS_DIR, exist_ok=True)

_MONITOR_ALERT_LOCK = threading.Lock()
_MONITOR_ALERT_TIMES: dict[str, float] = {}


def monitoring_destination_configured() -> bool:
    return bool(env_secret("MONITORING_ALERT_WEBHOOK_URL") or env_secret("MONITORING_ALERT_EMAIL"))


def _deliver_monitoring_alert(payload: dict) -> None:
    webhook_url = env_secret("MONITORING_ALERT_WEBHOOK_URL")
    alert_email = env_secret("MONITORING_ALERT_EMAIL")
    if webhook_url:
        try:
            body = json.dumps(payload, ensure_ascii=True).encode("utf-8")
            alert_request = Request(
                webhook_url,
                data=body,
                headers={"Content-Type": "application/json", "User-Agent": "BCB-SUSU-Monitor/1.0"},
                method="POST",
            )
            with urlopen(alert_request, timeout=5) as response:
                if int(getattr(response, "status", 200) or 200) >= 400:
                    raise RuntimeError(f"Monitoring webhook returned HTTP {response.status}")
        except Exception as exc:
            app.logger.error("Monitoring webhook delivery failed: %s", exc)
    if alert_email:
        try:
            summary = f"{payload.get('severity', 'error').upper()}: {payload.get('message', 'Portal alert')}"
            details = json.dumps(payload, indent=2, ensure_ascii=True)
            send_mail(alert_email, f"BCB SUSU alert: {payload.get('event', 'system')}", f"{summary}\n\n{details}", f"<h2>{summary}</h2><pre>{details}</pre>")
        except Exception as exc:
            app.logger.error("Monitoring email delivery failed: %s", exc)


def emit_production_alert(
    event: str,
    message: str,
    *,
    severity: str = "error",
    context: dict | None = None,
    throttle_key: str | None = None,
    throttle_seconds: int = 300,
) -> None:
    request_id = ""
    route = ""
    method = ""
    if has_request_context():
        request_id = str(getattr(g, "request_id", "") or "")
        route = request.path
        method = request.method
    safe_context = {
        str(key): value
        for key, value in (context or {}).items()
        if key not in {"password", "passwordHash", "token", "code", "accountNumber", "email", "phone"}
    }
    payload = {
        "event": str(event),
        "severity": str(severity),
        "message": str(message),
        "environment": env_secret("MONITORING_ENVIRONMENT") or ("render" if os.getenv("RENDER") else "local"),
        "service": "susu-collection-portal",
        "timestamp": int(time.time()),
        "requestId": request_id,
        "route": route,
        "method": method,
        "context": safe_context,
    }
    app.logger.error("MONITOR_EVENT %s", json.dumps(payload, ensure_ascii=True, sort_keys=True))
    if not monitoring_destination_configured():
        return
    key = throttle_key or f"{event}:{route}:{safe_context.get('status', '')}"
    now = time.time()
    with _MONITOR_ALERT_LOCK:
        last_sent = _MONITOR_ALERT_TIMES.get(key, 0)
        if now - last_sent < max(0, throttle_seconds):
            return
        _MONITOR_ALERT_TIMES[key] = now
    threading.Thread(target=_deliver_monitoring_alert, args=(payload,), daemon=True).start()


def pg_enabled() -> bool:
    return bool(DATABASE_URL)


def pg_connect():
    if not DATABASE_URL:
        return None
    try:
        import psycopg
    except ImportError as exc:
        raise RuntimeError("DATABASE_URL is configured but psycopg is not installed") from exc
    return psycopg.connect(DATABASE_URL, connect_timeout=5)


_PG_TABLE_READY = False
_PG_TABLE_LOCK = threading.Lock()


def ensure_pg_store_table() -> None:
    global _PG_TABLE_READY
    if not pg_enabled() or _PG_TABLE_READY:
        return
    with _PG_TABLE_LOCK:
        if _PG_TABLE_READY:
            return
        with pg_connect() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS portal_store (
                      store_key TEXT PRIMARY KEY,
                      payload JSONB NOT NULL,
                      updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                )
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS portal_schema_migrations (
                      migration_key TEXT PRIMARY KEY,
                      applied_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                )
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS susu_customers (
                      record_id TEXT PRIMARY KEY,
                      account_number TEXT NOT NULL UNIQUE CHECK (account_number ~ '^[0-9]{13}$'),
                      branch_name TEXT NOT NULL,
                      customer_status TEXT NOT NULL,
                      payload JSONB NOT NULL,
                      updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                )
                cur.execute("CREATE INDEX IF NOT EXISTS idx_susu_customers_branch ON susu_customers (branch_name)")
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS susu_collections (
                      record_id TEXT PRIMARY KEY,
                      customer_id TEXT NOT NULL,
                      transaction_reference TEXT NOT NULL UNIQUE,
                      account_number TEXT NOT NULL CHECK (account_number ~ '^[0-9]{13}$'),
                      amount NUMERIC(14, 2) NOT NULL CHECK (amount > 0),
                      agent_id TEXT NOT NULL,
                      branch_name TEXT NOT NULL,
                      transaction_date DATE NOT NULL,
                      collection_status TEXT NOT NULL,
                      idempotency_key TEXT,
                      payload JSONB NOT NULL,
                      created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                      updated_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                )
                cur.execute("CREATE INDEX IF NOT EXISTS idx_susu_collections_date ON susu_collections (transaction_date)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_susu_collections_agent ON susu_collections (agent_id)")
                cur.execute("CREATE INDEX IF NOT EXISTS idx_susu_collections_branch ON susu_collections (branch_name)")
                cur.execute(
                    """
                    CREATE UNIQUE INDEX IF NOT EXISTS idx_susu_collections_idempotency
                    ON susu_collections (agent_id, idempotency_key)
                    WHERE idempotency_key IS NOT NULL AND idempotency_key <> ''
                    """
                )
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS susu_deposit_guards (
                      customer_id TEXT NOT NULL,
                      transaction_date DATE NOT NULL,
                      agent_id TEXT NOT NULL,
                      idempotency_key TEXT,
                      collection_id TEXT NOT NULL UNIQUE,
                      created_at TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                      PRIMARY KEY (customer_id, transaction_date)
                    )
                    """
                )
                cur.execute(
                    """
                    CREATE UNIQUE INDEX IF NOT EXISTS idx_susu_deposit_guard_idempotency
                    ON susu_deposit_guards (agent_id, idempotency_key)
                    WHERE idempotency_key IS NOT NULL AND idempotency_key <> ''
                    """
                )
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS portal_uploads (
                      filename TEXT PRIMARY KEY,
                      content_type TEXT NOT NULL,
                      payload BYTEA NOT NULL,
                      created_at TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                    """
                )
                cur.execute(
                    "SELECT 1 FROM portal_schema_migrations WHERE migration_key = %s",
                    ("normalized-financial-records-v1",),
                )
                if not cur.fetchone():
                    cur.execute(
                        """
                        INSERT INTO susu_customers (
                          record_id, account_number, branch_name, customer_status, payload
                        )
                        SELECT
                          item->>'id', item->>'account_number',
                          UPPER(COALESCE(NULLIF(item->>'branch_name', ''), NULLIF(item->>'branch', ''), 'UNKNOWN')),
                          LOWER(COALESCE(NULLIF(item->>'customer_status', ''), 'active')),
                          item
                        FROM portal_store,
                             LATERAL jsonb_array_elements(payload) AS item
                        WHERE store_key = %s
                          AND COALESCE(item->>'id', '') <> ''
                          AND COALESCE(item->>'account_number', '') ~ '^[0-9]{13}$'
                        ON CONFLICT (record_id) DO NOTHING
                        """,
                        (os.path.basename(CUSTOMERS_STORE_PATH),),
                    )
                    cur.execute(
                        """
                        INSERT INTO susu_collections (
                          record_id, customer_id, transaction_reference, account_number,
                          amount, agent_id, branch_name, transaction_date,
                          collection_status, idempotency_key, payload
                        )
                        SELECT
                          item->>'id', item->>'customer_id', item->>'transaction_reference',
                          item->>'account_number', (item->>'amount')::numeric,
                          item->>'agent_id',
                          UPPER(COALESCE(NULLIF(item->>'branch_name', ''), NULLIF(item->>'branch_id', ''), 'UNKNOWN')),
                          (item->>'transaction_date')::date,
                          LOWER(COALESCE(NULLIF(item->>'status', ''), 'completed')),
                          NULLIF(item->>'idempotency_key', ''), item
                        FROM portal_store,
                             LATERAL jsonb_array_elements(payload) AS item
                        WHERE store_key = %s
                          AND COALESCE(item->>'id', '') <> ''
                          AND COALESCE(item->>'customer_id', '') <> ''
                          AND COALESCE(item->>'transaction_reference', '') <> ''
                          AND COALESCE(item->>'account_number', '') ~ '^[0-9]{13}$'
                          AND COALESCE(item->>'amount', '') ~ '^[0-9]+(\\.[0-9]+)?$'
                          AND (item->>'amount')::numeric > 0
                          AND COALESCE(item->>'agent_id', '') <> ''
                          AND COALESCE(item->>'transaction_date', '') ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}$'
                        ON CONFLICT (record_id) DO NOTHING
                        """,
                        (os.path.basename(COLLECTIONS_STORE_PATH),),
                    )
                    cur.execute(
                        """
                        INSERT INTO susu_deposit_guards (
                          customer_id, transaction_date, agent_id, idempotency_key, collection_id
                        )
                        SELECT DISTINCT ON (customer_id, transaction_date)
                          customer_id, transaction_date, agent_id, idempotency_key, record_id
                        FROM susu_collections
                        WHERE collection_status = 'completed'
                        ORDER BY customer_id, transaction_date, created_at
                        ON CONFLICT DO NOTHING
                        """
                    )
                    cur.execute(
                        "INSERT INTO portal_schema_migrations (migration_key) VALUES (%s)",
                        ("normalized-financial-records-v1",),
                    )
            conn.commit()
        _PG_TABLE_READY = True


def store_key_for_path(path: str) -> str:
    return os.path.basename(path)


class PortalDataLock:
    def __init__(self):
        self._local_lock = threading.RLock()
        self._state = threading.local()

    def __enter__(self):
        depth = int(getattr(self._state, "depth", 0) or 0)
        if depth == 0:
            self._local_lock.acquire()
            self._state.pg_conn = None
            if pg_enabled():
                ensure_pg_store_table()
                conn = pg_connect()
                conn.autocommit = True
                with conn.cursor() as cur:
                    cur.execute("SELECT pg_advisory_lock(%s)", (PG_LOCK_KEY,))
                self._state.pg_conn = conn
        self._state.depth = depth + 1
        return self

    def __exit__(self, exc_type, exc, tb):
        depth = int(getattr(self._state, "depth", 1) or 1) - 1
        self._state.depth = max(0, depth)
        if depth == 0:
            conn = getattr(self._state, "pg_conn", None)
            if conn is not None:
                try:
                    with conn.cursor() as cur:
                        cur.execute("SELECT pg_advisory_unlock(%s)", (PG_LOCK_KEY,))
                finally:
                    conn.close()
                    self._state.pg_conn = None
            self._local_lock.release()


DATA_LOCK = PortalDataLock()


STORE_DEFAULTS: dict[str, object] = {
    PRESENCE_STORE_PATH: {},
    PASSWORD_STORE_PATH: {},
    USERS_STORE_PATH: [],
    PENDING_VERIFICATIONS_PATH: {},
    RESET_TOKENS_PATH: {},
    AGENT_SETUP_TOKENS_PATH: {},
    SESSIONS_STORE_PATH: {},
    PORTAL_AUTHORIZATIONS_PATH: {},
    PRIVILEGED_MFA_PATH: {},
    TRUSTED_DEVICES_PATH: {},
    AUTH_RATE_LIMITS_PATH: {},
    NOTIFICATIONS_STORE_PATH: [],
    AUDIT_LOGS_STORE_PATH: [],
    CUSTOMERS_STORE_PATH: [],
    COLLECTIONS_STORE_PATH: [],
}


def initialize_data_directory() -> None:
    for path, default in STORE_DEFAULTS.items():
        if os.path.exists(path):
            continue
        legacy_path = os.path.join(BASE_DIR, os.path.basename(path))
        if path != legacy_path and os.path.exists(legacy_path):
            try:
                os.replace(legacy_path, path)
                continue
            except OSError:
                pass
        with open(path, "w", encoding="utf-8") as handle:
            json.dump(default, handle, ensure_ascii=True, indent=2)


initialize_data_directory()


def seed_password_store_if_needed() -> None:
    if not DEFAULT_INITIAL_PASSWORD:
        return
    existing = read_json_file(PASSWORD_STORE_PATH, {})
    passwords = existing if isinstance(existing, dict) else {}
    changed = False
    for user in [OWNER_ADMIN_USER, *INITIAL_USERS]:
        email = str(user.get("email", "")).strip().lower()
        if not email or passwords.get(email):
            continue
        passwords[email] = hash_password_for_storage(DEFAULT_INITIAL_PASSWORD)
        changed = True
    if changed:
        save_password_store(passwords)


def allowed_origins() -> set[str]:
    raw = os.getenv("ALLOWED_ORIGINS", "")
    return {item.strip() for item in raw.split(",") if item.strip()}


@app.before_request
def begin_request_tracking():
    g.request_id = request.headers.get("X-Request-ID") or secrets.token_hex(8)
    g.request_started_at = time.perf_counter()


@app.errorhandler(Exception)
def handle_unexpected_error(exc):
    if isinstance(exc, HTTPException):
        return exc
    app.logger.exception("Unhandled request failure [%s]", getattr(g, "request_id", "unknown"))
    g.unhandled_alerted = True
    if request.path == "/api/collections" and request.method == "POST":
        emit_production_alert(
            "DEPOSIT_PROCESSING_FAILED",
            "A deposit could not be processed because of an internal system error.",
            severity="critical",
            context={"exceptionType": type(exc).__name__},
            throttle_key="deposit-processing-failure",
            throttle_seconds=60,
        )
    emit_production_alert(
        "API_UNHANDLED_EXCEPTION",
        "An API request failed unexpectedly.",
        severity="critical",
        context={"exceptionType": type(exc).__name__},
        throttle_key=f"api-exception:{request.path}",
        throttle_seconds=60,
    )
    return jsonify({"error": "The server could not complete this request.", "requestId": getattr(g, "request_id", "")}), 500


@app.after_request
def add_cors_headers(response):
    origin = request.headers.get("Origin")
    origins = allowed_origins()
    if "*" in origins:
        response.headers["Access-Control-Allow-Origin"] = origin or "*"
        response.headers["Vary"] = "Origin"
    elif origin and origin in origins:
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Vary"] = "Origin"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
    if origin and ("*" in origins or origin in origins):
        response.headers["Access-Control-Allow-Credentials"] = "true"
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; script-src 'self'; style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data: blob: https:; font-src 'self' data:; connect-src 'self'; "
        "object-src 'none'; base-uri 'self'; frame-ancestors 'none'; form-action 'self'; "
        "worker-src 'self' blob:; manifest-src 'self'"
    )
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=(), payment=()"
    response.headers["Cross-Origin-Opener-Policy"] = "same-origin"
    if session_cookie_secure():
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    response.headers["X-Request-ID"] = str(getattr(g, "request_id", "") or "")
    elapsed = time.perf_counter() - float(getattr(g, "request_started_at", time.perf_counter()))
    if request.path.startswith("/api/") and response.status_code >= 500 and not getattr(g, "unhandled_alerted", False):
        emit_production_alert(
            "API_FAILURE",
            "An API endpoint returned a server error.",
            context={"status": response.status_code, "durationMs": round(elapsed * 1000)},
            throttle_key=f"api-failure:{request.path}:{response.status_code}",
            throttle_seconds=60,
        )
    return response


def require_json():
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return None, (jsonify({"error": "JSON body required"}), 400)
    return data, None


def extract_drive_file_id(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    if raw.startswith("DRIVE:"):
        return raw[6:].strip()
    if "drive.google.com" not in raw:
        return raw
    if "/d/" in raw:
        return raw.split("/d/")[1].split("/")[0].split("?")[0].strip()
    if "id=" in raw:
        return raw.split("id=")[1].split("&")[0].strip()
    return raw


def normalize_visibility_and_department(data: dict) -> tuple[str, str | None]:
    visibility = str(data.get("visibility", "General")).strip()
    if visibility not in {"General", "Department"}:
        raise ValueError("Visibility must be General or Department")
    department = str(data.get("department", "") or "").strip().upper() or None
    if visibility == "Department" and not department:
        raise ValueError("Department is required for department visibility")
    if visibility == "General":
        department = None
    return visibility, department


def normalize_scope_list(value: object, *, empty_default: list[str] | None = None) -> list[str]:
    if not isinstance(value, list):
        return list(empty_default or [])
    normalized: list[str] = []
    seen = set()
    for item in value:
        current = str(item or "").strip().upper()
        if not current:
            continue
        if current == "ALL":
            return ["ALL"]
        if current in seen:
            continue
        seen.add(current)
        normalized.append(current)
    return normalized or list(empty_default or [])


def default_permissions_for_role(role: str) -> dict[str, bool]:
    is_manager = role in GLOBAL_MANAGER_ROLES
    susu_permissions = {
        "customers": is_manager,
        "transactions": is_manager,
        "reports": is_manager,
        "agents": is_manager,
        "branches": is_manager,
        "auditLog": is_manager,
        "backupExport": is_manager,
    }
    if is_manager:
        return {
            "userManagement": True,
            **susu_permissions,
        }
    return {
        "userManagement": False,
        **susu_permissions,
    }


def normalize_user_permissions(value: object, role: str) -> dict[str, bool]:
    defaults = default_permissions_for_role(role)
    if role not in GLOBAL_MANAGER_ROLES and role != "Supervisor":
        return defaults
    if not isinstance(value, dict):
        return defaults
    normalized = dict(defaults)
    for key in defaults:
        if key in value:
            normalized[key] = bool(value.get(key))
    return normalized


def normalize_managed_departments_by_branch(value: object) -> dict[str, list[str]]:
    if not isinstance(value, dict):
        return {}
    normalized: dict[str, list[str]] = {}
    for branch, departments in value.items():
        branch_key = str(branch or "").strip().upper()
        if not branch_key:
            continue
        normalized_departments = normalize_scope_list(departments, empty_default=[])
        if normalized_departments:
            normalized[branch_key] = normalized_departments
    return normalized


def validate_supervisor_configuration(user: dict) -> None:
    if str(user.get("role", "")).strip() != "Supervisor":
        return
    managed_branches = normalize_scope_list(user.get("managedBranches"), empty_default=[])
    if not managed_branches:
        raise ValueError("Supervisors must be assigned at least one branch.")
    for branch in managed_branches:
        if branch == "ALL":
            raise ValueError("Supervisors cannot be assigned to all branches.")


def derive_content_scope(
    data: dict,
    *,
    visibility: str,
    department: str | None,
    existing: dict | None = None,
) -> tuple[list[str], list[str]]:
    branch_scope = normalize_scope_list(
        data.get("branchScope", existing.get("branchScope") if existing else None),
        empty_default=["ALL"],
    )
    department_scope = normalize_scope_list(
        data.get("departmentScope", existing.get("departmentScope") if existing else None),
        empty_default=[],
    )
    if not department_scope:
        department_scope = [department] if visibility == "Department" and department else ["ALL"]
    return branch_scope, department_scope


def normalize_non_empty_title(value: object, label: str) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        raise ValueError(f"{label} is required")
    return normalized


def normalize_storage_type(value: object) -> str:
    normalized = str(value or "Drive").strip()
    if normalized not in {"Drive", "Local"}:
        raise ValueError("Storage type must be Drive or Local")
    return normalized


def normalize_local_filename(value: object) -> str:
    filename = secure_filename(str(value or "").strip())
    if not filename:
        raise ValueError("A valid uploaded file is required")
    if not os.path.isfile(os.path.join(UPLOADS_DIR, filename)):
        raise ValueError("Uploaded file could not be found")
    return filename


def parse_session_token() -> str:
    header = str(request.headers.get("Authorization", "")).strip()
    if header.startswith("Bearer "):
        return header[7:].strip()
    return str(request.cookies.get(SESSION_COOKIE_NAME, "") or "").strip()


def validate_email(email: str) -> str:
    normalized = (email or "").strip().lower()
    settings = load_portal_settings_store()
    if not normalized.endswith(settings["emailDomain"]):
        raise ValueError("Only official Bawjiase email addresses are allowed")
    return normalized


VALID_CUSTOMER_STATUSES = {"active", "inactive", "suspended"}


def normalize_required_text(value: object, field_label: str) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        raise ValueError(f"{field_label} is required")
    return normalized


def normalize_phone(value: object) -> str:
    phone = str(value or "").strip()
    if not phone:
        return ""
    allowed = set("0123456789+ -()")
    if any(char not in allowed for char in phone):
        raise ValueError("Phone number can only contain numbers, spaces, +, -, and brackets")
    digits = "".join(char for char in phone if char.isdigit())
    if len(digits) < 7:
        raise ValueError("Phone number is too short")
    return phone


def normalize_account_number(value: object) -> str:
    account_number = str(value or "").strip().replace(" ", "")
    if not account_number.isdigit():
        raise ValueError("Account number must contain only digits")
    if len(account_number) != 13:
        raise ValueError("Account number must be exactly 13 digits")
    return account_number


def normalize_customer_status(value: object) -> str:
    status = str(value or "active").strip().lower()
    if status not in VALID_CUSTOMER_STATUSES:
        raise ValueError("Customer status must be active, inactive, or suspended")
    return status


def normalize_portal_branch_name(value: object) -> str:
    branch = str(value or "").strip().upper()
    if not branch:
        raise ValueError("Branch is required")
    settings = load_portal_settings_store()
    valid = {str(item).strip().upper() for item in settings.get("branches", [])}
    if valid and branch not in valid:
        raise ValueError("Branch must be selected from Portal Control")
    return branch


def normalize_portal_department_name(value: object) -> str:
    department = str(value or "").strip().upper()
    if not department:
        raise ValueError("Department is required")
    settings = load_portal_settings_store()
    valid = {str(item).strip().upper() for item in settings.get("departments", [])}
    if valid and department not in valid:
        raise ValueError("Department must be selected from Portal Control")
    return department


def role_from_department(department: str) -> str:
    return "GeneralStaff"


def is_global_manager(user: dict | None) -> bool:
    return bool(user) and str(user.get("role", "")).strip() in GLOBAL_MANAGER_ROLES


def is_owner_admin(user: dict | None) -> bool:
    return bool(user) and str(user.get("role", "")).strip() == OWNER_ADMIN_ROLE


def user_has_permission(user: dict, permission_key: str) -> bool:
    if is_global_manager(user):
        return True
    permissions = user.get("permissions")
    if not isinstance(permissions, dict):
        return False
    return bool(permissions.get(permission_key, False))


def now_ms() -> int:
    return int(time.time() * 1000)


def now_seconds() -> int:
    return int(time.time())


def pagination_requested() -> bool:
    return "page" in request.args or "pageSize" in request.args


def paginate_items(items: list, *, default_page_size: int = 25, max_page_size: int = 100) -> tuple[list, dict]:
    try:
        page = max(1, int(request.args.get("page", 1) or 1))
        page_size = int(request.args.get("pageSize", default_page_size) or default_page_size)
    except (TypeError, ValueError):
        page, page_size = 1, default_page_size
    page_size = max(5, min(max_page_size, page_size))
    total = len(items)
    total_pages = max(1, math.ceil(total / page_size))
    page = min(page, total_pages)
    start = (page - 1) * page_size
    return items[start:start + page_size], {
        "page": page,
        "pageSize": page_size,
        "total": total,
        "totalPages": total_pages,
        "hasPrevious": page > 1,
        "hasNext": page < total_pages,
    }


def paginated_response(key: str, items: list):
    page_items, pagination = paginate_items(items)
    return jsonify({key: page_items, "pagination": pagination})


def legacy_hash_password(password: str) -> str:
    h = 0
    for char in password:
        h = ((31 * h) + ord(char)) & 0xFFFFFFFF
        if h & 0x80000000:
            h -= 0x100000000
    return str(abs(h))


def is_secure_password_hash(value: str) -> bool:
    return value.startswith("pbkdf2:") or value.startswith("scrypt:")


def hash_password_for_storage(password: str) -> str:
    return generate_password_hash(password)


def verify_password(stored_value: str, password: str) -> bool:
    if is_secure_password_hash(stored_value):
        try:
            return check_password_hash(stored_value, password)
        except ValueError:
            return False
    return stored_value == legacy_hash_password(password)


def atomic_write_json(path: str, payload) -> None:
    with DATA_LOCK:
        if pg_enabled():
            ensure_pg_store_table()
            key = store_key_for_path(path)
            with pg_connect() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        """
                        INSERT INTO portal_store (store_key, payload, updated_at)
                        VALUES (%s, %s::jsonb, NOW())
                        ON CONFLICT (store_key)
                        DO UPDATE SET payload = EXCLUDED.payload, updated_at = NOW()
                        """,
                        (key, json.dumps(payload, ensure_ascii=True)),
                    )
                conn.commit()
            return
        directory = os.path.dirname(path)
        fd, tmp_path = tempfile.mkstemp(prefix="tmp-", suffix=".json", dir=directory)
        try:
            with os.fdopen(fd, "w", encoding="utf-8") as handle:
                json.dump(payload, handle, ensure_ascii=True, indent=2)
            last_error = None
            for attempt in range(8):
                try:
                    os.replace(tmp_path, path)
                    last_error = None
                    break
                except PermissionError as exc:
                    last_error = exc
                    time.sleep(0.05 * (attempt + 1))
            if last_error:
                raise last_error
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)


def read_json_file(path: str, default):
    if pg_enabled():
        ensure_pg_store_table()
        key = store_key_for_path(path)
        with pg_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT payload FROM portal_store WHERE store_key = %s", (key,))
                row = cur.fetchone()
                if row:
                    return row[0]
                initial_payload = default
                if os.path.exists(path):
                    try:
                        with open(path, "r", encoding="utf-8") as handle:
                            initial_payload = json.load(handle)
                    except Exception:
                        initial_payload = default
                cur.execute(
                    "INSERT INTO portal_store (store_key, payload) VALUES (%s, %s::jsonb) ON CONFLICT DO NOTHING",
                    (key, json.dumps(initial_payload, ensure_ascii=True)),
                )
            conn.commit()
        return initial_payload
    if not os.path.exists(path):
        return default
    try:
        with open(path, "r", encoding="utf-8") as handle:
            return json.load(handle)
    except Exception:
        return default


def normalize_user(raw: dict) -> dict:
    raw_email = str(raw.get("email", "")).strip().lower()
    login_username = str(raw.get("loginUsername", "") or raw.get("username", "")).strip().lower()
    if raw_email.endswith("@agents.local") or (not raw_email and login_username):
        email = raw_email or f"{login_username}@agents.local"
    else:
        email = validate_email(raw_email)
    department = str(raw.get("department", "")).strip().upper()
    if not department or department in {"SUSU AGENT", "SUSU SUPERVISOR"}:
        department = "SUSU"
    branch = str(raw.get("branch", "")).strip().upper()
    role = str(raw.get("role", role_from_department(department))).strip() or role_from_department(department)
    if email == "sitecreator@bawjiasecommunitybank.com" or str(raw.get("id", "")).strip() == "owner-admin-1":
        role = OWNER_ADMIN_ROLE
    return {
        "id": str(raw.get("id", "")).strip(),
        "fullname": str(raw.get("fullname", "")).strip(),
        "phone": str(raw.get("phone", "")).strip(),
        "email": email,
        "role": role,
        "position": str(raw.get("position", "")).strip() or "Staff",
        "department": department,
        "branch": branch,
        "imageFile": raw.get("imageFile"),
        "managedBranches": normalize_scope_list(
            raw.get("managedBranches"),
            empty_default=(
                ["ALL"]
                if role in GLOBAL_MANAGER_ROLES
                else ([branch] if role == "Supervisor" and branch else [])
            ),
        ),
        "managedDepartmentsByBranch": normalize_managed_departments_by_branch(
            raw.get("managedDepartmentsByBranch")
        ),
        "permissions": normalize_user_permissions(raw.get("permissions"), role),
        "isActive": bool(raw.get("isActive", True)),
        "isVerified": bool(raw.get("isVerified", True)),
        "lastSeen": normalize_last_seen_ms(raw.get("lastSeen", 0)),
        "registrationTime": int(raw.get("registrationTime", 0) or 0),
        "isArchived": bool(raw.get("isArchived", False)),
        "loginUsername": login_username,
        "createdBySupervisorId": str(raw.get("createdBySupervisorId", "") or "").strip(),
        "createdBySupervisorName": str(raw.get("createdBySupervisorName", "") or "").strip(),
        "forcePasswordChange": bool(raw.get("forcePasswordChange", False)),
        "setupComplete": bool(raw.get("setupComplete", True)),
        "setupReason": str(raw.get("setupReason", "") or "").strip().lower(),
        "isTestData": bool(raw.get("isTestData", False)),
    }


def load_user_store() -> list[dict]:
    raw = read_json_file(USERS_STORE_PATH, [])
    users_by_email = {}
    default_source = [OWNER_ADMIN_USER] if isinstance(raw, list) and raw else [OWNER_ADMIN_USER, *INITIAL_USERS]
    for default_user in default_source:
        normalized = normalize_user(default_user)
        users_by_email[normalized["email"]] = normalized
    if isinstance(raw, list):
        for item in raw:
            if isinstance(item, dict):
                try:
                    normalized = normalize_user(item)
                    users_by_email[normalized["email"]] = normalized
                except ValueError:
                    continue
    return list(users_by_email.values())


def save_user_store(users: list[dict]) -> None:
    normalized = []
    for user in users:
        try:
            normalized.append(normalize_user(user))
        except ValueError:
            continue
    atomic_write_json(USERS_STORE_PATH, normalized)


def find_user_by_email(users: list[dict], email: str):
    return next((user for user in users if user["email"] == email), None)


def find_user_by_id(users: list[dict], user_id: str):
    return next((user for user in users if user["id"] == user_id), None)


def load_presence_store() -> dict[str, int]:
    raw = read_json_file(PRESENCE_STORE_PATH, {})
    if not isinstance(raw, dict):
        return {}
    return {
        str(user_id): int(timestamp)
        for user_id, timestamp in raw.items()
        if str(user_id) and isinstance(timestamp, (int, float, str))
    }


def normalize_last_seen_ms(value: object) -> int:
    try:
        last_seen = int(value or 0)
    except (TypeError, ValueError):
        return 0
    if last_seen <= 0:
        return 0
    current = now_ms()
    if last_seen > current + 60_000:
        return 0
    return last_seen


def user_has_active_session(user_id: str) -> bool:
    normalized_user_id = str(user_id or "").strip()
    if not normalized_user_id:
        return False
    sessions = load_sessions()
    return any(str(session.get("userId", "")).strip() == normalized_user_id for session in sessions.values())


def presence_is_online(presence_timestamp: object, user_id: str | None = None) -> bool:
    value = normalize_presence_timestamp(int(presence_timestamp or 0))
    if value <= 0:
        return False
    return value >= now_seconds() - ONLINE_WINDOW_SECONDS


def set_user_last_seen(user_id: str, last_seen_ms: int | None) -> None:
    users = load_user_store()
    user = find_user_by_id(users, user_id)
    if not user:
        return
    user["lastSeen"] = normalize_last_seen_ms(last_seen_ms or 0)
    save_user_store(users)


def normalize_presence_timestamp(timestamp: int) -> int:
    value = int(timestamp or 0)
    if value <= 0:
        return 0
    # Older builds may have written milliseconds instead of seconds.
    if value > 10_000_000_000:
        value = value // 1000
    now = int(time.time())
    # Discard obviously broken future timestamps.
    if value > now + 60:
        return 0
    return value


def save_presence_store(store: dict[str, int]) -> None:
    atomic_write_json(PRESENCE_STORE_PATH, store)


def prune_presence(store: dict[str, int]) -> dict[str, int]:
    cutoff = int(time.time()) - PRESENCE_TTL_SECONDS
    return {
        str(user_id): normalize_presence_timestamp(timestamp)
        for user_id, timestamp in store.items()
        if str(user_id).strip()
        and normalize_presence_timestamp(timestamp) >= cutoff
    }


def serialize_user_with_presence(user: dict, presence: dict[str, int] | None = None) -> dict:
    presence_map = presence if presence is not None else prune_presence(load_presence_store())
    serialized = dict(user)
    user_id = str(serialized.get("id", "")).strip()
    last_seen = normalize_last_seen_ms(serialized.get("lastSeen", 0))
    serialized["lastSeen"] = last_seen
    serialized["isOnlineNow"] = presence_is_online(presence_map.get(user_id, 0), user_id)
    return serialized


def serialize_users_with_presence(users: list[dict]) -> list[dict]:
    presence = prune_presence(load_presence_store())
    save_presence_store(presence)
    return [serialize_user_with_presence(user, presence) for user in users]


def load_password_store() -> dict[str, str]:
    raw = read_json_file(PASSWORD_STORE_PATH, {})
    if not isinstance(raw, dict):
        return {}
    return {
        email.strip().lower(): password_hash
        for email, password_hash in raw.items()
        if isinstance(email, str) and isinstance(password_hash, str) and password_hash
    }


def save_password_store(store: dict[str, str]) -> None:
    normalized = {
        str(email).strip().lower(): str(password_hash).strip()
        for email, password_hash in store.items()
        if str(email).strip() and str(password_hash).strip()
    }
    atomic_write_json(PASSWORD_STORE_PATH, normalized)


seed_password_store_if_needed()


def load_pending_verifications() -> dict[str, dict]:
    raw = read_json_file(PENDING_VERIFICATIONS_PATH, {})
    if not isinstance(raw, dict):
        return {}
    pending = {}
    current = int(time.time())
    for email, item in raw.items():
        if not isinstance(item, dict):
            continue
        try:
            normalized_email = validate_email(email)
        except ValueError:
            continue
        expires_at = int(item.get("expiresAt", 0) or 0)
        if expires_at <= current:
            continue
        user = item.get("user")
        password_hash = str(item.get("passwordHash", "")).strip()
        code = "".join(ch for ch in str(item.get("code", "")) if ch.isdigit())
        if not isinstance(user, dict) or len(code) != 6 or not password_hash:
            continue
        try:
            pending[normalized_email] = {
                "user": normalize_user(user),
                "passwordHash": password_hash,
                "code": code,
                "expiresAt": expires_at,
            }
        except ValueError:
            continue
    return pending


def save_pending_verifications(store: dict[str, dict]) -> None:
    atomic_write_json(PENDING_VERIFICATIONS_PATH, store)


def load_reset_tokens() -> dict[str, dict]:
    raw = read_json_file(RESET_TOKENS_PATH, {})
    if not isinstance(raw, dict):
        return {}
    current = int(time.time())
    tokens = {}
    for token, item in raw.items():
        if not isinstance(token, str) or not isinstance(item, dict):
            continue
        expires_at = int(item.get("expiresAt", 0) or 0)
        if expires_at <= current:
            continue
        try:
            email = validate_email(str(item.get("email", "")))
        except ValueError:
            continue
        tokens[token] = {
            "email": email,
            "expiresAt": expires_at,
        }
    return tokens


def save_reset_tokens(store: dict[str, dict]) -> None:
    atomic_write_json(RESET_TOKENS_PATH, store)


def load_agent_setup_tokens() -> dict[str, dict]:
    raw = read_json_file(AGENT_SETUP_TOKENS_PATH, {})
    if not isinstance(raw, dict):
        return {}
    current = int(time.time())
    tokens = {}
    for username, item in raw.items():
        if not isinstance(username, str) or not isinstance(item, dict):
            continue
        try:
            normalized_username = normalize_agent_username(username)
        except ValueError:
            continue
        expires_at = int(item.get("expiresAt", 0) or 0)
        code = "".join(ch for ch in str(item.get("code", "")) if ch.isdigit())
        if expires_at <= current or len(code) != 6:
            continue
        tokens[normalized_username] = {
            "code": code,
            "expiresAt": expires_at,
            "phone": normalize_phone(item.get("phone")),
        }
    return tokens


def save_agent_setup_tokens(store: dict[str, dict]) -> None:
    atomic_write_json(AGENT_SETUP_TOKENS_PATH, store)


def issue_agent_setup_token(username: str, phone: str) -> dict:
    tokens = load_agent_setup_tokens()
    code = generate_verification_code()
    normalized_username = normalize_agent_username(username)
    expires_at = int(time.time()) + int(load_portal_settings_store()["verificationMinutes"]) * 60
    tokens[normalized_username] = {
        "code": code,
        "phone": normalize_phone(phone),
        "expiresAt": expires_at,
    }
    save_agent_setup_tokens(tokens)
    return tokens[normalized_username]


def normalized_financial_table(path: str) -> str | None:
    absolute_path = os.path.abspath(path)
    if absolute_path == os.path.abspath(CUSTOMERS_STORE_PATH):
        return "susu_customers"
    if absolute_path == os.path.abspath(COLLECTIONS_STORE_PATH):
        return "susu_collections"
    return None


def load_normalized_financial_store(table_name: str) -> list[dict]:
    ensure_pg_store_table()
    order_column = "updated_at" if table_name == "susu_customers" else "created_at"
    with pg_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(f"SELECT payload FROM {table_name} ORDER BY {order_column}, record_id")
            return [row[0] for row in cur.fetchall() if isinstance(row[0], dict)]


def save_normalized_financial_store(table_name: str, items: list[dict]) -> None:
    ensure_pg_store_table()
    normalized_items = [item for item in items if isinstance(item, dict)]
    record_ids = []
    with DATA_LOCK:
        with pg_connect() as conn:
            with conn.cursor() as cur:
                for item in normalized_items:
                    record_id = str(item.get("id") or "").strip()
                    account_number = str(item.get("account_number") or "").strip()
                    if not record_id:
                        raise ValueError("Financial records must have an id")
                    if len(account_number) != 13 or not account_number.isdigit():
                        raise ValueError("Financial records must have a valid 13-digit account number")
                    record_ids.append(record_id)
                    if table_name == "susu_customers":
                        cur.execute(
                            """
                            INSERT INTO susu_customers (
                              record_id, account_number, branch_name, customer_status, payload, updated_at
                            ) VALUES (%s, %s, %s, %s, %s::jsonb, NOW())
                            ON CONFLICT (record_id) DO UPDATE SET
                              account_number = EXCLUDED.account_number,
                              branch_name = EXCLUDED.branch_name,
                              customer_status = EXCLUDED.customer_status,
                              payload = EXCLUDED.payload,
                              updated_at = NOW()
                            """,
                            (
                                record_id,
                                account_number,
                                str(item.get("branch_name") or item.get("branch") or "UNKNOWN").strip().upper(),
                                str(item.get("customer_status") or "active").strip().lower(),
                                json.dumps(item, ensure_ascii=True),
                            ),
                        )
                    else:
                        transaction_date = str(item.get("transaction_date") or "").strip()
                        try:
                            time.strptime(transaction_date, "%Y-%m-%d")
                            amount = round(float(item.get("amount") or 0), 2)
                        except (TypeError, ValueError):
                            raise ValueError("Collection records must have a valid amount and transaction date")
                        if amount <= 0:
                            raise ValueError("Collection amounts must be greater than zero")
                        cur.execute(
                            """
                            INSERT INTO susu_collections (
                              record_id, customer_id, transaction_reference, account_number,
                              amount, agent_id, branch_name, transaction_date,
                              collection_status, idempotency_key, payload, updated_at
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s::date, %s, %s, %s::jsonb, NOW())
                            ON CONFLICT (record_id) DO UPDATE SET
                              customer_id = EXCLUDED.customer_id,
                              transaction_reference = EXCLUDED.transaction_reference,
                              account_number = EXCLUDED.account_number,
                              amount = EXCLUDED.amount,
                              agent_id = EXCLUDED.agent_id,
                              branch_name = EXCLUDED.branch_name,
                              transaction_date = EXCLUDED.transaction_date,
                              collection_status = EXCLUDED.collection_status,
                              idempotency_key = EXCLUDED.idempotency_key,
                              payload = EXCLUDED.payload,
                              updated_at = NOW()
                            """,
                            (
                                record_id,
                                str(item.get("customer_id") or "").strip(),
                                str(item.get("transaction_reference") or "").strip(),
                                account_number,
                                amount,
                                str(item.get("agent_id") or "").strip(),
                                str(item.get("branch_name") or item.get("branch_id") or "UNKNOWN").strip().upper(),
                                transaction_date,
                                str(item.get("status") or "completed").strip().lower(),
                                str(item.get("idempotency_key") or "").strip() or None,
                                json.dumps(item, ensure_ascii=True),
                            ),
                        )
                if record_ids:
                    cur.execute(
                        f"DELETE FROM {table_name} WHERE NOT (record_id = ANY(%s))",
                        (record_ids,),
                    )
                else:
                    cur.execute(f"DELETE FROM {table_name}")
                if table_name == "susu_collections":
                    cur.execute(
                        """
                        DELETE FROM susu_deposit_guards AS guard
                        WHERE NOT EXISTS (
                          SELECT 1 FROM susu_collections AS collection
                          WHERE collection.record_id = guard.collection_id
                            AND collection.collection_status = 'completed'
                        )
                        """
                    )
                    cur.execute(
                        """
                        INSERT INTO susu_deposit_guards (
                          customer_id, transaction_date, agent_id, idempotency_key, collection_id
                        )
                        SELECT DISTINCT ON (customer_id, transaction_date)
                          customer_id, transaction_date, agent_id, idempotency_key, record_id
                        FROM susu_collections
                        WHERE collection_status = 'completed'
                        ORDER BY customer_id, transaction_date, created_at
                        ON CONFLICT DO NOTHING
                        """
                    )
            conn.commit()


def load_json_list_store(path: str) -> list[dict]:
    table_name = normalized_financial_table(path) if pg_enabled() else None
    if table_name:
        return load_normalized_financial_store(table_name)
    raw = read_json_file(path, [])
    return raw if isinstance(raw, list) else []


def save_json_list_store(path: str, items: list[dict]) -> None:
    table_name = normalized_financial_table(path) if pg_enabled() else None
    if table_name:
        save_normalized_financial_store(table_name, items)
        return
    atomic_write_json(path, items)


def normalize_portal_branches(values) -> list[str]:
    seen = set()
    branches = []
    if not isinstance(values, list):
        values = []
    for value in values:
        branch = str(value or "").strip().upper()
        if not branch or branch in seen:
            continue
        seen.add(branch)
        branches.append(branch)
    return branches or list(DEFAULT_PORTAL_BRANCHES)


def normalize_portal_list(values, fallback: list[str], uppercase: bool = False) -> list[str]:
    seen = set()
    items = []
    if not isinstance(values, list):
        values = []
    for value in values:
        item = str(value or "").strip()
        if uppercase:
            item = item.upper()
        key = item.upper()
        if not item or key in seen:
            continue
        seen.add(key)
        items.append(item)
    return items or list(fallback)


def merge_missing_portal_defaults(values: list[str], defaults: list[str], uppercase: bool = False) -> list[str]:
    items = normalize_portal_list(values, defaults, uppercase)
    seen = {str(item).strip().upper() for item in items}
    for value in defaults:
        item = str(value or "").strip()
        if uppercase:
            item = item.upper()
        key = item.upper()
        if item and key not in seen:
            items.append(item)
            seen.add(key)
    return items


def normalize_portal_departments(values) -> list[str]:
    raw_departments = normalize_portal_list(values, DEFAULT_PORTAL_DEPARTMENTS, True)
    departments = []
    seen = set()
    for value in raw_departments:
        item = str(value or "").strip().upper()
        if item in {"SUSU AGENT", "SUSU SUPERVISOR"}:
            item = "SUSU"
        if item and item not in seen:
            departments.append(item)
            seen.add(item)
    for value in REQUIRED_PORTAL_DEPARTMENTS:
        item = str(value or "").strip().upper()
        if item and item not in seen:
            departments.append(item)
            seen.add(item)
    return departments


def normalize_portal_rename_map(value: object) -> dict[str, str]:
    if not isinstance(value, dict):
        return {}
    renames: dict[str, str] = {}
    for old_value, new_value in value.items():
        old_item = str(old_value or "").strip().upper()
        new_item = str(new_value or "").strip().upper()
        if old_item and new_item and old_item != new_item:
            renames[old_item] = new_item
    return renames


def rename_scoped_values(values: object, renames: dict[str, str]) -> list[str]:
    normalized = normalize_scope_list(values, empty_default=[])
    updated = []
    seen = set()
    for value in normalized:
        key = str(value or "").strip().upper()
        item = renames.get(key, key)
        if item and item not in seen:
            updated.append(item)
            seen.add(item)
    return updated


def rename_managed_department_map(value: object, branch_renames: dict[str, str], department_renames: dict[str, str]) -> dict[str, list[str]]:
    current = normalize_managed_departments_by_branch(value)
    updated: dict[str, list[str]] = {}
    for branch, departments in current.items():
        branch_key = branch_renames.get(str(branch or "").strip().upper(), str(branch or "").strip().upper())
        department_values = rename_scoped_values(departments, department_renames)
        if branch_key and department_values:
            updated[branch_key] = department_values
    return updated


def rename_item_fields(item: dict, renames: dict[str, str], field_names: list[str]) -> bool:
    changed = False
    for field in field_names:
        value = str(item.get(field, "") or "").strip().upper()
        if value in renames:
            item[field] = renames[value]
            changed = True
    return changed


def rename_scope_fields(item: dict, branch_renames: dict[str, str], department_renames: dict[str, str]) -> bool:
    changed = False
    if "branchScope" in item:
        next_scope = rename_scoped_values(item.get("branchScope"), branch_renames)
        if next_scope != item.get("branchScope"):
            item["branchScope"] = next_scope
            changed = True
    if "departmentScope" in item:
        next_scope = rename_scoped_values(item.get("departmentScope"), department_renames)
        if next_scope != item.get("departmentScope"):
            item["departmentScope"] = next_scope
            changed = True
    return changed


def apply_portal_renames(branch_renames: dict[str, str], department_renames: dict[str, str]) -> dict[str, int]:
    summary = {"users": 0, "customers": 0, "customerImports": 0, "collections": 0, "dailyCloses": 0, "content": 0}
    if not branch_renames and not department_renames:
        return summary

    users = load_user_store()
    users_changed = False
    for user in users:
        changed = rename_item_fields(user, branch_renames, ["branch"])
        changed = rename_item_fields(user, department_renames, ["department"]) or changed
        next_managed_branches = rename_scoped_values(user.get("managedBranches"), branch_renames)
        if next_managed_branches != user.get("managedBranches"):
            user["managedBranches"] = next_managed_branches
            changed = True
        next_managed_departments = rename_managed_department_map(
            user.get("managedDepartmentsByBranch"),
            branch_renames,
            department_renames,
        )
        if next_managed_departments != normalize_managed_departments_by_branch(user.get("managedDepartmentsByBranch")):
            user["managedDepartmentsByBranch"] = next_managed_departments
            changed = True
        if changed:
            summary["users"] += 1
            users_changed = True
    if users_changed:
        save_user_store(users)

    simple_stores = [
        (CUSTOMERS_STORE_PATH, "customers", ["branch", "branch_name", "branch_id"]),
        (CUSTOMER_IMPORTS_STORE_PATH, "customerImports", ["branch", "branch_name", "branch_id"]),
        (COLLECTIONS_STORE_PATH, "collections", ["branch", "branch_name", "branch_id", "agent_branch", "customer_branch"]),
        (DAILY_CLOSES_STORE_PATH, "dailyCloses", ["branch", "branch_name", "branch_id"]),
    ]
    for path, summary_key, branch_fields in simple_stores:
        items = load_json_list_store(path)
        changed_count = 0
        for item in items:
            if rename_item_fields(item, branch_renames, branch_fields):
                changed_count += 1
        if changed_count:
            save_json_list_store(path, items)
            summary[summary_key] = changed_count

    return summary


def persist_normalized_susu_departments() -> int:
    raw_users = read_json_file(USERS_STORE_PATH, [])
    if not isinstance(raw_users, list):
        return 0
    legacy_count = sum(
        1
        for user in raw_users
        if isinstance(user, dict)
        and str(user.get("department", "")).strip().upper() in {"SUSU AGENT", "SUSU SUPERVISOR"}
    )
    if legacy_count:
        save_user_store(load_user_store())
    return legacy_count


def portal_list_changes(previous_values: object, next_values: object) -> dict[str, list[str]]:
    previous = normalize_portal_list(previous_values, [], True)
    current = normalize_portal_list(next_values, [], True)
    previous_set = set(previous)
    current_set = set(current)
    return {
        "added": [item for item in current if item not in previous_set],
        "removed": [item for item in previous if item not in current_set],
    }
def normalize_email_domain(value) -> str:
    domain = str(value or "").strip().lower()
    if not domain:
        return OFFICIAL_EMAIL_DOMAIN
    domain = domain.replace(" ", "")
    if not domain.startswith("@"):
        domain = f"@{domain}"
    return domain


def normalize_positive_number(value, fallback: int) -> int:
    try:
        number = int(float(value))
    except (TypeError, ValueError):
        return fallback
    return number if number > 0 else fallback


def load_portal_settings_store() -> dict:
    raw = read_json_file(PORTAL_SETTINGS_STORE_PATH, {})
    if not isinstance(raw, dict):
        raw = {}
    return {
        "bankName": str(raw.get("bankName") or DEFAULT_PORTAL_SETTINGS["bankName"]).strip(),
        "shortBankName": str(raw.get("shortBankName") or DEFAULT_PORTAL_SETTINGS["shortBankName"]).strip(),
        "portalName": str(raw.get("portalName") or DEFAULT_PORTAL_SETTINGS["portalName"]).strip(),
        "emailDomain": normalize_email_domain(raw.get("emailDomain")),
        "branches": normalize_portal_branches(raw.get("branches")),
        "departments": normalize_portal_departments(raw.get("departments")),
        "formCategories": [],
        "departmentChangeTypes": [],
        "transferLocations": [],
        "loginSubtitle": str(raw.get("loginSubtitle") or DEFAULT_PORTAL_SETTINGS["loginSubtitle"]),
        "loginButtonText": str(raw.get("loginButtonText") or DEFAULT_PORTAL_SETTINGS["loginButtonText"]),
        "authorizedAccessText": str(raw.get("authorizedAccessText") or DEFAULT_PORTAL_SETTINGS["authorizedAccessText"]),
        "itAccessCode": str(raw.get("itAccessCode") or DEFAULT_PORTAL_SETTINGS["itAccessCode"]),
        "hrAccessCode": str(raw.get("hrAccessCode") or DEFAULT_PORTAL_SETTINGS["hrAccessCode"]),
        "sessionMinutes": min(60, normalize_positive_number(raw.get("sessionMinutes"), DEFAULT_PORTAL_SETTINGS["sessionMinutes"])),
        "absoluteSessionHours": min(24, normalize_positive_number(raw.get("absoluteSessionHours"), DEFAULT_PORTAL_SETTINGS["absoluteSessionHours"])),
        "sensitiveReauthMinutes": min(30, normalize_positive_number(raw.get("sensitiveReauthMinutes"), DEFAULT_PORTAL_SETTINGS["sensitiveReauthMinutes"])),
        "verificationMinutes": normalize_positive_number(raw.get("verificationMinutes"), DEFAULT_PORTAL_SETTINGS["verificationMinutes"]),
        "passwordResetMinutes": normalize_positive_number(raw.get("passwordResetMinutes"), DEFAULT_PORTAL_SETTINGS["passwordResetMinutes"]),
        "auditRetentionDays": max(365, min(3650, normalize_positive_number(raw.get("auditRetentionDays"), DEFAULT_PORTAL_SETTINGS["auditRetentionDays"]))),
        "notificationRetentionDays": max(7, min(365, normalize_positive_number(raw.get("notificationRetentionDays"), DEFAULT_PORTAL_SETTINGS["notificationRetentionDays"]))),
        "verificationRetentionHours": max(1, min(168, normalize_positive_number(raw.get("verificationRetentionHours"), DEFAULT_PORTAL_SETTINGS["verificationRetentionHours"]))),
        "expiredSessionRetentionDays": max(1, min(90, normalize_positive_number(raw.get("expiredSessionRetentionDays"), DEFAULT_PORTAL_SETTINGS["expiredSessionRetentionDays"]))),
        "securityReviewStatus": str(raw.get("securityReviewStatus") or DEFAULT_PORTAL_SETTINGS["securityReviewStatus"]).strip().lower(),
        "securityReviewProvider": str(raw.get("securityReviewProvider") or "").strip()[:160],
        "securityReviewReference": str(raw.get("securityReviewReference") or "").strip()[:240],
        "securityReviewDate": str(raw.get("securityReviewDate") or "").strip()[:10],
        "dashboardLabel": str(raw.get("dashboardLabel") or DEFAULT_PORTAL_SETTINGS["dashboardLabel"]),
        "formsLabel": str(raw.get("formsLabel") or DEFAULT_PORTAL_SETTINGS["formsLabel"]),
        "appMode": "live" if str(raw.get("appMode", DEFAULT_PORTAL_SETTINGS["appMode"])).strip().lower() == "live" else "test",
        "publicRegistrationEnabled": bool(raw.get("publicRegistrationEnabled", DEFAULT_PORTAL_SETTINGS["publicRegistrationEnabled"])),
        "profileLabel": str(raw.get("profileLabel") or DEFAULT_PORTAL_SETTINGS["profileLabel"]),
        "activeStaffLabel": str(raw.get("activeStaffLabel") or DEFAULT_PORTAL_SETTINGS["activeStaffLabel"]),
        "branchCoverageLabel": str(raw.get("branchCoverageLabel") or DEFAULT_PORTAL_SETTINGS["branchCoverageLabel"]),
        "openOperationsLabel": str(raw.get("openOperationsLabel") or DEFAULT_PORTAL_SETTINGS["openOperationsLabel"]),
        "resolutionRateLabel": str(raw.get("resolutionRateLabel") or DEFAULT_PORTAL_SETTINGS["resolutionRateLabel"]),
        "updatedAt": int(raw.get("updatedAt", 0) or 0),
        "updatedBy": raw.get("updatedBy") if isinstance(raw.get("updatedBy"), dict) else None,
    }


def save_portal_settings_store(settings: dict) -> None:
    atomic_write_json(PORTAL_SETTINGS_STORE_PATH, settings)


def public_portal_settings(settings: dict | None = None) -> dict:
    data = dict(settings or load_portal_settings_store())
    data.pop("portalControlPassword", None)
    data.pop("itAccessCode", None)
    data.pop("hrAccessCode", None)
    return data


def load_portal_authorizations() -> dict[str, dict]:
    raw = read_json_file(PORTAL_AUTHORIZATIONS_PATH, {})
    if not isinstance(raw, dict):
        return {}
    current = now_seconds()
    return {
        str(token_hash): entry
        for token_hash, entry in raw.items()
        if isinstance(entry, dict)
        and str(entry.get("userId", "")).strip()
        and str(entry.get("sessionHash", "")).strip()
        and int(entry.get("expiresAt", 0) or 0) > current
    }


def issue_portal_authorization(user: dict, session_token: str) -> tuple[str, int]:
    token = secrets.token_urlsafe(32)
    expires_at = now_seconds() + PORTAL_AUTHORIZATION_MINUTES * 60
    authorizations = load_portal_authorizations()
    authorizations[session_token_hash(token)] = {
        "userId": user["id"],
        "sessionHash": session_token_hash(session_token),
        "expiresAt": expires_at,
    }
    atomic_write_json(PORTAL_AUTHORIZATIONS_PATH, authorizations)
    return token, expires_at


def portal_authorization_error(data: dict, user: dict, session_token: str):
    token = str(data.get("portalAuthorization", "") or "").strip()
    entry = load_portal_authorizations().get(session_token_hash(token)) if token else None
    valid = bool(
        entry
        and hmac.compare_digest(str(entry.get("userId", "")), str(user.get("id", "")))
        and hmac.compare_digest(str(entry.get("sessionHash", "")), session_token_hash(session_token))
    )
    if not valid:
        return jsonify({"error": "Portal Control authorization expired. Unlock Portal Control again."}), 403
    return None


def next_content_id(items: list[dict], floor: int = 1000) -> int:
    current = floor - 1
    for item in items:
        try:
            current = max(current, int(item.get("id", 0) or 0))
        except Exception:
            continue
    return current + 1


def request_ip_address() -> str:
    forwarded_for = request.headers.get("X-Forwarded-For", "")
    if str(os.getenv("TRUST_PROXY_HEADERS", "")).strip().lower() in {"1", "true", "yes"} and forwarded_for:
        return forwarded_for.split(",")[0].strip() or "unknown"
    return str(request.remote_addr or "unknown")


def rate_limit_key(scope: str, identifier: object) -> str:
    return f"{scope}:{request_ip_address()}:{str(identifier or '').strip().lower()}"


def load_auth_rate_limits() -> dict[str, list[int]]:
    raw = read_json_file(AUTH_RATE_LIMITS_PATH, {})
    if not isinstance(raw, dict):
        return {}
    cutoff = now_seconds() - RATE_LIMIT_WINDOW_SECONDS
    return {
        str(key): [int(stamp) for stamp in stamps if int(stamp or 0) >= cutoff]
        for key, stamps in raw.items()
        if isinstance(stamps, list)
    }


def auth_rate_limited(key: str) -> bool:
    limits = load_auth_rate_limits()
    attempts = limits.get(key, [])
    atomic_write_json(AUTH_RATE_LIMITS_PATH, {item: stamps for item, stamps in limits.items() if stamps})
    return len(attempts) >= RATE_LIMIT_MAX_ATTEMPTS


def user_lockout_status(user: dict, limits: dict[str, list[int]] | None = None) -> tuple[bool, int, int]:
    rate_limits = limits if limits is not None else load_auth_rate_limits()
    identifiers = {
        str(user.get("email", "")).strip().lower(),
        str(user.get("loginUsername", "")).strip().lower(),
    } - {""}
    matching = [
        stamps for key, stamps in rate_limits.items()
        if any(str(key).lower().endswith(f":{identifier}") for identifier in identifiers)
    ]
    attempts = max((len(stamps) for stamps in matching), default=0)
    last_attempt = max((max(stamps) for stamps in matching if stamps), default=0)
    locked = attempts >= RATE_LIMIT_MAX_ATTEMPTS and last_attempt + RATE_LIMIT_WINDOW_SECONDS > now_seconds()
    return locked, attempts, (last_attempt + RATE_LIMIT_WINDOW_SECONDS) * 1000 if locked else 0


def record_auth_failure(key: str) -> None:
    with DATA_LOCK:
        limits = load_auth_rate_limits()
        attempts = limits.get(key, [])
        attempts.append(now_seconds())
        limits[key] = attempts[-RATE_LIMIT_MAX_ATTEMPTS:]
        atomic_write_json(AUTH_RATE_LIMITS_PATH, limits)


def clear_auth_failures(key: str) -> None:
    with DATA_LOCK:
        limits = load_auth_rate_limits()
        if key in limits:
            limits.pop(key, None)
            atomic_write_json(AUTH_RATE_LIMITS_PATH, limits)


def compact_audit_target(target: object) -> str:
    if isinstance(target, str):
        return target
    if not isinstance(target, dict):
        return json.dumps(target, ensure_ascii=True, sort_keys=True)
    parts = []
    for key in [
        "customerId",
        "collectionId",
        "accountNumber",
        "accountName",
        "staffId",
        "staffName",
        "email",
        "username",
        "branch",
        "date",
        "amount",
        "createdCount",
        "skippedCount",
        "created",
        "skipped",
        "action",
        "reason",
    ]:
        if key in target and target.get(key) not in (None, "", [], {}):
            parts.append(f"{key}: {target.get(key)}")
    before = target.get("before")
    after = target.get("after")
    if isinstance(before, dict) and isinstance(after, dict):
        changed = [
            key
            for key in sorted(set(before.keys()) | set(after.keys()))
            if before.get(key) != after.get(key) and key not in {"lastSeen", "updatedAt"}
        ]
        if changed:
            parts.append(f"changed: {', '.join(changed[:8])}")
    return "; ".join(parts) or json.dumps(target, ensure_ascii=True, sort_keys=True)


def load_audit_logs_store() -> list[dict]:
    items = load_json_list_store(AUDIT_LOGS_STORE_PATH)
    normalized = []
    for item in items:
        try:
            normalized.append(
                {
                    "id": int(item.get("id", 0) or 0),
                    "actorId": str(item.get("actorId", "") or "system"),
                    "actorName": str(item.get("actorName", "") or "System"),
                    "action": str(item.get("action", "")).strip(),
                    "target": str(item.get("target", "")).strip(),
                    "ipAddress": str(item.get("ipAddress", "") or "unknown"),
                    "timestamp": int(item.get("timestamp", 0) or 0),
                    "isArchived": bool(item.get("isArchived", False)),
                    "archivedAt": int(item.get("archivedAt", 0) or 0),
                    "archivedBy": str(item.get("archivedBy", "") or ""),
                }
            )
        except Exception:
            continue
    return [
        item
        for item in normalized
        if item["id"] > 0 and item["action"] and item["target"] and item["timestamp"] > 0
    ]


def save_audit_logs_store(items: list[dict]) -> None:
    save_json_list_store(AUDIT_LOGS_STORE_PATH, items)


def merge_audit_logs(existing: list[dict], imported: list[dict]) -> list[dict]:
    merged = list(existing)
    signatures = {
        (
            int(item.get("timestamp", 0) or 0),
            str(item.get("actorId", "")),
            str(item.get("action", "")),
            str(item.get("target", "")),
        )
        for item in merged
    }
    next_id = next_content_id(merged, floor=1)
    for item in imported:
        if not isinstance(item, dict):
            continue
        signature = (
            int(item.get("timestamp", 0) or 0),
            str(item.get("actorId", "")),
            str(item.get("action", "")),
            str(item.get("target", "")),
        )
        if not signature[0] or not signature[2] or signature in signatures:
            continue
        restored = dict(item)
        restored["id"] = next_id
        next_id += 1
        merged.append(restored)
        signatures.add(signature)
    merged.sort(key=lambda item: int(item.get("timestamp", 0) or 0), reverse=True)
    return merged


def record_audit_log(
    actor: dict | None,
    action: str,
    target: object,
    ip_address: str | None = None,
) -> dict:
    logs = load_audit_logs_store()
    target_text = compact_audit_target(target)
    entry = {
        "id": next_content_id(logs, floor=1),
        "actorId": str(actor.get("id", "system") if actor else "system"),
        "actorName": str(actor.get("fullname", "System") if actor else "System"),
        "action": str(action or "").strip().upper(),
        "target": str(target_text or "").strip(),
        "ipAddress": ip_address or request_ip_address(),
        "timestamp": now_ms(),
    }
    logs.insert(0, entry)
    save_audit_logs_store(logs)
    return entry


def has_recent_backup_export(actor: dict, max_age_ms: int = 15 * 60 * 1000) -> bool:
    actor_id = str(actor.get("id", "") or "")
    cutoff = now_ms() - max_age_ms
    return any(
        entry.get("action") == "EXPORT_PRODUCTION_BACKUP"
        and str(entry.get("actorId", "")) == actor_id
        and int(entry.get("timestamp", 0) or 0) >= cutoff
        for entry in load_audit_logs_store()
    )


def backup_confirmation_error(actor: dict, confirmed: object):
    session_token = parse_session_token()
    reauth_error = recent_reauthentication_error(session_token) if session_token else None
    if reauth_error:
        return reauth_error
    if not confirmed:
        return jsonify({"error": "Export a backup before continuing."}), 400
    if not has_recent_backup_export(actor):
        return jsonify({"error": "Your backup confirmation has expired. Export a fresh backup before continuing."}), 400
    return None


def staff_audit_target(user: dict, extra: dict | None = None) -> dict:
    target = {
        "staffId": str(user.get("id", "")),
        "staffName": str(user.get("fullname", "")),
        "email": str(user.get("email", "")),
        "role": str(user.get("role", "")),
        "department": str(user.get("department", "")),
        "branch": str(user.get("branch", "")),
    }
    if extra:
        target.update(extra)
    return target


def session_token_hash(token: str) -> str:
    return hashlib.sha256(str(token or "").encode("utf-8")).hexdigest()


def trusted_device_browser_hash() -> str:
    user_agent = str(request.headers.get("User-Agent", "") or "").strip()
    return hashlib.sha256(user_agent.encode("utf-8")).hexdigest()


def load_trusted_devices() -> dict[str, dict]:
    raw = read_json_file(TRUSTED_DEVICES_PATH, {})
    if not isinstance(raw, dict):
        return {}
    current = now_seconds()
    return {
        str(token_hash): entry
        for token_hash, entry in raw.items()
        if isinstance(entry, dict)
        and str(entry.get("userId", "")).strip()
        and int(entry.get("expiresAt", 0) or 0) > current
        and str(entry.get("browserHash", "")).strip()
    }


def save_trusted_devices(store: dict[str, dict]) -> None:
    atomic_write_json(TRUSTED_DEVICES_PATH, store)


def issue_trusted_device(user_id: str) -> str:
    token = secrets.token_urlsafe(32)
    current = now_seconds()
    devices = load_trusted_devices()
    devices[session_token_hash(token)] = {
        "userId": user_id,
        "browserHash": trusted_device_browser_hash(),
        "createdAt": current,
        "expiresAt": current + TRUSTED_DEVICE_TTL_SECONDS,
    }
    save_trusted_devices(devices)
    return token


def trusted_device_is_valid(user_id: str) -> bool:
    token = str(request.cookies.get(TRUSTED_DEVICE_COOKIE_NAME, "") or "").strip()
    if not token:
        return False
    entry = load_trusted_devices().get(session_token_hash(token))
    return bool(
        entry
        and str(entry.get("userId", "")) == str(user_id)
        and hmac.compare_digest(str(entry.get("browserHash", "")), trusted_device_browser_hash())
    )


def revoke_user_trusted_devices(user_id: str) -> None:
    devices = load_trusted_devices()
    filtered = {
        token_hash: entry
        for token_hash, entry in devices.items()
        if str(entry.get("userId", "")) != str(user_id)
    }
    if filtered != devices:
        save_trusted_devices(filtered)


def revoke_current_trusted_device() -> bool:
    token = str(request.cookies.get(TRUSTED_DEVICE_COOKIE_NAME, "") or "").strip()
    if not token:
        return False
    devices = load_trusted_devices()
    removed = devices.pop(session_token_hash(token), None) is not None
    if removed:
        save_trusted_devices(devices)
    return removed


def load_sessions() -> dict[str, dict]:
    raw = read_json_file(SESSIONS_STORE_PATH, {})
    if not isinstance(raw, dict):
        return {}
    current = now_seconds()
    sessions = {}
    for token, item in raw.items():
        if not isinstance(token, str) or not isinstance(item, dict):
            continue
        # Sessions created before idle/absolute expiry tracking are intentionally
        # invalidated so an old 30-day token cannot bypass the tightened policy.
        if "absoluteExpiresAt" not in item or "lastActivityAt" not in item:
            continue
        user_id = str(item.get("userId", "")).strip()
        expires_at = int(item.get("expiresAt", 0) or 0)
        absolute_expires_at = int(item.get("absoluteExpiresAt", expires_at) or 0)
        last_activity_at = int(item.get("lastActivityAt", current) or 0)
        recent_auth_at = int(item.get("recentAuthAt", last_activity_at) or 0)
        if not user_id or expires_at <= current or absolute_expires_at <= current:
            continue
        sessions[token] = {
            "userId": user_id,
            "expiresAt": expires_at,
            "absoluteExpiresAt": absolute_expires_at,
            "lastActivityAt": last_activity_at,
            "recentAuthAt": recent_auth_at,
            "createdAt": int(item.get("createdAt", last_activity_at) or last_activity_at),
            "ipAddress": str(item.get("ipAddress", "unknown") or "unknown")[:80],
            "userAgent": str(item.get("userAgent", "Unknown device") or "Unknown device")[:300],
        }
    return sessions


def save_sessions(store: dict[str, dict]) -> None:
    normalized = {}
    for token, session in store.items():
        key = str(token or "").strip()
        if not key:
            continue
        if len(key) != 64 or any(ch not in "0123456789abcdef" for ch in key.lower()):
            key = session_token_hash(key)
        normalized[key] = session
    atomic_write_json(SESSIONS_STORE_PATH, normalized)


def issue_session(user_id: str) -> str:
    sessions = load_sessions()
    token = secrets.token_urlsafe(32)
    settings = load_portal_settings_store()
    current = now_seconds()
    sessions[session_token_hash(token)] = {
        "userId": user_id,
        "expiresAt": current + int(settings["sessionMinutes"]) * 60,
        "absoluteExpiresAt": current + int(settings["absoluteSessionHours"]) * 60 * 60,
        "lastActivityAt": current,
        "recentAuthAt": current,
        "createdAt": current,
        "ipAddress": request_ip_address() if has_request_context() else "system",
        "userAgent": str(request.headers.get("User-Agent", "Unknown device") or "Unknown device")[:300] if has_request_context() else "System test client",
    }
    save_sessions(sessions)
    return token


def session_cookie_secure() -> bool:
    configured_url = portal_public_url()
    if configured_url:
        return urlparse(configured_url).scheme.lower() == "https"
    forwarded_proto = str(request.headers.get("X-Forwarded-Proto", "") or "").split(",")[0].strip().lower()
    return request.is_secure or forwarded_proto == "https"


def set_session_cookie(response, token: str):
    max_age = int(load_portal_settings_store()["absoluteSessionHours"]) * 60 * 60
    response.set_cookie(
        SESSION_COOKIE_NAME,
        token,
        max_age=max_age,
        secure=session_cookie_secure(),
        httponly=True,
        samesite="Strict",
        path="/",
    )
    return response


def clear_session_cookie(response):
    response.delete_cookie(
        SESSION_COOKIE_NAME,
        secure=session_cookie_secure(),
        httponly=True,
        samesite="Strict",
        path="/",
    )
    return response


def set_trusted_device_cookie(response, token: str):
    response.set_cookie(
        TRUSTED_DEVICE_COOKIE_NAME,
        token,
        max_age=TRUSTED_DEVICE_TTL_SECONDS,
        secure=session_cookie_secure(),
        httponly=True,
        samesite="Strict",
        path="/",
    )
    return response


def clear_trusted_device_cookie(response):
    response.delete_cookie(
        TRUSTED_DEVICE_COOKIE_NAME,
        secure=session_cookie_secure(),
        httponly=True,
        samesite="Strict",
        path="/",
    )
    return response


def authenticated_response(user: dict, token: str):
    return set_session_cookie(jsonify({"ok": True, "user": user}), token)


def revoke_session(token: str) -> None:
    sessions = load_sessions()
    hashed = session_token_hash(token)
    if hashed in sessions or token in sessions:
        sessions.pop(hashed, None)
        sessions.pop(token, None)
        save_sessions(sessions)


def revoke_user_sessions(user_id: str) -> None:
    sessions = load_sessions()
    filtered = {
        token: session
        for token, session in sessions.items()
        if session.get("userId") != user_id
    }
    if filtered != sessions:
        save_sessions(filtered)
    revoke_user_trusted_devices(user_id)


_RETENTION_LOCK = threading.Lock()
_LAST_RETENTION_RUN = 0


def apply_retention_rules() -> dict:
    settings = load_portal_settings_store()
    now_s = now_seconds()
    now_millis = now_ms()
    summary = {}

    audit_items = load_audit_logs_store()
    audit_cutoff = now_millis - int(settings["auditRetentionDays"]) * 86400000
    kept_audits = [item for item in audit_items if int(item.get("timestamp", 0) or 0) >= audit_cutoff]
    summary["auditLogs"] = len(audit_items) - len(kept_audits)
    if len(kept_audits) != len(audit_items):
        save_audit_logs_store(kept_audits)

    notifications = load_json_list_store(NOTIFICATIONS_STORE_PATH)
    notification_cutoff = now_millis - int(settings["notificationRetentionDays"]) * 86400000
    kept_notifications = [
        item for item in notifications
        if not bool(item.get("isRead", False))
        or int(item.get("createdAt", 0) or 0) >= notification_cutoff
    ]
    summary["notifications"] = len(notifications) - len(kept_notifications)
    if len(kept_notifications) != len(notifications):
        save_json_list_store(NOTIFICATIONS_STORE_PATH, kept_notifications)

    verification_grace = int(settings["verificationRetentionHours"]) * 3600
    token_paths = [
        PENDING_VERIFICATIONS_PATH,
        RESET_TOKENS_PATH,
        AGENT_SETUP_TOKENS_PATH,
        PRIVILEGED_MFA_PATH,
        PORTAL_AUTHORIZATIONS_PATH,
    ]
    removed_tokens = 0
    for path in token_paths:
        raw = read_json_file(path, {})
        if not isinstance(raw, dict):
            continue
        kept = {
            key: value for key, value in raw.items()
            if isinstance(value, dict)
            and int(value.get("expiresAt", 0) or 0) + verification_grace >= now_s
        }
        removed_tokens += len(raw) - len(kept)
        if len(kept) != len(raw):
            atomic_write_json(path, kept)
    summary["verificationRecords"] = removed_tokens

    raw_sessions = read_json_file(SESSIONS_STORE_PATH, {})
    session_grace = int(settings["expiredSessionRetentionDays"]) * 86400
    if isinstance(raw_sessions, dict):
        kept_sessions = {
            key: value for key, value in raw_sessions.items()
            if isinstance(value, dict)
            and int(value.get("absoluteExpiresAt", value.get("expiresAt", 0)) or 0) + session_grace >= now_s
        }
        summary["sessions"] = len(raw_sessions) - len(kept_sessions)
        if len(kept_sessions) != len(raw_sessions):
            save_sessions(kept_sessions)
    else:
        summary["sessions"] = 0
    summary["completedAt"] = now_millis
    return summary


@app.before_request
def run_periodic_retention_cleanup():
    global _LAST_RETENTION_RUN
    current = now_seconds()
    if current - _LAST_RETENTION_RUN < 6 * 3600:
        return None
    if not _RETENTION_LOCK.acquire(blocking=False):
        return None
    try:
        if current - _LAST_RETENTION_RUN >= 6 * 3600:
            apply_retention_rules()
            _LAST_RETENTION_RUN = current
    except Exception:
        app.logger.exception("Retention cleanup failed")
    finally:
        _RETENTION_LOCK.release()
    return None


@app.route("/api/owner/retention/run", methods=["POST", "OPTIONS"])
def owner_run_retention():
    preflight = handle_options()
    if preflight:
        return preflight
    session_token, auth_user, error = require_owner_admin()
    if error:
        return error
    reauth_error = recent_reauthentication_error(session_token)
    if reauth_error:
        return reauth_error
    summary = apply_retention_rules()
    record_audit_log(auth_user, "RUN_RETENTION_CLEANUP", summary)
    return jsonify({"ok": True, "summary": summary})

def require_authenticated_user():
    token = parse_session_token()
    if not token:
        return None, None, (jsonify({"error": "Authentication required"}), 401)
    sessions = load_sessions()
    session = sessions.get(session_token_hash(token)) or sessions.get(token)
    if not session:
        return None, None, (jsonify({"error": "Invalid or expired session"}), 401)
    users = load_user_store()
    user = find_user_by_id(users, session["userId"])
    if not user or user["isArchived"] or not user["isActive"] or not user["isVerified"]:
        revoke_session(token)
        return None, None, (jsonify({"error": "Invalid or expired session"}), 401)
    current = now_seconds()
    if current - int(session.get("lastActivityAt", current) or current) >= 60:
        settings = load_portal_settings_store()
        session["lastActivityAt"] = current
        session["expiresAt"] = min(
            current + int(settings["sessionMinutes"]) * 60,
            int(session.get("absoluteExpiresAt", current) or current),
        )
        sessions[session_token_hash(token)] = session
        sessions.pop(token, None)
        save_sessions(sessions)
    return token, user, None


def recent_reauthentication_error(session_token: str):
    sessions = load_sessions()
    session = sessions.get(session_token_hash(session_token)) or sessions.get(session_token)
    settings = load_portal_settings_store()
    cutoff = now_seconds() - int(settings["sensitiveReauthMinutes"]) * 60
    if not session or int(session.get("recentAuthAt", 0) or 0) < cutoff:
        return jsonify({
            "error": "Confirm your password before continuing with this sensitive action.",
            "code": "REAUTHENTICATION_REQUIRED",
        }), 428
    return None


def require_staff_manager():
    token, user, error = require_authenticated_user()
    if error:
        return token, user, error
    if not is_owner_admin(user):
        return token, user, (jsonify({"error": "Admin access required"}), 403)
    return token, user, None


def require_owner_admin():
    token, user, error = require_authenticated_user()
    if error:
        return token, user, error
    if not is_owner_admin(user):
        return token, user, (jsonify({"error": "Owner admin access required"}), 403)
    return token, user, None


def generate_verification_code() -> str:
    return f"{secrets.randbelow(900000) + 100000:06d}"


def build_reset_url(base_url: str, token: str) -> str:
    trusted_base = build_portal_link("/reset-password")
    if not trusted_base:
        raise ValueError("PORTAL_PUBLIC_URL must be configured before password reset links can be sent")
    parsed = urlparse(trusted_base)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise ValueError("PORTAL_PUBLIC_URL must be a valid HTTPS or HTTP URL")
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query["token"] = token
    return urlunparse(parsed._replace(query=urlencode(query)))


def mail_config() -> dict[str, str | int]:
    required = {
        "MAIL_SERVER": os.getenv("MAIL_SERVER", ""),
        "MAIL_USERNAME": os.getenv("MAIL_USERNAME", ""),
        "MAIL_PASSWORD": os.getenv("MAIL_PASSWORD", ""),
        "MAIL_DEFAULT_SENDER": os.getenv("MAIL_DEFAULT_SENDER", ""),
    }
    missing = [key for key, value in required.items() if not value]
    if missing:
        raise RuntimeError(f"Missing mail configuration: {', '.join(missing)}")
    return {
        **required,
        "MAIL_PORT": int(os.getenv("MAIL_PORT", "465")),
    }


def send_mail(to_email: str, subject: str, text_body: str, html_body: str):
    cfg = mail_config()
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = str(cfg["MAIL_DEFAULT_SENDER"])
    msg["To"] = to_email
    msg.set_content(text_body)
    msg.add_alternative(html_body, subtype="html")

    with smtplib.SMTP_SSL(str(cfg["MAIL_SERVER"]), int(cfg["MAIL_PORT"]), timeout=30) as smtp:
        smtp.login(str(cfg["MAIL_USERNAME"]), str(cfg["MAIL_PASSWORD"]))
        smtp.send_message(msg)


def send_sms_token(phone: str, code: str, purpose: str = "agent_setup") -> bool:
    webhook = env_secret("SMS_WEBHOOK_URL")
    if not webhook:
        return False
    payload = json.dumps({
        "phone": phone,
        "message": f"Your BCB SUSU verification code is {code}. It expires shortly.",
        "code": code,
        "purpose": purpose,
    }).encode("utf-8")
    headers = {"Content-Type": "application/json"}
    api_key = env_secret("SMS_WEBHOOK_API_KEY")
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    req = Request(webhook, data=payload, headers=headers, method="POST")
    try:
        with urlopen(req, timeout=15) as response:
            return 200 <= int(response.status) < 300
    except Exception as exc:
        app.logger.warning("SMS token delivery failed: %s", exc)
        return False


def send_verification_code_email(email: str, code: str) -> None:
    text_body = (
        "Dear Staff,\n\n"
        f"Your Bawjiase SUSU Collection Portal verification code is: {code}\n\n"
        "This code expires in 15 minutes.\n\n"
        "Thank you.\nBawjiase Community Bank PLC"
    )
    html_body = f"""
    <html>
      <body style="font-family: Arial, sans-serif; color: #1f2937; line-height: 1.6;">
        <div style="max-width: 600px; margin: 0 auto; padding: 24px; border: 1px solid #e5e7eb; border-radius: 12px;">
          <h2 style="color: #15803d; text-align: center;">Email Verification</h2>
          <p>Dear Staff,</p>
          <p>Use this code to verify your email address for the <strong>Bawjiase SUSU Collection Portal</strong>:</p>
          <div style="text-align: center; margin: 28px 0;">
            <span style="display: inline-block; border: 2px solid #15803d; color: #15803d; padding: 14px 28px; font-size: 24px; font-weight: 700; border-radius: 8px; letter-spacing: 5px;">{code}</span>
          </div>
          <p>This code expires in 15 minutes.</p>
          <p>If you did not request this code, please ignore this email.</p>
          <p style="font-weight: 700; color: #4b5563;">Bawjiase Community Bank PLC</p>
        </div>
      </body>
    </html>
    """
    send_mail(email, "Bawjiase SUSU Collection Portal - Email Verification Code", text_body, html_body)


def privileged_mfa_required(user: dict) -> bool:
    return str(user.get("role", "")).strip() in {OWNER_ADMIN_ROLE, "Supervisor"}


def load_privileged_mfa_challenges() -> dict[str, dict]:
    raw = read_json_file(PRIVILEGED_MFA_PATH, {})
    if not isinstance(raw, dict):
        return {}
    current = now_seconds()
    return {
        str(challenge_id): entry
        for challenge_id, entry in raw.items()
        if isinstance(entry, dict)
        and str(entry.get("userId", "")).strip()
        and int(entry.get("expiresAt", 0) or 0) > current
    }


def privileged_mfa_code_hash(challenge_id: str, code: str) -> str:
    return hashlib.sha256(f"{challenge_id}:{code}".encode("utf-8")).hexdigest()


def issue_privileged_mfa_challenge(user: dict, purpose: str = "login") -> dict:
    challenge_id = secrets.token_urlsafe(24)
    code = generate_verification_code()
    expires_at = now_seconds() + 10 * 60
    challenges = load_privileged_mfa_challenges()
    challenges[challenge_id] = {
        "userId": user["id"],
        "codeHash": privileged_mfa_code_hash(challenge_id, code),
        "expiresAt": expires_at,
        "attempts": 0,
        "purpose": purpose,
    }
    atomic_write_json(PRIVILEGED_MFA_PATH, challenges)
    text_body = (
        f"Your BCB SUSU security verification code is {code}.\n\n"
        "It expires in 10 minutes. If you did not request this code, contact the system owner immediately."
    )
    html_body = (
        "<div style='font-family:Arial,sans-serif;max-width:560px;margin:auto;padding:24px'>"
        "<h2 style='color:#047857'>Security verification</h2>"
        f"<p>Your verification code is:</p><p style='font-size:28px;font-weight:700;letter-spacing:6px'>{code}</p>"
        "<p>This code expires in 10 minutes.</p></div>"
    )
    is_test_mode = str(load_portal_settings_store().get("appMode", "test")).strip().lower() != "live"
    if not is_test_mode:
        try:
            send_mail(user["email"], "BCB SUSU security verification", text_body, html_body)
        except Exception:
            challenges.pop(challenge_id, None)
            atomic_write_json(PRIVILEGED_MFA_PATH, challenges)
            raise
    result = {"challengeId": challenge_id, "expiresAt": expires_at}
    if is_test_mode:
        result["testCode"] = code
    return result


def send_password_reset_link_email(email: str, reset_url: str) -> None:
    text_body = (
        "Dear Staff,\n\n"
        "Use the link below to reset your Bawjiase SUSU Collection Portal password:\n"
        f"{reset_url}\n\n"
        "This link expires in 30 minutes.\n\n"
        "Bawjiase Community Bank PLC"
    )
    html_body = f"""
    <html>
      <body style="font-family: Arial, sans-serif; color: #1f2937; line-height: 1.6;">
        <div style="max-width: 600px; margin: 0 auto; padding: 24px; border: 1px solid #e5e7eb; border-radius: 12px;">
          <h2 style="color: #15803d; text-align: center;">Password Reset</h2>
          <p>Dear Staff,</p>
          <p>Use the button below to reset your Bawjiase SUSU Collection Portal password.</p>
          <p style="text-align: center; margin: 28px 0;">
            <a href="{reset_url}" style="background: #15803d; color: #ffffff; padding: 12px 22px; border-radius: 8px; text-decoration: none; font-weight: 700;">Reset Password</a>
          </p>
          <p>This link expires in 30 minutes.</p>
          <p style="font-weight: 700; color: #4b5563;">Bawjiase Community Bank PLC</p>
        </div>
      </body>
    </html>
    """
    send_mail(email, "Bawjiase SUSU Collection Portal - Password Reset", text_body, html_body)


def portal_public_url() -> str:
    return os.getenv("PORTAL_PUBLIC_URL", "").strip().rstrip("/")


def build_portal_link(path: str) -> str | None:
    base = portal_public_url()
    if not base:
        return None
    return f"{base}{path if path.startswith('/') else f'/{path}'}"


def value_in_scope(scope: list[str], current_value: str) -> bool:
    if "ALL" in scope:
        return True
    return str(current_value or "").strip().upper() in scope


def branch_allowed_for_user(user: dict, branch: str) -> bool:
    if is_global_manager(user):
        return True
    managed_branches = normalize_scope_list(user.get("managedBranches"), empty_default=[])
    return value_in_scope(managed_branches, branch)


def department_allowed_for_user(user: dict, branch: str, department: str) -> bool:
    if is_global_manager(user):
        return True
    managed = normalize_managed_departments_by_branch(user.get("managedDepartmentsByBranch"))
    branch_departments = managed.get(str(branch or "").strip().upper())
    if not branch_departments:
        return False
    return value_in_scope(branch_departments, department)


def can_manage_scope(user: dict, branch_scope: list[str], department_scope: list[str]) -> bool:
    if is_global_manager(user):
        return True
    if "ALL" in branch_scope:
        return False
    normalized_departments = department_scope if department_scope else ["ALL"]
    managed_departments = normalize_managed_departments_by_branch(
        user.get("managedDepartmentsByBranch")
    )
    for branch in [item for item in branch_scope if item != "ALL"]:
        if not branch_allowed_for_user(user, branch):
            return False
        branch_managed_departments = managed_departments.get(branch, [])
        for department in normalized_departments:
            if department == "ALL":
                if "ALL" not in branch_managed_departments:
                    return False
            elif not department_allowed_for_user(user, branch, department):
                return False
    return True


def is_assigned_supervisor(user: dict | None) -> bool:
    if not user or str(user.get("role", "")).strip() != "Supervisor":
        return False
    return bool(normalize_scope_list(user.get("managedBranches"), empty_default=[]))


def is_susu_agent(user: dict | None) -> bool:
    if not user:
        return False
    department = str(user.get("department", "")).strip().upper()
    role = str(user.get("role", "")).strip()
    return department in SUSU_DEPARTMENTS and role not in {OWNER_ADMIN_ROLE, "Supervisor", "SuperAdmin"}


def can_manage_agents_and_customers(user: dict | None) -> bool:
    return is_global_manager(user) or is_assigned_supervisor(user)


def managed_branch_for_user(user: dict, requested_branch: object = None) -> str:
    requested = str(requested_branch or "").strip().upper()
    if is_global_manager(user):
        return normalize_portal_branch_name(requested or user.get("branch"))
    managed = normalize_scope_list(user.get("managedBranches"), empty_default=[])
    branch = requested or str(user.get("branch") or "").strip().upper()
    if not branch and managed:
        branch = managed[0]
    branch = normalize_portal_branch_name(branch)
    if not branch_allowed_for_user(user, branch):
        raise ValueError("You can only manage records for your assigned branch.")
    return branch


def normalize_agent_username(value: object) -> str:
    username = str(value or "").strip().lower()
    if len(username) < 3:
        raise ValueError("Username must be at least 3 characters.")
    allowed = set("abcdefghijklmnopqrstuvwxyz0123456789._-")
    if any(char not in allowed for char in username):
        raise ValueError("Username can only contain letters, numbers, dot, dash, and underscore.")
    return username


def agent_password_key(username: str) -> str:
    return f"username:{normalize_agent_username(username)}"


def find_user_by_username(users: list[dict], username: str):
    normalized = normalize_agent_username(username)
    return next((user for user in users if str(user.get("loginUsername", "")).strip().lower() == normalized), None)


def find_user_by_username_safe(users: list[dict], username: object):
    try:
        return find_user_by_username(users, normalize_agent_username(username))
    except ValueError:
        return None


def can_view_operational_record(user: dict, item: dict) -> bool:
    if is_global_manager(user):
        return True
    branch = str(item.get("branch_name") or item.get("branch_id") or item.get("branch") or "").strip().upper()
    if is_assigned_supervisor(user):
        return branch_allowed_for_user(user, branch)
    is_customer_record = "account_number" in item and "amount" not in item and "agent_id" not in item
    if is_susu_agent(user) and is_customer_record:
        return branch == str(user.get("branch") or "").strip().upper()
    owner_ids = {
        str(item.get("createdById", "") or ""),
        str(item.get("agent_id", "") or ""),
        str(item.get("recordedById", "") or ""),
    }
    owner_emails = {
        str(item.get("createdByEmail", "") or "").strip().lower(),
        str(item.get("agent_email", "") or "").strip().lower(),
        str(item.get("recordedByEmail", "") or "").strip().lower(),
    }
    owner_names = {
        str(item.get("createdBy", "") or "").strip().lower(),
        str(item.get("agent_name", "") or "").strip().lower(),
        str(item.get("recorded_by", "") or "").strip().lower(),
    }
    return (
        str(user.get("id", "")) in owner_ids
        or str(user.get("email", "")).strip().lower() in owner_emails
        or str(user.get("fullname", "")).strip().lower() in owner_names
    )


def can_view_staff_record(viewer: dict, staff_user: dict) -> bool:
    if is_global_manager(viewer):
        return True
    if str(viewer.get("id")) == str(staff_user.get("id")):
        return True
    if not is_assigned_supervisor(viewer):
        return False
    return branch_allowed_for_user(viewer, staff_user.get("branch", ""))


def manageable_scope_message(user: dict) -> str:
    if is_global_manager(user):
        return "You can manage all SUSU branches."
    managed_branches = normalize_scope_list(user.get("managedBranches"), empty_default=[])
    if not managed_branches:
        return "No supervisor branch scope is assigned to your account."
    return f"You can only manage these SUSU branches: {', '.join(managed_branches)}."


def scoped_access_denial(user: dict):
    return jsonify({"error": manageable_scope_message(user)}), 403


def create_notifications_for_users(
    users: list[dict],
    *,
    kind: str,
    title: str,
    message: str,
    link_to: str | None,
) -> int:
    items = load_json_list_store(NOTIFICATIONS_STORE_PATH)
    created_at = now_ms()
    count = 0
    for user in users:
        items.insert(
            0,
            {
                "id": next_content_id(items, floor=1),
                "userId": user["id"],
                "kind": kind,
                "title": title,
                "message": message,
                "linkTo": link_to,
                "isRead": False,
                "createdAt": created_at,
            },
        )
        count += 1
    save_json_list_store(NOTIFICATIONS_STORE_PATH, items)
    return count


def notify_active_managers(*, kind: str, title: str, message: str, link_to: str | None) -> int:
    users = [
        user
        for user in load_user_store()
        if user.get("isActive")
        and user.get("isVerified")
        and not user.get("isArchived")
        and (
            str(user.get("role")) in GLOBAL_MANAGER_ROLES
            or user_has_permission(user, "userManagement")
        )
    ]
    return create_notifications_for_users(
        users,
        kind=kind,
        title=title,
        message=message,
        link_to=link_to,
    )


def find_user_by_local_image(filename: str) -> dict | None:
    users = load_user_store()
    return next(
        (
            user
            for user in users
            if is_local_upload_ref(user.get("imageFile"), filename)
        ),
        None,
    )


def local_filename_from_ref(value: object) -> str:
    raw = str(value or "").strip()
    if not raw.startswith("LOCAL:"):
        return ""
    return raw.replace("LOCAL:", "", 1).strip()


def remove_uploaded_file_if_unused(filename: str) -> None:
    if not filename:
        return
    if find_user_by_local_image(filename):
        return
    file_path = os.path.join(UPLOADS_DIR, filename)
    if os.path.isfile(file_path):
        os.remove(file_path)


def handle_options():
    if request.method == "OPTIONS":
        return ("", 204)
    return None


def send_persisted_upload(filename: str):
    if pg_enabled():
        ensure_pg_store_table()
        with pg_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT content_type, payload FROM portal_uploads WHERE filename = %s", (filename,))
                row = cur.fetchone()
        if row:
            return send_file(
                io.BytesIO(bytes(row[1])),
                mimetype=str(row[0] or "application/octet-stream"),
                download_name=filename,
                conditional=True,
                max_age=3600,
            )
    local_path = os.path.join(UPLOADS_DIR, filename)
    if os.path.isfile(local_path):
        return send_from_directory(UPLOADS_DIR, filename, conditional=True)
    return jsonify({"error": "File not found"}), 404


@app.route("/api/health", methods=["GET"])
def health():
    started_at = time.perf_counter()
    database_required = str(os.getenv("PORTAL_REQUIRE_DATABASE", "")).strip().lower() in {"1", "true", "yes", "on"}
    if not pg_enabled():
        payload = {
            "ok": not database_required,
            "status": "degraded" if not database_required else "unhealthy",
            "storageBackend": "json-file",
            "database": {"configured": False, "reachable": False},
            "durationMs": round((time.perf_counter() - started_at) * 1000),
        }
        return jsonify(payload), 200 if payload["ok"] else 503
    try:
        with pg_connect() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
                row = cur.fetchone()
        reachable = bool(row and row[0] == 1)
    except Exception as exc:
        app.logger.error("Database readiness check failed: %s", exc)
        reachable = False
    payload = {
        "ok": reachable,
        "status": "healthy" if reachable else "unhealthy",
        "storageBackend": "postgres",
        "database": {"configured": True, "reachable": reachable},
        "durationMs": round((time.perf_counter() - started_at) * 1000),
    }
    return jsonify(payload), 200 if reachable else 503


@app.route("/api/production-status", methods=["GET"])
def production_status():
    _, _, error = require_owner_admin()
    if error:
        return error
    test_staff_count = sum(1 for item in load_user_store() if bool(item.get("isTestData")))
    test_customer_count = sum(1 for item in load_json_list_store(CUSTOMERS_STORE_PATH) if bool(item.get("isTestData")))
    settings = load_portal_settings_store()
    security_review_complete = (
        settings.get("securityReviewStatus") == "completed"
        and bool(settings.get("securityReviewProvider"))
        and bool(settings.get("securityReviewReference"))
        and bool(settings.get("securityReviewDate"))
    )
    checks = {
        "database": {
            "ok": pg_enabled(),
            "label": "PostgreSQL DATABASE_URL configured",
        },
        "portalPublicUrl": {
            "ok": bool(portal_public_url()),
            "label": "PORTAL_PUBLIC_URL configured",
        },
        "mail": {
            "ok": all(env_secret(name) for name in ["MAIL_SERVER", "MAIL_USERNAME", "MAIL_PASSWORD", "MAIL_DEFAULT_SENDER"]),
            "label": "Mail server configured",
        },
        "sms": {
            "ok": bool(env_secret("SMS_WEBHOOK_URL")),
            "label": "SMS webhook configured",
        },
        "monitoring": {
            "ok": monitoring_destination_configured(),
            "label": "Monitoring webhook or alert email configured",
        },
        "testDataRemoved": {
            "ok": test_staff_count == 0 and test_customer_count == 0,
            "label": f"Test data removed ({test_staff_count} staff, {test_customer_count} customers remaining)",
        },
        "independentSecurityReview": {
            "ok": security_review_complete,
            "label": "Independent security review completed with evidence",
        },
    }
    required = ["database", "portalPublicUrl", "mail", "monitoring", "testDataRemoved", "independentSecurityReview"]
    live_ready = all(checks[key]["ok"] for key in required)
    return jsonify({
        "storageBackend": "postgres" if pg_enabled() else "json-file",
        "liveReady": live_ready,
        "checks": checks,
        "required": required,
    })


@app.route("/uploads/<path:filename>", methods=["GET"])
def get_uploaded_media(filename: str):
    safe_name = secure_filename(filename)
    if not safe_name or safe_name != filename:
        return jsonify({"error": "Invalid file name"}), 400
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    if find_user_by_local_image(safe_name):
        return send_persisted_upload(safe_name)
    return jsonify({"error": "File not found"}), 404


@app.route("/mail-api/uploads/<path:filename>", methods=["GET"])
def get_uploaded_media_legacy(filename: str):
    return get_uploaded_media(filename)


@app.route("/mail-api/api/<path:path>", methods=["GET", "POST", "OPTIONS"])
def legacy_mail_api(path: str):
    destination = f"/api/{path}"
    query = request.query_string.decode("utf-8", errors="ignore").strip()
    if query:
        destination = f"{destination}?{query}"
    return redirect(destination, code=307)


@app.route("/api/presence", methods=["GET"])
def get_presence():
    _, _, error = require_authenticated_user()
    if error:
        return error
    store = prune_presence(load_presence_store())
    save_presence_store(store)
    return jsonify({"presence": store})


@app.route("/api/presence/ping", methods=["POST", "OPTIONS"])
def ping_presence():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    user_id = str(data.get("userId", "")).strip()
    if not user_id:
        return jsonify({"error": "userId is required"}), 400
    if auth_user["id"] != user_id:
        return jsonify({"error": "Cannot update another user's presence"}), 403
    current_ms = now_ms()
    set_user_last_seen(user_id, current_ms)
    store = prune_presence(load_presence_store())
    store[user_id] = int(time.time())
    save_presence_store(store)
    return jsonify({"ok": True, "lastSeen": current_ms})


@app.route("/api/presence/logout", methods=["POST", "OPTIONS"])
def logout_presence():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    user_id = str(data.get("userId", "")).strip()
    if not user_id:
        return jsonify({"error": "userId is required"}), 400
    if auth_user["id"] != user_id:
        return jsonify({"error": "Cannot update another user's presence"}), 403
    set_user_last_seen(user_id, now_ms())
    store = prune_presence(load_presence_store())
    store.pop(user_id, None)
    save_presence_store(store)
    return jsonify({"ok": True})


@app.route("/api/notifications", methods=["GET"])
def get_notifications():
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    items = load_json_list_store(NOTIFICATIONS_STORE_PATH)
    user_items = [
        item
        for item in items
        if str(item.get("userId", "")).strip() == auth_user["id"]
    ]
    user_items.sort(key=lambda item: int(item.get("createdAt", 0) or 0), reverse=True)
    return jsonify({"notifications": user_items})


@app.route("/api/notifications/unread-count", methods=["GET"])
def get_unread_notification_count():
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    items = load_json_list_store(NOTIFICATIONS_STORE_PATH)
    count = sum(
        1
        for item in items
        if str(item.get("userId", "")).strip() == auth_user["id"]
        and not bool(item.get("isRead", False))
    )
    return jsonify({"count": count})


@app.route("/api/notifications/<int:item_id>/read", methods=["POST", "OPTIONS"])
def mark_notification_read(item_id: int):
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    items = load_json_list_store(NOTIFICATIONS_STORE_PATH)
    notification = next(
        (
            item
            for item in items
            if int(item.get("id", 0) or 0) == item_id
            and str(item.get("userId", "")).strip() == auth_user["id"]
        ),
        None,
    )
    if not notification:
        return jsonify({"error": "Notification not found"}), 404
    notification["isRead"] = True
    save_json_list_store(NOTIFICATIONS_STORE_PATH, items)
    return jsonify({"ok": True})


@app.route("/api/notifications/read-all", methods=["POST", "OPTIONS"])
def mark_all_notifications_read():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    items = load_json_list_store(NOTIFICATIONS_STORE_PATH)
    changed = False
    for item in items:
        if str(item.get("userId", "")).strip() == auth_user["id"] and not bool(item.get("isRead", False)):
            item["isRead"] = True
            changed = True
    if changed:
        save_json_list_store(NOTIFICATIONS_STORE_PATH, items)
    return jsonify({"ok": True})


@app.route("/api/notifications/<int:item_id>/delete", methods=["POST", "OPTIONS"])
def delete_notification(item_id: int):
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    items = load_json_list_store(NOTIFICATIONS_STORE_PATH)
    filtered = [
        item
        for item in items
        if not (
            int(item.get("id", 0) or 0) == item_id
            and str(item.get("userId", "")).strip() == auth_user["id"]
        )
    ]
    if len(filtered) == len(items):
        return jsonify({"error": "Notification not found"}), 404
    save_json_list_store(NOTIFICATIONS_STORE_PATH, filtered)
    record_audit_log(
        auth_user,
        "DELETE_NOTIFICATION",
        {"notificationId": item_id},
    )
    return jsonify({"ok": True})


@app.route("/api/audit-logs", methods=["GET"])
def get_audit_logs():
    _, _, error = require_owner_admin()
    if error:
        return error
    logs = sorted(
        [item for item in load_audit_logs_store() if not bool(item.get("isArchived"))],
        key=lambda item: int(item["timestamp"]),
        reverse=True,
    )
    if pagination_requested():
        return paginated_response("logs", logs)
    return jsonify({"logs": logs})


@app.route("/api/audit-logs", methods=["POST", "OPTIONS"])
def create_audit_log():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_owner_admin()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    target = str(data.get("target", "")).strip()
    if not target:
        return jsonify({"error": "A manual audit note is required"}), 400
    entry = record_audit_log(
        auth_user,
        "MANUAL_AUDIT_NOTE",
        target[:1000],
        request_ip_address(),
    )
    return jsonify({"ok": True, "log": entry})


@app.route("/api/audit-logs/<int:item_id>/delete", methods=["POST", "OPTIONS"])
def delete_audit_log(item_id: int):
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_owner_admin()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    backup_error = backup_confirmation_error(auth_user, data.get("backupConfirmed"))
    if backup_error:
        return backup_error
    logs = load_audit_logs_store()
    item = next((entry for entry in logs if int(entry.get("id", 0) or 0) == item_id), None)
    if not item:
        return jsonify({"error": "Log entry not found"}), 404
    item["isArchived"] = True
    item["archivedAt"] = now_ms()
    item["archivedBy"] = auth_user["id"]
    save_audit_logs_store(logs)
    record_audit_log(auth_user, "ARCHIVE_AUDIT_LOG", {"archivedId": item_id})
    return jsonify({"ok": True})


@app.route("/api/audit-logs/delete", methods=["POST", "OPTIONS"])
def delete_audit_logs():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_owner_admin()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    backup_error = backup_confirmation_error(auth_user, data.get("backupConfirmed"))
    if backup_error:
        return backup_error
    ids = {
        int(item)
        for item in data.get("ids", [])
        if isinstance(item, (int, float, str)) and str(item).strip().isdigit()
    }
    if not ids:
        return jsonify({"ok": True})
    logs = load_audit_logs_store()
    archived_at = now_ms()
    for item in logs:
        if int(item.get("id", 0) or 0) in ids:
            item["isArchived"] = True
            item["archivedAt"] = archived_at
            item["archivedBy"] = auth_user["id"]
    save_audit_logs_store(logs)
    record_audit_log(auth_user, "ARCHIVE_AUDIT_LOGS", {"archivedIds": sorted(ids), "count": len(ids)})
    return jsonify({"ok": True})


@app.route("/api/portal-settings", methods=["GET"])
def get_portal_settings():
    return jsonify({"settings": public_portal_settings()})


@app.route("/api/portal-settings/unlock", methods=["POST", "OPTIONS"])
def unlock_portal_settings():
    preflight = handle_options()
    if preflight:
        return preflight
    session_token, auth_user, error = require_owner_admin()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    password = str(data.get("password", "") or "")
    passwords = load_password_store()
    stored_password = passwords.get(str(auth_user.get("email", "") or "").strip().lower())
    if not stored_password or not verify_password(stored_password, password):
        record_audit_log(auth_user, "PORTAL_CONTROL_UNLOCK_FAILED", {"reason": "invalid_password"})
        return jsonify({"error": "Your account password is incorrect."}), 401
    try:
        challenge = issue_privileged_mfa_challenge(auth_user, purpose="portal-control")
    except Exception as exc:
        app.logger.error("Portal Control MFA delivery failed for %s: %s", auth_user.get("email"), exc)
        return jsonify({"error": "The verification code could not be delivered. Try again shortly."}), 503
    record_audit_log(auth_user, "PORTAL_CONTROL_MFA_REQUIRED", {"ok": True})
    return jsonify({"ok": True, "requiresMfa": True, **challenge})


@app.route("/api/portal-settings/unlock/verify", methods=["POST", "OPTIONS"])
def verify_portal_settings_unlock():
    preflight = handle_options()
    if preflight:
        return preflight
    session_token, auth_user, error = require_owner_admin()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    challenge_id = str(data.get("challengeId", "") or "").strip()
    code = "".join(ch for ch in str(data.get("code", "") or "") if ch.isdigit())
    limit_key = rate_limit_key("portal-control-mfa", challenge_id)
    if auth_rate_limited(limit_key):
        return jsonify({"error": "Too many verification attempts. Start the unlock again."}), 429
    challenges = load_privileged_mfa_challenges()
    challenge = challenges.get(challenge_id)
    valid = bool(
        challenge
        and str(challenge.get("userId", "")) == str(auth_user.get("id", ""))
        and str(challenge.get("purpose", "login")) == "portal-control"
        and len(code) == 6
        and hmac.compare_digest(
            str(challenge.get("codeHash", "")),
            privileged_mfa_code_hash(challenge_id, code),
        )
    )
    if not valid:
        record_auth_failure(limit_key)
        record_audit_log(auth_user, "PORTAL_CONTROL_MFA_FAILED", {"challengeId": challenge_id})
        return jsonify({"error": "Invalid or expired verification code."}), 400
    challenges.pop(challenge_id, None)
    atomic_write_json(PRIVILEGED_MFA_PATH, challenges)
    clear_auth_failures(limit_key)
    record_audit_log(auth_user, "PORTAL_CONTROL_UNLOCKED", {"mfaVerified": True})
    authorization_token, expires_at = issue_portal_authorization(auth_user, session_token)
    return jsonify({"ok": True, "authorizationToken": authorization_token, "expiresAt": expires_at})


@app.route("/api/portal-settings", methods=["POST", "OPTIONS"])
def update_portal_settings():
    preflight = handle_options()
    if preflight:
        return preflight
    session_token, auth_user, error = require_owner_admin()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    authorization_error = portal_authorization_error(data, auth_user, session_token)
    if authorization_error:
        return authorization_error
    current_settings = load_portal_settings_store()
    requested_mode = "live" if str(data.get("appMode", "test")).strip().lower() == "live" else "test"
    if requested_mode == "live":
        has_test_staff = any(bool(item.get("isTestData")) for item in load_user_store())
        has_test_customers = any(bool(item.get("isTestData")) for item in load_json_list_store(CUSTOMERS_STORE_PATH))
        if has_test_staff or has_test_customers:
            return jsonify({"error": "Remove all loaded test staff and test customers before switching to Live Mode."}), 400
    branches = normalize_portal_branches(data.get("branches"))
    departments = normalize_portal_departments(data.get("departments"))
    branch_renames = normalize_portal_rename_map(data.get("branchRenames"))
    department_renames = normalize_portal_rename_map(data.get("departmentRenames"))
    branch_renames = {
        old_value: new_value
        for old_value, new_value in branch_renames.items()
        if old_value in {str(item).strip().upper() for item in current_settings.get("branches", [])}
        and new_value in {str(item).strip().upper() for item in branches}
    }
    department_renames = {
        old_value: new_value
        for old_value, new_value in department_renames.items()
        if old_value in {str(item).strip().upper() for item in current_settings.get("departments", [])}
        and new_value in {str(item).strip().upper() for item in departments}
    }
    branch_changes = portal_list_changes(current_settings.get("branches", []), branches)
    department_changes = portal_list_changes(current_settings.get("departments", []), departments)
    dangerous_list_change = bool(
        branch_changes["removed"]
        or department_changes["removed"]
        or branch_renames
        or department_renames
    )
    if dangerous_list_change:
        backup_error = backup_confirmation_error(auth_user, data.get("backupConfirmed"))
        if backup_error:
            return backup_error
    rename_summary = apply_portal_renames(branch_renames, department_renames)
    normalized_susu_users = persist_normalized_susu_departments()
    review_status = str(data.get("securityReviewStatus") or "not-scheduled").strip().lower()
    if review_status not in {"not-scheduled", "scheduled", "in-progress", "completed", "remediation-required"}:
        return jsonify({"error": "Security review status is invalid."}), 400
    settings = {
        "bankName": str(data.get("bankName") or DEFAULT_PORTAL_SETTINGS["bankName"]).strip(),
        "shortBankName": str(data.get("shortBankName") or DEFAULT_PORTAL_SETTINGS["shortBankName"]).strip(),
        "portalName": str(data.get("portalName") or DEFAULT_PORTAL_SETTINGS["portalName"]).strip(),
        "emailDomain": normalize_email_domain(data.get("emailDomain")),
        "branches": branches,
        "departments": departments,
        "formCategories": [],
        "departmentChangeTypes": [],
        "transferLocations": [],
        "loginSubtitle": str(data.get("loginSubtitle") or DEFAULT_PORTAL_SETTINGS["loginSubtitle"]),
        "loginButtonText": str(data.get("loginButtonText") or DEFAULT_PORTAL_SETTINGS["loginButtonText"]),
        "authorizedAccessText": str(data.get("authorizedAccessText") or DEFAULT_PORTAL_SETTINGS["authorizedAccessText"]),
        "itAccessCode": str(data.get("itAccessCode") or ""),
        "hrAccessCode": str(data.get("hrAccessCode") or ""),
        "sessionMinutes": min(60, normalize_positive_number(data.get("sessionMinutes"), DEFAULT_PORTAL_SETTINGS["sessionMinutes"])),
        "absoluteSessionHours": min(24, normalize_positive_number(data.get("absoluteSessionHours"), DEFAULT_PORTAL_SETTINGS["absoluteSessionHours"])),
        "sensitiveReauthMinutes": min(30, normalize_positive_number(data.get("sensitiveReauthMinutes"), DEFAULT_PORTAL_SETTINGS["sensitiveReauthMinutes"])),
        "verificationMinutes": normalize_positive_number(data.get("verificationMinutes"), DEFAULT_PORTAL_SETTINGS["verificationMinutes"]),
        "passwordResetMinutes": normalize_positive_number(data.get("passwordResetMinutes"), DEFAULT_PORTAL_SETTINGS["passwordResetMinutes"]),
        "auditRetentionDays": max(365, min(3650, normalize_positive_number(data.get("auditRetentionDays"), DEFAULT_PORTAL_SETTINGS["auditRetentionDays"]))),
        "notificationRetentionDays": max(7, min(365, normalize_positive_number(data.get("notificationRetentionDays"), DEFAULT_PORTAL_SETTINGS["notificationRetentionDays"]))),
        "verificationRetentionHours": max(1, min(168, normalize_positive_number(data.get("verificationRetentionHours"), DEFAULT_PORTAL_SETTINGS["verificationRetentionHours"]))),
        "expiredSessionRetentionDays": max(1, min(90, normalize_positive_number(data.get("expiredSessionRetentionDays"), DEFAULT_PORTAL_SETTINGS["expiredSessionRetentionDays"]))),
        "securityReviewStatus": review_status,
        "securityReviewProvider": str(data.get("securityReviewProvider") or "").strip()[:160],
        "securityReviewReference": str(data.get("securityReviewReference") or "").strip()[:240],
        "securityReviewDate": str(data.get("securityReviewDate") or "").strip()[:10],
        "dashboardLabel": str(data.get("dashboardLabel") or DEFAULT_PORTAL_SETTINGS["dashboardLabel"]),
        "formsLabel": str(data.get("formsLabel") or DEFAULT_PORTAL_SETTINGS["formsLabel"]),
        "appMode": requested_mode,
        "publicRegistrationEnabled": bool(data.get("publicRegistrationEnabled", False)),
        "profileLabel": str(data.get("profileLabel") or DEFAULT_PORTAL_SETTINGS["profileLabel"]),
        "activeStaffLabel": str(data.get("activeStaffLabel") or DEFAULT_PORTAL_SETTINGS["activeStaffLabel"]),
        "branchCoverageLabel": str(data.get("branchCoverageLabel") or DEFAULT_PORTAL_SETTINGS["branchCoverageLabel"]),
        "openOperationsLabel": str(data.get("openOperationsLabel") or DEFAULT_PORTAL_SETTINGS["openOperationsLabel"]),
        "resolutionRateLabel": str(data.get("resolutionRateLabel") or DEFAULT_PORTAL_SETTINGS["resolutionRateLabel"]),
        "updatedAt": now_ms(),
        "updatedBy": {
            "id": auth_user["id"],
            "fullname": auth_user["fullname"],
            "email": auth_user["email"],
        },
    }
    save_portal_settings_store(settings)
    record_audit_log(
        auth_user,
        "UPDATE_PORTAL_SETTINGS",
        {
            "branches": branches,
            "departments": settings["departments"],
            "branchRenames": branch_renames,
            "departmentRenames": department_renames,
            "branchChanges": branch_changes,
            "departmentChanges": department_changes,
            "renamedRecords": rename_summary,
            "normalizedSusuUsers": normalized_susu_users,
            "emailDomain": settings["emailDomain"],
            "appMode": settings["appMode"],
            "updatedAt": settings["updatedAt"],
        },
    )
    notify_active_managers(
        kind="portal_control",
        title="Portal settings updated",
        message=f"{auth_user['fullname']} updated portal branches, departments, labels, or access settings.",
        link_to="/portal-control",
    )
    return jsonify({"ok": True, "settings": public_portal_settings(settings)})


@app.route("/api/maintenance/normalize-susu-departments", methods=["POST", "OPTIONS"])
def normalize_stored_susu_departments():
    preflight = handle_options()
    if preflight:
        return preflight
    session_token, auth_user, error = require_owner_admin()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    authorization_error = portal_authorization_error(data, auth_user, session_token)
    if authorization_error:
        return authorization_error
    backup_error = backup_confirmation_error(auth_user, data.get("backupConfirmed"))
    if backup_error:
        return backup_error

    legacy_names = {"SUSU AGENT": "SUSU", "SUSU SUPERVISOR": "SUSU"}
    normalized_users = persist_normalized_susu_departments()
    migrated_records = apply_portal_renames({}, legacy_names)
    settings = load_portal_settings_store()
    settings["departments"] = ["SUSU"]
    settings["updatedAt"] = now_ms()
    settings["updatedBy"] = {
        "id": auth_user["id"],
        "fullname": auth_user["fullname"],
        "email": auth_user["email"],
    }
    save_portal_settings_store(settings)
    record_audit_log(
        auth_user,
        "NORMALIZE_SUSU_DEPARTMENTS",
        {"normalizedUsers": normalized_users, "migratedRecords": migrated_records},
    )
    return jsonify({
        "ok": True,
        "normalizedUsers": normalized_users,
        "migratedRecords": migrated_records,
        "settings": public_portal_settings(settings),
    })


def reconciliation_snapshot() -> dict:
    customers = load_json_list_store(CUSTOMERS_STORE_PATH)
    collections = load_json_list_store(COLLECTIONS_STORE_PATH)
    closes = load_json_list_store(DAILY_CLOSES_STORE_PATH)
    completed = [
        item for item in collections
        if str(item.get("status", "completed")).strip().lower() not in {"reversed", "rejected"}
    ]
    collection_by_customer = {}
    for item in completed:
        customer_id = str(item.get("customer_id", "")).strip()
        collection_by_customer[customer_id] = round(
            collection_by_customer.get(customer_id, 0.0) + float(item.get("amount") or 0), 2
        )
    issues = []
    customer_ids = {str(item.get("id", "")).strip() for item in customers}
    for customer in customers:
        customer_id = str(customer.get("id", "")).strip()
        stored = round(float(customer.get("total_deposits") or 0), 2)
        calculated = round(collection_by_customer.get(customer_id, 0.0), 2)
        if stored != calculated:
            issues.append({
                "type": "CUSTOMER_TOTAL_MISMATCH",
                "customerId": customer_id,
                "accountNumber": str(customer.get("account_number", "")),
                "stored": stored,
                "calculated": calculated,
                "difference": round(stored - calculated, 2),
            })
    orphaned = [item for item in completed if str(item.get("customer_id", "")).strip() not in customer_ids]
    for item in orphaned:
        issues.append({
            "type": "ORPHANED_DEPOSIT",
            "collectionId": str(item.get("id", "")),
            "reference": str(item.get("transaction_reference", "")),
            "amount": round(float(item.get("amount") or 0), 2),
        })
    for close in closes:
        agent_id = str(close.get("agentId", "")).strip()
        date_key = str(close.get("date", "")).strip()
        matching = [
            item for item in completed
            if str(item.get("agent_id", "")).strip() == agent_id
            and str(item.get("transaction_date", "")).strip() == date_key
        ]
        calculated_total = round(sum(float(item.get("amount") or 0) for item in matching), 2)
        stored_total = round(float(close.get("totalAmount") or 0), 2)
        stored_count = int(close.get("transactionCount", 0) or 0)
        if stored_total != calculated_total or stored_count != len(matching):
            issues.append({
                "type": "DAILY_CLOSE_MISMATCH",
                "closeId": str(close.get("id", "")),
                "agentId": agent_id,
                "date": date_key,
                "storedTotal": stored_total,
                "calculatedTotal": calculated_total,
                "storedCount": stored_count,
                "calculatedCount": len(matching),
            })
    deposit_total = round(sum(float(item.get("amount") or 0) for item in completed), 2)
    customer_total = round(sum(float(item.get("total_deposits") or 0) for item in customers), 2)
    return {
        "ok": not issues and deposit_total == customer_total,
        "generatedAt": now_ms(),
        "summary": {
            "customers": len(customers),
            "deposits": len(completed),
            "dailyCloses": len(closes),
            "depositTotal": deposit_total,
            "customerTotal": customer_total,
            "reportTotal": deposit_total,
            "difference": round(customer_total - deposit_total, 2),
            "issueCount": len(issues),
        },
        "issues": issues,
    }


@app.route("/api/owner/reconciliation", methods=["GET"])
def owner_reconciliation():
    _, auth_user, error = require_owner_admin()
    if error:
        return error
    result = reconciliation_snapshot()
    record_audit_log(auth_user, "RUN_RECONCILIATION", {"issueCount": result["summary"]["issueCount"]})
    return jsonify(result)


def session_device_label(user_agent: str) -> str:
    value = str(user_agent or "Unknown device")
    browser = next((name for token, name in [("Edg/", "Edge"), ("Chrome/", "Chrome"), ("Firefox/", "Firefox"), ("Safari/", "Safari")] if token in value), "Browser")
    platform = next((name for token, name in [("Android", "Android"), ("iPhone", "iPhone"), ("iPad", "iPad"), ("Windows", "Windows"), ("Macintosh", "macOS"), ("Linux", "Linux")] if token in value), "Unknown device")
    return f"{browser} on {platform}"


@app.route("/api/owner/sessions", methods=["GET"])
def owner_sessions():
    current_token, _, error = require_owner_admin()
    if error:
        return error
    users = {str(item.get("id")): item for item in load_user_store()}
    current_hash = session_token_hash(current_token)
    sessions = []
    for token_hash, item in load_sessions().items():
        user = users.get(str(item.get("userId")), {})
        sessions.append({
            "id": token_hash,
            "userId": str(item.get("userId", "")),
            "userName": str(user.get("fullname", "Unknown user")),
            "email": str(user.get("email", "")),
            "role": str(user.get("role", "")),
            "device": session_device_label(item.get("userAgent", "")),
            "ipAddress": str(item.get("ipAddress", "unknown")),
            "createdAt": int(item.get("createdAt", 0) or 0) * 1000,
            "lastActivityAt": int(item.get("lastActivityAt", 0) or 0) * 1000,
            "expiresAt": int(item.get("absoluteExpiresAt", 0) or 0) * 1000,
            "isCurrent": token_hash == current_hash,
        })
    sessions.sort(key=lambda item: item["lastActivityAt"], reverse=True)
    if pagination_requested():
        return paginated_response("sessions", sessions)
    return jsonify({"sessions": sessions})


@app.route("/api/owner/sessions/<session_id>/revoke", methods=["POST", "OPTIONS"])
def owner_revoke_session(session_id: str):
    preflight = handle_options()
    if preflight:
        return preflight
    current_token, auth_user, error = require_owner_admin()
    if error:
        return error
    sessions = load_sessions()
    session = sessions.pop(str(session_id).strip(), None)
    if not session:
        return jsonify({"error": "Active session not found"}), 404
    save_sessions(sessions)
    record_audit_log(auth_user, "REVOKE_DEVICE_SESSION", {"staffId": session.get("userId"), "sessionId": str(session_id)[:12]})
    response = jsonify({"ok": True, "revokedCurrent": session_token_hash(current_token) == session_id})
    return clear_session_cookie(response) if session_token_hash(current_token) == session_id else response


@app.route("/api/backup/export", methods=["GET"])
def export_production_backup():
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    if not can_manage_agents_and_customers(auth_user):
        return jsonify({"error": "Only owner admin or an assigned supervisor can export a backup."}), 403
    owner_export = is_owner_admin(auth_user)
    all_users = load_user_store()
    visible_users = all_users if owner_export else [
        item for item in all_users if can_view_staff_record(auth_user, item)
    ]
    visible_user_ids = {str(item.get("id", "")) for item in visible_users}
    all_presence = load_presence_store()
    customers = load_json_list_store(CUSTOMERS_STORE_PATH)
    customer_imports = load_json_list_store(CUSTOMER_IMPORTS_STORE_PATH)
    collections = load_json_list_store(COLLECTIONS_STORE_PATH)
    daily_closes = load_json_list_store(DAILY_CLOSES_STORE_PATH)
    if not owner_export:
        customers = [item for item in customers if can_view_operational_record(auth_user, item)]
        customer_imports = [item for item in customer_imports if can_view_operational_record(auth_user, item)]
        collections = [item for item in collections if can_view_operational_record(auth_user, item)]
        daily_closes = [item for item in daily_closes if can_view_operational_record(auth_user, item)]
    backup = {
        "metadata": {
            "app": "bawjiase-staff-portal",
            "generatedAt": now_ms(),
            "generatedBy": {
                "id": auth_user["id"],
                "fullname": auth_user["fullname"],
                "email": auth_user["email"],
                "role": auth_user["role"],
            },
            "dataDir": DATA_DIR,
            "schemaVersion": 1,
            "scope": "all-branches" if owner_export else manageable_scope_message(auth_user),
        },
        "stores": {
            "users": visible_users,
            "passwords": load_password_store() if owner_export else {},
            "presence": {
                user_id: timestamp
                for user_id, timestamp in all_presence.items()
                if user_id in visible_user_ids
            },
            "notifications": load_json_list_store(NOTIFICATIONS_STORE_PATH) if owner_export else [],
            "auditLogs": load_audit_logs_store() if owner_export else [],
            "portalSettings": public_portal_settings(),
            "customers": customers,
            "customerImports": customer_imports,
            "collections": collections,
            "dailyCloses": daily_closes,
        },
    }
    response = jsonify(backup)
    stamp = time.strftime("%Y%m%d-%H%M%S", time.gmtime())
    response.headers["Content-Disposition"] = (
        f'attachment; filename="bawjiase-portal-backup-{stamp}.json"'
    )
    response.headers["X-Backup-Filename"] = f"bawjiase-portal-backup-{stamp}.json"
    record_audit_log(
        auth_user,
        "EXPORT_PRODUCTION_BACKUP",
        {
            "stores": list(backup["stores"].keys()),
            "generatedAt": backup["metadata"]["generatedAt"],
        },
    )
    return response


@app.route("/api/backup/import", methods=["POST", "OPTIONS"])
def import_production_backup():
    preflight = handle_options()
    if preflight:
        return preflight
    session_token, auth_user, error = require_owner_admin()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    authorization_error = portal_authorization_error(data, auth_user, session_token)
    if authorization_error:
        return authorization_error
    stores = data.get("stores") if isinstance(data.get("stores"), dict) else None
    if not stores:
        return jsonify({"error": "Backup file is missing stores data."}), 400
    metadata = data.get("metadata") if isinstance(data.get("metadata"), dict) else {}
    backup_scope = str(metadata.get("scope", "") or "").strip()
    if backup_scope and backup_scope != "all-branches":
        return jsonify({
            "error": "Only a full Owner backup can restore the complete system. Branch-scoped backups are export-only."
        }), 400

    current_backup = {
        "users": load_user_store(),
        "passwords": load_password_store(),
        "presence": load_presence_store(),
        "notifications": load_json_list_store(NOTIFICATIONS_STORE_PATH),
        "auditLogs": load_audit_logs_store(),
        "portalSettings": public_portal_settings(),
        "customers": load_json_list_store(CUSTOMERS_STORE_PATH),
        "customerImports": load_json_list_store(CUSTOMER_IMPORTS_STORE_PATH),
        "collections": load_json_list_store(COLLECTIONS_STORE_PATH),
        "dailyCloses": load_json_list_store(DAILY_CLOSES_STORE_PATH),
    }

    try:
        with DATA_LOCK:
            if "users" in stores and isinstance(stores.get("users"), list):
                imported_users = [
                    normalize_user(item)
                    for item in stores.get("users") or []
                    if isinstance(item, dict)
                ]
                if not any(is_owner_admin(item) for item in imported_users):
                    imported_users.insert(0, normalize_user(OWNER_ADMIN_USER))
                save_user_store(imported_users)
            if "passwords" in stores and isinstance(stores.get("passwords"), dict):
                save_password_store(stores.get("passwords") or {})
            if "presence" in stores and isinstance(stores.get("presence"), dict):
                save_presence_store(stores.get("presence") or {})
            if "notifications" in stores:
                save_json_list_store(NOTIFICATIONS_STORE_PATH, stores.get("notifications") or [])
            if "auditLogs" in stores and isinstance(stores.get("auditLogs"), list):
                save_audit_logs_store(
                    merge_audit_logs(current_backup["auditLogs"], stores.get("auditLogs") or [])
                )
            if "portalSettings" in stores and isinstance(stores.get("portalSettings"), dict):
                imported_settings = {**DEFAULT_PORTAL_SETTINGS, **stores.get("portalSettings")}
                imported_settings.pop("portalControlPassword", None)
                imported_settings.pop("itAccessCode", None)
                imported_settings.pop("hrAccessCode", None)
                imported_settings["departments"] = normalize_portal_departments(imported_settings.get("departments"))
                imported_settings["itAccessCode"] = ""
                imported_settings["hrAccessCode"] = ""
                save_portal_settings_store(imported_settings)
            if "customers" in stores:
                save_json_list_store(CUSTOMERS_STORE_PATH, stores.get("customers") or [])
            if "customerImports" in stores:
                save_json_list_store(CUSTOMER_IMPORTS_STORE_PATH, stores.get("customerImports") or [])
            if "collections" in stores:
                save_json_list_store(COLLECTIONS_STORE_PATH, stores.get("collections") or [])
            if "dailyCloses" in stores:
                save_json_list_store(DAILY_CLOSES_STORE_PATH, stores.get("dailyCloses") or [])
    except Exception as exc:
        save_user_store(current_backup["users"])
        save_password_store(current_backup["passwords"])
        save_presence_store(current_backup["presence"])
        save_json_list_store(NOTIFICATIONS_STORE_PATH, current_backup["notifications"])
        save_audit_logs_store(current_backup["auditLogs"])
        save_portal_settings_store(current_backup["portalSettings"])
        save_json_list_store(CUSTOMERS_STORE_PATH, current_backup["customers"])
        save_json_list_store(CUSTOMER_IMPORTS_STORE_PATH, current_backup["customerImports"])
        save_json_list_store(COLLECTIONS_STORE_PATH, current_backup["collections"])
        save_json_list_store(DAILY_CLOSES_STORE_PATH, current_backup["dailyCloses"])
        return jsonify({"error": f"Backup import failed and current data was restored: {exc}"}), 400

    record_audit_log(auth_user, "IMPORT_BACKUP", {"stores": list(stores.keys())})
    return jsonify({"ok": True, "settings": public_portal_settings()})


@app.route("/api/maintenance/clear-test-data", methods=["POST", "OPTIONS"])
def clear_test_data():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_owner_admin()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    settings = load_portal_settings_store()
    if str(settings.get("appMode", "test")).lower() != "test":
        return jsonify({"error": "Switch the portal to Test Mode before clearing test data."}), 400
    backup_error = backup_confirmation_error(auth_user, data.get("backupConfirmed"))
    if backup_error:
        return backup_error
    with DATA_LOCK:
        save_json_list_store(CUSTOMERS_STORE_PATH, [])
        save_json_list_store(COLLECTIONS_STORE_PATH, [])
        save_json_list_store(NOTIFICATIONS_STORE_PATH, [])
        save_json_list_store(DAILY_CLOSES_STORE_PATH, [])
        record_audit_log(auth_user, "CLEAR_TEST_DATA", {"cleared": ["customers", "collections", "notifications", "dailyCloses"], "auditPreserved": True})
    return jsonify({"ok": True})


@app.route("/api/maintenance/remove-test-customers", methods=["POST", "OPTIONS"])
def remove_test_customers():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_owner_admin()
    if error:
        return error
    settings = load_portal_settings_store()
    if str(settings.get("appMode", "test")).lower() != "test":
        return jsonify({"error": "Switch the portal to Test Mode before removing test customers."}), 400
    customers = load_json_list_store(CUSTOMERS_STORE_PATH)
    before_count = len(customers)
    filtered = [item for item in customers if not bool(item.get("isTestData"))]
    removed_count = before_count - len(filtered)
    if removed_count:
        save_json_list_store(CUSTOMERS_STORE_PATH, filtered)
    record_audit_log(auth_user, "REMOVE_TEST_CUSTOMERS", {"removedCount": removed_count})
    return jsonify({"ok": True, "removedCount": removed_count})


@app.route("/api/maintenance/seed-test-customers", methods=["POST", "OPTIONS"])
def seed_test_customers():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_owner_admin()
    if error:
        return error
    settings = load_portal_settings_store()
    if str(settings.get("appMode", "test")).lower() != "test":
        return jsonify({"error": "Switch the portal to Test Mode before loading test customers."}), 400

    valid_branches = {str(item or "").strip().upper() for item in settings.get("branches", []) if str(item or "").strip()}
    fallback_branch = next(iter(valid_branches), "HEAD OFFICE")
    customers = load_json_list_store(CUSTOMERS_STORE_PATH)
    existing_numbers = {str(item.get("account_number", "")).strip() for item in customers}
    created = []
    skipped = []
    timestamp = now_ms()

    for index, (name, account_number, phone, branch) in enumerate(TEST_CUSTOMER_SEED_ROWS, start=1):
        branch_name = branch if branch in valid_branches else fallback_branch
        if account_number in existing_numbers:
            skipped.append({"account_number": account_number, "reason": "already exists"})
            continue
        customer = {
            "id": f"test-cust-{timestamp}-{index:02d}",
            "account_name": name,
            "account_number": account_number,
            "phone": phone,
            "branch_id": branch_name,
            "branch_name": branch_name,
            "customer_status": "active",
            "address": "TEST DATA",
            "total_deposits": 0,
            "last_deposit_date": None,
            "createdAt": timestamp + index,
            "createdBy": auth_user["fullname"],
            "createdById": auth_user["id"],
            "createdByEmail": auth_user["email"],
            "isTestData": True,
        }
        customers.append(customer)
        created.append(customer)
        existing_numbers.add(account_number)

    if created:
        save_json_list_store(CUSTOMERS_STORE_PATH, customers)
    record_audit_log(
        auth_user,
        "SEED_TEST_CUSTOMERS",
        {"createdCount": len(created), "skippedCount": len(skipped), "testMode": settings.get("appMode", "test")},
    )
    return jsonify({"ok": True, "createdCount": len(created), "skipped": skipped, "customers": created})


@app.route("/api/maintenance/seed-test-staff", methods=["POST", "OPTIONS"])
def seed_test_staff():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_owner_admin()
    if error:
        return error
    settings = load_portal_settings_store()
    if str(settings.get("appMode", "test")).lower() != "test":
        return jsonify({"error": "Switch the portal to Test Mode before loading test staff."}), 400
    if not DEFAULT_INITIAL_PASSWORD:
        return jsonify({"error": "PORTAL_DEFAULT_INITIAL_PASSWORD is not configured on the server."}), 500

    valid_branches = {
        str(item or "").strip().upper()
        for item in settings.get("branches", [])
        if str(item or "").strip()
    }
    fallback_branch = next(iter(valid_branches), "HEAD OFFICE")
    users = load_user_store()
    passwords = load_password_store()
    existing_emails = {str(item.get("email", "")).strip().lower() for item in users}
    created = []
    skipped = []

    for seed_user in INITIAL_USERS:
        email = str(seed_user.get("email", "")).strip().lower()
        if email in existing_emails:
            skipped.append({"email": email, "reason": "already exists"})
            continue
        branch = str(seed_user.get("branch", "")).strip().upper()
        user = normalize_user({
            **seed_user,
            "branch": branch if branch in valid_branches else fallback_branch,
            "lastSeen": 0,
            "registrationTime": now_ms(),
            "isTestData": True,
        })
        users.append(user)
        existing_emails.add(email)
        if not passwords.get(email):
            passwords[email] = hash_password_for_storage(DEFAULT_INITIAL_PASSWORD)
        created.append(user)

    if created:
        save_user_store(users)
        save_password_store(passwords)
    record_audit_log(
        auth_user,
        "SEED_TEST_STAFF",
        {
            "createdCount": len(created),
            "skippedCount": len(skipped),
            "emails": [item.get("email") for item in created],
        },
    )
    return jsonify({
        "ok": True,
        "createdCount": len(created),
        "skippedCount": len(skipped),
        "users": created,
    })


@app.route("/api/maintenance/remove-test-staff", methods=["POST", "OPTIONS"])
def remove_test_staff():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_owner_admin()
    if error:
        return error
    settings = load_portal_settings_store()
    if str(settings.get("appMode", "test")).lower() != "test":
        return jsonify({"error": "Switch the portal to Test Mode before removing test staff."}), 400

    users = load_user_store()
    removed_users = [item for item in users if bool(item.get("isTestData")) and not is_owner_admin(item)]
    removed_ids = {str(item.get("id", "")) for item in removed_users}
    removed_emails = {str(item.get("email", "")).strip().lower() for item in removed_users}
    if removed_users:
        save_user_store([item for item in users if str(item.get("id", "")) not in removed_ids])
        passwords = load_password_store()
        save_password_store({email: value for email, value in passwords.items() if email not in removed_emails})
        for user_id in removed_ids:
            revoke_user_sessions(user_id)
    record_audit_log(
        auth_user,
        "REMOVE_TEST_STAFF",
        {"removedCount": len(removed_users), "emails": sorted(removed_emails)},
    )
    return jsonify({"ok": True, "removedCount": len(removed_users)})


@app.route("/api/daily-close", methods=["GET"])
def get_daily_close():
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    date_key = str(request.args.get("date") or time.strftime("%Y-%m-%d")).strip()
    agent_id = str(request.args.get("agentId") or auth_user["id"]).strip()
    if agent_id != auth_user["id"]:
        users = load_user_store()
        agent = find_user_by_id(users, agent_id)
        if not agent or not can_view_operational_record(auth_user, {"branch_name": agent.get("branch"), "agent_id": agent_id}):
            return scoped_access_denial(auth_user)
    closes = load_json_list_store(DAILY_CLOSES_STORE_PATH)
    item = next((entry for entry in closes if str(entry.get("agentId")) == agent_id and str(entry.get("date")) == date_key), None)
    return jsonify({"closed": bool(item), "close": item})


@app.route("/api/daily-close", methods=["POST", "OPTIONS"])
def close_daily_collections():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    if not is_susu_agent(auth_user):
        return jsonify({"error": "Only SUSU agent users can close a collection day."}), 403
    data, error = require_json()
    if error:
        return error
    date_key = time.strftime("%Y-%m-%d")
    counted_amount_raw = data.get("cashCountedAmount")
    counted_amount = None
    if counted_amount_raw not in (None, ""):
        try:
            counted_amount = round(float(counted_amount_raw), 2)
        except (TypeError, ValueError):
            return jsonify({"error": "Cash counted amount must be a valid number."}), 400
    with DATA_LOCK:
        collections = [
            item for item in load_json_list_store(COLLECTIONS_STORE_PATH)
            if str(item.get("agent_id")) == auth_user["id"] and str(item.get("transaction_date")) == date_key and str(item.get("status")) != "reversed"
        ]
        closes = load_json_list_store(DAILY_CLOSES_STORE_PATH)
        existing = next((entry for entry in closes if str(entry.get("agentId")) == auth_user["id"] and str(entry.get("date")) == date_key), None)
        if existing:
            return jsonify({"ok": True, "close": existing})
        total_amount = round(sum(float(item.get("amount") or 0) for item in collections), 2)
        close = {
            "id": f"close-{now_ms()}-{secrets.token_hex(6)}",
            "date": date_key,
            "agentId": auth_user["id"],
            "agentName": auth_user["fullname"],
            "branch": auth_user.get("branch", ""),
            "transactionCount": len(collections),
            "totalAmount": total_amount,
            "cashCountedAmount": counted_amount,
            "variance": round((counted_amount - total_amount), 2) if counted_amount is not None else None,
            "closedAt": now_ms(),
            "supervisorSignoffStatus": "pending",
        }
        closes.append(close)
        save_json_list_store(DAILY_CLOSES_STORE_PATH, closes)
        record_audit_log(auth_user, "CLOSE_DAILY_COLLECTIONS", close)
    notify_active_managers(
        kind="daily_close",
        title="Daily collections closed",
        message=f"{auth_user['fullname']} closed {date_key} with GHS {close['totalAmount']:,.2f}.",
        link_to="/reports",
    )
    return jsonify({"ok": True, "close": close})


@app.route("/api/daily-close/reopen", methods=["POST", "OPTIONS"])
def reopen_daily_collections():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    if auth_user.get("role") not in {"OwnerAdmin", "Supervisor"}:
        return jsonify({"error": "Only supervisors or owner admin can reopen a closed collection day."}), 403
    data, error = require_json()
    if error:
        return error
    date_key = str(data.get("date") or "").strip()
    agent_id = str(data.get("agentId") or "").strip()
    if not date_key or not agent_id:
        return jsonify({"error": "Date and agent are required to reopen a closed day."}), 400
    users = load_user_store()
    agent = find_user_by_id(users, agent_id)
    if not agent or not is_susu_agent(agent):
        return jsonify({"error": "Select a valid SUSU agent."}), 404
    if not can_view_operational_record(auth_user, {"branch_name": agent.get("branch"), "agent_id": agent_id}):
        return scoped_access_denial(auth_user)
    closes = load_json_list_store(DAILY_CLOSES_STORE_PATH)
    kept = [
        entry for entry in closes
        if not (str(entry.get("agentId")) == agent_id and str(entry.get("date")) == date_key)
    ]
    removed_count = len(closes) - len(kept)
    if removed_count:
        save_json_list_store(DAILY_CLOSES_STORE_PATH, kept)
    record_audit_log(
        auth_user,
        "REOPEN_DAILY_COLLECTIONS",
        {"date": date_key, "agentId": agent_id, "staffName": agent.get("fullname"), "removedCount": removed_count},
    )
    notify_active_managers(
        kind="daily_close",
        title="Collection day reopened",
        message=f"{auth_user['fullname']} reopened {date_key} for {agent.get('fullname')}.",
        link_to="/transactions",
    )
    return jsonify({"ok": True, "removedCount": removed_count})


@app.route("/api/customers", methods=["GET"])
def get_customers():
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    customers = load_json_list_store(CUSTOMERS_STORE_PATH)
    customers = [item for item in customers if can_view_operational_record(auth_user, item)]
    customers.sort(key=lambda item: int(item.get("createdAt", 0) or 0), reverse=True)
    if pagination_requested():
        return paginated_response("customers", customers)
    return jsonify({"customers": customers})


@app.route("/api/customers", methods=["POST", "OPTIONS"])
def create_customer():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    if not can_manage_agents_and_customers(auth_user):
        return jsonify({"error": "Only supervisors or owner admin can add customers."}), 403
    customers = load_json_list_store(CUSTOMERS_STORE_PATH)
    try:
        account_name = normalize_required_text(data.get("account_name"), "Account name")
        account_number = normalize_account_number(data.get("account_number"))
        phone = normalize_phone(data.get("phone"))
        requested_branch = normalize_portal_branch_name(
            data.get("branch_name") or data.get("branch") or auth_user.get("branch")
        )
        customer_status = normalize_customer_status(data.get("customer_status"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    if any(str(item.get("account_number", "")).strip() == account_number for item in customers):
        return jsonify({"error": "Customer account number already exists"}), 400
    try:
        branch_name = managed_branch_for_user(auth_user, requested_branch)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 403
    customer = {
        "id": f"cust-{now_ms()}-{secrets.token_hex(3)}",
        "account_name": account_name,
        "account_number": account_number,
        "phone": phone,
        "branch_id": branch_name,
        "branch_name": branch_name,
        "customer_status": customer_status,
        "address": str(data.get("address") or "").strip(),
        "total_deposits": 0,
        "last_deposit_date": None,
        "createdAt": now_ms(),
        "createdBy": auth_user["fullname"],
        "createdById": auth_user["id"],
        "createdByEmail": auth_user["email"],
    }
    customers.append(customer)
    save_json_list_store(CUSTOMERS_STORE_PATH, customers)
    record_audit_log(auth_user, "CREATE_CUSTOMER", {"customerId": customer["id"], "accountName": customer["account_name"]})
    notify_active_managers(
        kind="customer",
        title="New customer added",
        message=f"{auth_user['fullname']} added {customer['account_name']} at {customer['branch_name']}.",
        link_to="/customers",
    )
    return jsonify({"ok": True, "customer": customer})


@app.route("/api/customers/<customer_id>", methods=["POST", "OPTIONS"])
def update_customer(customer_id: str):
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    customers = load_json_list_store(CUSTOMERS_STORE_PATH)
    customer = next((item for item in customers if str(item.get("id")) == customer_id), None)
    if not customer:
        return jsonify({"error": "Customer not found"}), 404
    if not can_view_operational_record(auth_user, customer):
        return jsonify({"error": "Access denied"}), 403
    if not can_manage_agents_and_customers(auth_user):
        return jsonify({"error": "Only supervisors or owner admin can edit customers."}), 403
    before = dict(customer)
    try:
        if "account_name" in data:
            customer["account_name"] = normalize_required_text(data.get("account_name"), "Account name")
        if "account_number" in data:
            account_number = normalize_account_number(data.get("account_number"))
            duplicate = next(
                (
                    item
                    for item in customers
                    if str(item.get("id")) != customer_id
                    and str(item.get("account_number", "")).strip() == account_number
                ),
                None,
            )
            if duplicate:
                return jsonify({"error": "Another customer already uses this account number"}), 400
            customer["account_number"] = account_number
        if "phone" in data:
            customer["phone"] = normalize_phone(data.get("phone"))
        if "customer_status" in data:
            customer["customer_status"] = normalize_customer_status(data.get("customer_status"))
        if "branch_name" in data or "branch" in data:
            customer["branch_name"] = normalize_portal_branch_name(data.get("branch_name") or data.get("branch"))
        if "address" in data:
            customer["address"] = str(data.get("address") or "").strip()
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    customer["branch_id"] = customer.get("branch_name", "")
    if is_assigned_supervisor(auth_user) and not branch_allowed_for_user(auth_user, customer["branch_name"]):
        return jsonify({"error": "You can only manage customers in your assigned branch."}), 403
    customer["updatedAt"] = now_ms()
    save_json_list_store(CUSTOMERS_STORE_PATH, customers)
    if before.get("customer_status") != customer.get("customer_status"):
        record_audit_log(
            auth_user,
            "CUSTOMER_STATUS_CHANGE",
            {
                "customerId": customer["id"],
                "accountNumber": customer.get("account_number"),
                "before": before.get("customer_status"),
                "after": customer.get("customer_status"),
            },
        )
    record_audit_log(
        auth_user,
        "UPDATE_CUSTOMER",
        {
            "customerId": customer["id"],
            "accountNumber": customer.get("account_number"),
            "before": before,
            "after": customer,
        },
    )
    notify_active_managers(
        kind="customer",
        title="Customer updated",
        message=f"{auth_user['fullname']} updated {customer.get('account_name')}.",
        link_to="/customers",
    )
    return jsonify({"ok": True, "customer": customer})


def parse_customer_import_file(file_storage) -> list[dict]:
    if not file_storage or not getattr(file_storage, "filename", ""):
        raise ValueError("Choose a CSV or XLSX file.")
    filename = secure_filename(str(file_storage.filename))
    extension = os.path.splitext(filename)[1].lower()
    if extension not in {".csv", ".xlsx"}:
        raise ValueError("Only CSV and XLSX customer files are accepted.")
    max_bytes = 5 * 1024 * 1024
    content = file_storage.stream.read(max_bytes + 1)
    if len(content) > max_bytes:
        raise ValueError("Customer import files must be 5 MB or smaller.")
    if not content:
        raise ValueError("The customer import file is empty.")

    if extension == ".csv":
        try:
            decoded = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise ValueError("CSV files must use UTF-8 encoding.") from exc
        return [
            {**dict(row), "_row": index}
            for index, row in enumerate(csv.DictReader(io.StringIO(decoded)), start=2)
        ]

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header_values = next(iterator, None)
        if not header_values:
            raise ValueError("The XLSX worksheet is empty.")
        headers = [str(value or "").strip() for value in header_values]
        rows = []
        for row_number, values in enumerate(iterator, start=2):
            row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
            if any(str(value or "").strip() for value in row.values()):
                rows.append({**row, "_row": row_number})
        workbook.close()
        return rows
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError("The XLSX file could not be read safely.") from exc


def customer_import_value(row: dict, *aliases: str) -> str:
    normalized = {str(key).strip().lower(): value for key, value in row.items()}
    for alias in aliases:
        value = normalized.get(alias.lower())
        if value is None:
            continue
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        return str(value).strip()
    return ""


@app.route("/api/customers/import", methods=["POST", "OPTIONS"])
def import_customers():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    if not can_manage_agents_and_customers(auth_user):
        return jsonify({"error": "Only supervisors or owner admin can import customers."}), 403
    is_multipart = bool(request.files.get("file"))
    if is_multipart:
        data = request.form.to_dict()
        try:
            rows = parse_customer_import_file(request.files.get("file"))
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    else:
        data, error = require_json()
        if error:
            return error
        rows = data.get("customers")
    if not isinstance(rows, list) or not rows:
        return jsonify({"error": "Upload must contain at least one customer row."}), 400
    try:
        import_branch = managed_branch_for_user(auth_user, data.get("branch"))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 403
    customers = load_json_list_store(CUSTOMERS_STORE_PATH)
    existing_numbers = {str(item.get("account_number", "")).strip() for item in customers}
    valid_rows = []
    skipped = []
    for index, row in enumerate(rows, start=1):
        row_number = int(row.get("_row", index) or index) if isinstance(row, dict) else index
        if not isinstance(row, dict):
            skipped.append({"row": row_number, "reason": "Invalid row"})
            continue
        try:
            account_name = normalize_required_text(
                customer_import_value(row, "account name", "account_name", "name", "customer name"),
                "Account name",
            )
            account_number = normalize_account_number(
                customer_import_value(row, "account number", "account_number", "account no", "account no.", "account"),
            )
            branch_name = managed_branch_for_user(
                auth_user,
                customer_import_value(row, "branch", "branch name") or import_branch,
            )
        except ValueError as exc:
            skipped.append({"row": row_number, "reason": str(exc)})
            continue
        if account_number in existing_numbers:
            skipped.append({"row": row_number, "reason": "Duplicate account number"})
            continue
        valid_rows.append({
            "rowNumber": row_number,
            "account_name": account_name,
            "account_number": account_number,
            "branch": branch_name,
        })
        existing_numbers.add(account_number)

    if str(data.get("preview", "")).strip().lower() in {"1", "true", "yes"}:
        return jsonify({"ok": True, "validRows": valid_rows, "skipped": skipped})

    created = []
    for row in valid_rows:
        account_name = row["account_name"]
        account_number = row["account_number"]
        branch_name = row["branch"]
        customer = {
            "id": f"cust-{now_ms()}-{secrets.token_hex(3)}",
            "account_name": account_name,
            "account_number": account_number,
            "phone": "",
            "branch_id": branch_name,
            "branch_name": branch_name,
            "customer_status": "active",
            "address": "",
            "total_deposits": 0,
            "last_deposit_date": None,
            "createdAt": now_ms(),
            "createdBy": auth_user["fullname"],
            "createdById": auth_user["id"],
            "createdByEmail": auth_user["email"],
            "importedBy": auth_user["fullname"],
        }
        customers.append(customer)
        created.append(customer)
    if created:
        save_json_list_store(CUSTOMERS_STORE_PATH, customers)
        imports = load_json_list_store(CUSTOMER_IMPORTS_STORE_PATH)
        imports.insert(
            0,
            {
                "id": f"import-{now_ms()}-{secrets.token_hex(3)}",
                "branch": import_branch,
                "createdCount": len(created),
                "skippedCount": len(skipped),
                "createdCustomerIds": [item["id"] for item in created],
                "skipped": skipped[:50],
                "uploadedBy": auth_user["fullname"],
                "uploadedById": auth_user["id"],
                "uploadedByEmail": auth_user["email"],
                "uploadedAt": now_ms(),
            },
        )
        save_json_list_store(CUSTOMER_IMPORTS_STORE_PATH, imports)
        record_audit_log(
            auth_user,
            "IMPORT_CUSTOMERS",
            {"branch": import_branch, "created": len(created), "skipped": len(skipped)},
        )
    return jsonify({"ok": True, "created": created, "createdCount": len(created), "skipped": skipped})


@app.route("/api/customers/imports", methods=["GET"])
def get_customer_imports():
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    if not can_manage_agents_and_customers(auth_user):
        return jsonify({"error": "Only supervisors or owner admin can view customer import history."}), 403
    imports = [
        item for item in load_json_list_store(CUSTOMER_IMPORTS_STORE_PATH)
        if is_global_manager(auth_user) or branch_allowed_for_user(auth_user, item.get("branch", ""))
    ]
    imports.sort(key=lambda item: int(item.get("uploadedAt", 0) or 0), reverse=True)
    return jsonify({"imports": imports})


def persist_postgres_deposit(collection: dict, customer: dict) -> tuple[str, dict | None]:
    ensure_pg_store_table()
    with pg_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO susu_deposit_guards (
                  customer_id, transaction_date, agent_id, idempotency_key, collection_id
                ) VALUES (%s, %s::date, %s, %s, %s)
                ON CONFLICT DO NOTHING
                RETURNING collection_id
                """,
                (
                    collection["customer_id"],
                    collection["transaction_date"],
                    collection["agent_id"],
                    collection.get("idempotency_key") or None,
                    collection["id"],
                ),
            )
            inserted = cur.fetchone()
            if not inserted:
                if collection.get("idempotency_key"):
                    cur.execute(
                        """
                        SELECT collection_id FROM susu_deposit_guards
                        WHERE agent_id = %s AND idempotency_key = %s
                        """,
                        (collection["agent_id"], collection["idempotency_key"]),
                    )
                    existing = cur.fetchone()
                    if existing:
                        cur.execute("SELECT payload FROM susu_collections WHERE record_id = %s", (existing[0],))
                        existing_payload = cur.fetchone()
                        conn.rollback()
                        return "idempotent", existing_payload[0] if existing_payload else None
                cur.execute(
                    """
                    SELECT collection_id FROM susu_deposit_guards
                    WHERE customer_id = %s AND transaction_date = %s::date
                    """,
                    (collection["customer_id"], collection["transaction_date"]),
                )
                existing = cur.fetchone()
                conn.rollback()
                return "duplicate", {"id": str(existing[0])} if existing else None
            cur.execute(
                """
                INSERT INTO susu_collections (
                  record_id, customer_id, transaction_reference, account_number,
                  amount, agent_id, branch_name, transaction_date,
                  collection_status, idempotency_key, payload, updated_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s::date, %s, %s, %s::jsonb, NOW())
                ON CONFLICT (record_id) DO NOTHING
                """,
                (
                    collection["id"],
                    collection["customer_id"],
                    collection["transaction_reference"],
                    collection["account_number"],
                    collection["amount"],
                    collection["agent_id"],
                    collection["branch_name"],
                    collection["transaction_date"],
                    collection["status"],
                    collection.get("idempotency_key") or None,
                    json.dumps(collection, ensure_ascii=True),
                ),
            )
            cur.execute(
                """
                INSERT INTO susu_customers (
                  record_id, account_number, branch_name, customer_status, payload, updated_at
                ) VALUES (%s, %s, %s, %s, %s::jsonb, NOW())
                ON CONFLICT (record_id) DO UPDATE SET
                  account_number = EXCLUDED.account_number,
                  branch_name = EXCLUDED.branch_name,
                  customer_status = EXCLUDED.customer_status,
                  payload = EXCLUDED.payload,
                  updated_at = NOW()
                """,
                (
                    customer["id"],
                    customer["account_number"],
                    str(customer.get("branch_name") or customer.get("branch") or "UNKNOWN").strip().upper(),
                    str(customer.get("customer_status") or "active").strip().lower(),
                    json.dumps(customer, ensure_ascii=True),
                ),
            )
        conn.commit()
    return "created", collection


def release_deposit_guard(collection_id: str) -> None:
    if not pg_enabled() or not collection_id:
        return
    ensure_pg_store_table()
    with pg_connect() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM susu_deposit_guards WHERE collection_id = %s", (collection_id,))
        conn.commit()


@app.route("/api/collections", methods=["GET"])
def get_collections():
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    collections = load_json_list_store(COLLECTIONS_STORE_PATH)
    collections = [item for item in collections if can_view_operational_record(auth_user, item)]
    collections.sort(key=lambda item: int(item.get("created_date", 0) or 0), reverse=True)
    if pagination_requested():
        return paginated_response("collections", collections)
    return jsonify({"collections": collections})


@app.route("/api/collections", methods=["POST", "OPTIONS"])
def create_collection():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    if not is_susu_agent(auth_user):
        return jsonify({"error": "Only SUSU agent users can record deposits."}), 403
    data, error = require_json()
    if error:
        return error
    try:
        amount = float(data.get("amount") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "Amount must be a valid number"}), 400
    if amount <= 0:
        return jsonify({"error": "Amount must be greater than zero"}), 400
    if amount > 100000:
        return jsonify({"error": "Amount is above the allowed single-deposit limit."}), 400
    transaction_date = time.strftime("%Y-%m-%d")
    current_time = time.strftime("%H:%M:%S")
    customer_id = str(data.get("customer_id") or "").strip()
    idempotency_key = str(data.get("idempotency_key") or request.headers.get("Idempotency-Key") or "").strip()[:120]
    with DATA_LOCK:
        closes = load_json_list_store(DAILY_CLOSES_STORE_PATH)
        if any(str(entry.get("agentId")) == auth_user["id"] and str(entry.get("date")) == transaction_date for entry in closes):
            return jsonify({"error": "This collection day is closed. Reopen through supervisor support before recording more deposits."}), 400
        collections = load_json_list_store(COLLECTIONS_STORE_PATH)
        if idempotency_key:
            existing_idempotent = next(
                (
                    item for item in collections
                    if str(item.get("agent_id")) == auth_user["id"]
                    and str(item.get("idempotency_key", "")) == idempotency_key
                ),
                None,
            )
            if existing_idempotent:
                return jsonify({"ok": True, "collection": existing_idempotent, "idempotent": True})
        customers = load_json_list_store(CUSTOMERS_STORE_PATH)
        customer = next((item for item in customers if item.get("id") == customer_id), None)
        if not customer:
            return jsonify({"error": "Customer not found"}), 404
        if str(customer.get("customer_status", "active")).lower() != "active":
            return jsonify({"error": "Only active customers can receive deposits."}), 400
        try:
            customer_account_number = normalize_account_number(customer.get("account_number"))
        except ValueError:
            return jsonify({"error": "This customer has an invalid account number. Ask a supervisor to correct it before collecting a deposit."}), 409
        if not can_view_operational_record(auth_user, customer):
            return jsonify({"error": "You can only record deposits for customers assigned to you."}), 403
        branch_name = str((customer or {}).get("branch_name") or auth_user.get("branch") or "").strip().upper()
        if any(
            str(item.get("customer_id", "")).strip() == customer_id
            and str(item.get("transaction_date", "")).strip() == transaction_date
            and str(item.get("status", "completed")).strip().lower() == "completed"
            for item in collections
        ):
            return jsonify({"error": "This customer already has a completed deposit for today."}), 400
        collection_id = f"col-{now_ms()}-{secrets.token_hex(6)}"
        transaction_reference = f"SUSU-{time.strftime('%Y%m%d')}-{secrets.token_hex(5).upper()}"
        existing_refs = {str(item.get("transaction_reference", "")).strip() for item in collections}
        while transaction_reference in existing_refs:
            transaction_reference = f"SUSU-{time.strftime('%Y%m%d')}-{secrets.token_hex(5).upper()}"
        collection = {
            "id": collection_id,
            "customer_id": customer_id,
            "account_name": str(customer.get("account_name") or "").strip(),
            "account_number": customer_account_number,
            "amount": round(amount, 2),
            "agent_name": auth_user["fullname"],
            "agent_id": auth_user["id"],
            "agent_email": auth_user["email"],
            "branch_id": branch_name,
            "branch_name": branch_name,
            "transaction_date": transaction_date,
            "transaction_time": current_time,
            "transaction_reference": transaction_reference,
            "status": "completed",
            "supervisor_review_status": "pending",
            "idempotency_key": idempotency_key,
            "notes": str(data.get("notes") or "").strip()[:500],
            "recorded_by": auth_user["fullname"],
            "recordedById": auth_user["id"],
            "recordedByEmail": auth_user["email"],
            "created_date": now_ms(),
        }
        customer["total_deposits"] = round(float(customer.get("total_deposits") or 0) + round(amount, 2), 2)
        customer["last_deposit_date"] = collection["transaction_date"]
        if pg_enabled():
            guard_status, existing = persist_postgres_deposit(collection, customer)
        else:
            guard_status, existing = "created", None
        if guard_status == "idempotent":
            if existing:
                return jsonify({"ok": True, "collection": existing, "idempotent": True})
            return jsonify({"error": "The original deposit is still being saved. Please wait and refresh."}), 409
        if guard_status == "duplicate":
            return jsonify({"error": "This customer already has a completed deposit for today."}), 409
        if not pg_enabled():
            collections.append(collection)
            save_json_list_store(COLLECTIONS_STORE_PATH, collections)
            save_json_list_store(CUSTOMERS_STORE_PATH, customers)
        record_audit_log(auth_user, "CREATE_COLLECTION", {
            "collectionId": collection["id"],
            "amount": collection["amount"],
            "customerId": customer_id,
            "accountNumber": collection["account_number"],
            "branch": branch_name,
        })
    notify_active_managers(
        kind="collection",
        title="New deposit recorded",
        message=f"{auth_user['fullname']} recorded GHS {amount:,.2f} for {collection['account_name']}.",
        link_to="/transactions",
    )
    return jsonify({"ok": True, "collection": collection})


@app.route("/api/collections/<collection_id>/review", methods=["POST", "OPTIONS"])
def review_collection(collection_id: str):
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    if auth_user.get("role") not in {"OwnerAdmin", "Supervisor"}:
        return jsonify({"error": "Only supervisors or owner admin can review transactions."}), 403
    data, error = require_json()
    if error:
        return error
    status = str(data.get("supervisor_review_status") or "").strip().lower()
    if status not in {"pending", "approved", "queried", "rejected"}:
        return jsonify({"error": "Review status must be pending, approved, queried, or rejected."}), 400
    note = str(data.get("correction_note") or "").strip()
    if status in {"queried", "rejected"} and not note:
        return jsonify({"error": "Enter a correction reason before querying or rejecting a transaction."}), 400
    with DATA_LOCK:
        collections = load_json_list_store(COLLECTIONS_STORE_PATH)
        item = next((entry for entry in collections if str(entry.get("id")) == str(collection_id)), None)
        if not item:
            return jsonify({"error": "Transaction not found"}), 404
        if not can_view_operational_record(auth_user, item):
            return scoped_access_denial(auth_user)
        if str(item.get("status", "")).lower() == "reversed" and status != "rejected":
            return jsonify({"error": "A reversed transaction cannot be approved again. Record a replacement deposit if needed."}), 400
        before = {
            "status": item.get("status"),
            "supervisor_review_status": item.get("supervisor_review_status"),
            "correction_note": item.get("correction_note"),
            "reversal_applied": item.get("reversal_applied"),
        }
        reversal_applied = False
        if status == "rejected" and str(item.get("status", "completed")).lower() != "reversed":
            item["status"] = "reversed"
            item["reversal_applied"] = True
            item["reversed_at"] = now_ms()
            item["reversed_by"] = auth_user["fullname"]
            item["reversed_by_id"] = auth_user["id"]
            customers = load_json_list_store(CUSTOMERS_STORE_PATH)
            customer = next((entry for entry in customers if str(entry.get("id")) == str(item.get("customer_id"))), None)
            if customer:
                amount = float(item.get("amount") or 0)
                customer["total_deposits"] = round(max(0, float(customer.get("total_deposits") or 0) - amount), 2)
                completed_for_customer = [
                    entry for entry in collections
                    if str(entry.get("customer_id")) == str(customer.get("id"))
                    and str(entry.get("status", "completed")).lower() == "completed"
                ]
                completed_for_customer.sort(key=lambda entry: int(entry.get("created_date", 0) or 0), reverse=True)
                customer["last_deposit_date"] = completed_for_customer[0].get("transaction_date") if completed_for_customer else None
                save_json_list_store(CUSTOMERS_STORE_PATH, customers)
            release_deposit_guard(str(item.get("id") or ""))
            reversal_applied = True
        item["supervisor_review_status"] = status
        item["reviewed_by"] = auth_user["fullname"]
        item["reviewed_by_id"] = auth_user["id"]
        item["reviewed_at"] = now_ms()
        if status in {"queried", "rejected"}:
            item["correction_note"] = note[:500]
        elif status == "approved":
            item["correction_note"] = ""
        save_json_list_store(COLLECTIONS_STORE_PATH, collections)
        record_audit_log(
            auth_user,
            "REVIEW_COLLECTION",
            {
                "collectionId": item["id"],
                "accountName": item.get("account_name"),
                "before": before,
                "after": {
                    "status": item.get("status"),
                    "supervisor_review_status": item.get("supervisor_review_status"),
                    "correction_note": item.get("correction_note"),
                    "reversal_applied": reversal_applied,
                },
            },
        )
    if status == "queried" and item.get("agent_id"):
        users = load_user_store()
        agent_user = find_user_by_id(users, str(item.get("agent_id")))
        if agent_user:
            create_notifications_for_users(
                [agent_user],
                kind="correction",
                title="Correction requested",
                message=f"{auth_user['fullname']} requested a correction for {item.get('account_name', 'a transaction')}: {note}",
                link_to="/transactions",
            )
    return jsonify({"ok": True, "collection": item})


@app.route("/api/users", methods=["GET"])
def list_users():
    _, _, error = require_owner_admin()
    if error:
        return error
    return jsonify({"users": serialize_users_with_presence(load_user_store())})


@app.route("/api/users/<user_id>", methods=["GET"])
def get_user(user_id: str):
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    if auth_user["id"] != user_id and not is_global_manager(auth_user):
        return jsonify({"error": "Access denied"}), 403
    users = load_user_store()
    user = find_user_by_id(users, user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404
    presence = prune_presence(load_presence_store())
    save_presence_store(presence)
    return jsonify({"user": serialize_user_with_presence(user, presence)})


@app.route("/api/users/<user_id>/profile", methods=["POST", "OPTIONS"])
def update_profile(user_id: str):
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    if auth_user["id"] != user_id and not is_global_manager(auth_user):
        return jsonify({"error": "Access denied"}), 403
    data, error = require_json()
    if error:
        return error
    users = load_user_store()
    user = find_user_by_id(users, user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404

    can_manage_org_fields = is_owner_admin(auth_user)
    before_profile = {
        "fullname": user.get("fullname"),
        "phone": user.get("phone"),
        "position": user.get("position"),
        "department": user.get("department"),
        "branch": user.get("branch"),
    }
    previous_image = str(user.get("imageFile") or "").strip()
    if "fullname" in data:
        try:
            user["fullname"] = normalize_required_text(data.get("fullname"), "Full name")
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    if "phone" in data:
        try:
            user["phone"] = normalize_phone(data.get("phone")) or user["phone"]
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    if "position" in data:
        user["position"] = str(data.get("position", "")).strip() or user["position"]
    user["department"] = "SUSU"
    if "branch" in data and can_manage_org_fields:
        try:
            user["branch"] = normalize_portal_branch_name(data.get("branch"))
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    if "imageFile" in data:
        image_file = data.get("imageFile")
        user["imageFile"] = str(image_file) if image_file else None
        if previous_image.startswith("LOCAL:") and previous_image != user["imageFile"]:
            remove_uploaded_file_if_unused(previous_image.replace("LOCAL:", "", 1).strip())
    save_user_store(users)
    after_profile = {
        "fullname": user.get("fullname"),
        "phone": user.get("phone"),
        "position": user.get("position"),
        "department": user.get("department"),
        "branch": user.get("branch"),
    }
    if after_profile != before_profile:
        record_audit_log(
            auth_user,
            "UPDATE_PROFILE",
            staff_audit_target(
                user,
                {
                    "changedBySelf": auth_user["id"] == user["id"],
                    "before": before_profile,
                    "after": after_profile,
                },
            ),
        )
    current_image = str(user.get("imageFile") or "").strip()
    if "imageFile" in data and current_image != previous_image:
        if previous_image and current_image:
            action = "CHANGE_PROFILE_PHOTO"
        elif current_image:
            action = "ADD_PROFILE_PHOTO"
        else:
            action = "REMOVE_PROFILE_PHOTO"
        record_audit_log(
            auth_user,
            action,
            staff_audit_target(
                user,
                {"changedBySelf": auth_user["id"] == user["id"]},
            ),
        )
    return jsonify({"ok": True, "user": user})


@app.route("/api/staff/active", methods=["GET"])
def get_active_staff():
    _, _, error = require_authenticated_user()
    if error:
        return error
    users = load_user_store()
    active_users = [
        user for user in users
        if user["isActive"]
        and not user["isArchived"]
        and user["fullname"] not in {"MASTER ADMIN", "System Admin"}
        and user["role"] != OWNER_ADMIN_ROLE
    ]
    serialized = serialize_users_with_presence(active_users)
    if pagination_requested():
        return paginated_response("users", serialized)
    return jsonify({"users": serialized})


@app.route("/api/agents", methods=["GET"])
def get_agents():
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    if not can_manage_agents_and_customers(auth_user):
        return jsonify({"error": "Only supervisors or owner admin can view agent management."}), 403
    users = [
        user for user in load_user_store()
        if is_susu_agent(user)
        and user.get("isActive")
        and not user.get("isArchived")
        and can_view_staff_record(auth_user, user)
    ]
    users.sort(key=lambda item: str(item.get("fullname", "")).lower())
    serialized = serialize_users_with_presence(users)
    if pagination_requested():
        return paginated_response("users", serialized)
    return jsonify({"users": serialized})


def account_lifecycle_record(user: dict, limits: dict[str, list[int]], session_counts: dict[str, int]) -> dict:
    locked, failed_attempts, locked_until = user_lockout_status(user, limits)
    setup_reason = str(user.get("setupReason", "") or "").strip().lower()
    force_change = bool(user.get("forcePasswordChange", False))
    setup_complete = bool(user.get("setupComplete", True))
    if bool(user.get("isArchived", False)):
        status = "archived"
    elif locked:
        status = "locked"
    elif not bool(user.get("isActive", True)):
        status = "inactive"
    elif force_change and (setup_reason == "password-reset" or setup_complete):
        status = "password-reset-required"
    elif force_change or not setup_complete:
        status = "first-login-pending"
    else:
        status = "active"
    return {
        "id": str(user.get("id", "")),
        "fullname": str(user.get("fullname", "")),
        "email": str(user.get("email", "")),
        "username": str(user.get("loginUsername", "")),
        "phone": str(user.get("phone", "")),
        "role": str(user.get("role", "")),
        "branch": str(user.get("branch", "")),
        "status": status,
        "isTestData": bool(user.get("isTestData", False)),
        "failedAttempts": failed_attempts,
        "lockedUntil": locked_until,
        "activeSessions": session_counts.get(str(user.get("id", "")), 0),
        "lastSeen": int(user.get("lastSeen", 0) or 0),
        "registrationTime": int(user.get("registrationTime", 0) or 0),
    }


@app.route("/api/owner/accounts", methods=["GET"])
def owner_account_status():
    _, _, error = require_owner_admin()
    if error:
        return error
    limits = load_auth_rate_limits()
    session_counts = {}
    for session in load_sessions().values():
        user_id = str(session.get("userId", ""))
        session_counts[user_id] = session_counts.get(user_id, 0) + 1
    records = [account_lifecycle_record(user, limits, session_counts) for user in load_user_store()]
    summary = {}
    for record in records:
        summary[record["status"]] = summary.get(record["status"], 0) + 1
    search = str(request.args.get("search", "") or "").strip().lower()
    status_filter = str(request.args.get("status", "") or "").strip().lower()
    branch_filter = str(request.args.get("branch", "") or "").strip().upper()
    if search:
        records = [
            item for item in records
            if search in " ".join([item["fullname"], item["email"], item["username"], item["phone"]]).lower()
        ]
    if status_filter:
        records = [item for item in records if item["status"] == status_filter]
    if branch_filter:
        records = [item for item in records if item["branch"].upper() == branch_filter]
    records.sort(key=lambda item: (item["status"] != "locked", item["fullname"].lower()))
    page_items, pagination = paginate_items(records)
    return jsonify({"accounts": page_items, "pagination": pagination, "summary": summary})


@app.route("/api/staff/archived", methods=["GET"])
def get_archived_staff():
    _, auth_user, error = require_owner_admin()
    if error:
        return error
    users = load_user_store()
    return jsonify({"users": [user for user in users if user["isArchived"] and can_view_staff_record(auth_user, user)]})


@app.route("/api/staff/stats", methods=["GET"])
def get_staff_stats():
    _, _, error = require_staff_manager()
    if error:
        return error
    users = load_user_store()
    active = [user for user in users if user["isActive"] and not user["isArchived"]]
    by_department = {}
    by_branch = {}
    by_role = {}
    for user in active:
        by_department[user["department"]] = by_department.get(user["department"], 0) + 1
        by_branch[user["branch"]] = by_branch.get(user["branch"], 0) + 1
        by_role[user["role"]] = by_role.get(user["role"], 0) + 1
    return jsonify({
        "total": len(users),
        "active": len(active),
        "archived": len([user for user in users if user["isArchived"]]),
        "byDepartment": by_department,
        "byBranch": by_branch,
        "byRole": by_role,
    })


@app.route("/api/staff/<user_id>", methods=["GET"])
def get_staff_member(user_id: str):
    _, _, error = require_authenticated_user()
    if error:
        return error
    users = load_user_store()
    user = find_user_by_id(users, user_id)
    if not user:
        return jsonify({"error": "Staff member not found"}), 404
    return jsonify({"user": user})


@app.route("/api/staff/<user_id>/update", methods=["POST", "OPTIONS"])
def update_staff(user_id: str):
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    users = load_user_store()
    user = find_user_by_id(users, user_id)
    if not user:
        return jsonify({"error": "Staff member not found"}), 404
    owner_actor = is_owner_admin(auth_user)
    if not owner_actor and not (can_manage_agents_and_customers(auth_user) and is_susu_agent(user)):
        return jsonify({"error": "Only owner admin can edit this staff member."}), 403
    privileged_fields = {"role", "managedBranches", "managedDepartmentsByBranch", "permissions", "department", "isActive"}
    if not owner_actor and any(field in data for field in privileged_fields):
        return jsonify({"error": "Only owner admin can change staff roles, access scope, department, or account status."}), 403
    if not can_view_staff_record(auth_user, user):
        return scoped_access_denial(auth_user)
    previous_active = bool(user.get("isActive", False))
    previous_supervisor_access = {
        "role": str(user.get("role", "")),
        "managedBranches": normalize_scope_list(user.get("managedBranches"), empty_default=[]),
        "managedDepartmentsByBranch": normalize_managed_departments_by_branch(
            user.get("managedDepartmentsByBranch")
        ),
        "permissions": normalize_user_permissions(user.get("permissions"), str(user.get("role", ""))),
    }

    try:
        requested_department = normalize_portal_department_name(data.get("department", user["department"]))
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    before_staff = dict(user)
    if "fullname" in data:
        try:
            user["fullname"] = normalize_required_text(data.get("fullname"), "Full name")
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    if "phone" in data:
        try:
            user["phone"] = normalize_phone(data.get("phone")) or user["phone"]
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
    if "position" in data:
        user["position"] = str(data.get("position", "")).strip() or user["position"]
    if "department" in data and requested_department and owner_actor:
        user["department"] = requested_department
        if user.get("role") != "Supervisor":
            user["role"] = role_from_department(requested_department)
    if "branch" in data:
        try:
            user["branch"] = normalize_portal_branch_name(data.get("branch"))
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        if not owner_actor and not branch_allowed_for_user(auth_user, user["branch"]):
            return jsonify({"error": "You can only assign agents inside your branch scope."}), 403
    if "role" in data and owner_actor:
        requested_role = str(data.get("role", "")).strip()
        if requested_role in {"GeneralStaff", "Supervisor"}:
            user["role"] = requested_role
    if "managedBranches" in data and owner_actor:
        user["managedBranches"] = normalize_scope_list(
            data.get("managedBranches"),
            empty_default=["ALL"] if user["role"] in GLOBAL_MANAGER_ROLES else [],
        )
    if "managedDepartmentsByBranch" in data and owner_actor:
        user["managedDepartmentsByBranch"] = normalize_managed_departments_by_branch(
            data.get("managedDepartmentsByBranch")
        )
    if owner_actor and user["role"] == "Supervisor" and "managedBranches" not in data:
        supervisor_branch = str(user.get("branch") or "").strip().upper()
        if supervisor_branch:
            previous_managed_departments = normalize_managed_departments_by_branch(
                user.get("managedDepartmentsByBranch")
            )
            carried_scope = previous_managed_departments.get(supervisor_branch)
            if not carried_scope and previous_managed_departments:
                carried_scope = next(iter(previous_managed_departments.values()))
            user["managedBranches"] = [supervisor_branch]
            if "managedDepartmentsByBranch" not in data:
                user["managedDepartmentsByBranch"] = {supervisor_branch: carried_scope or ["ALL"]}
    if owner_actor and user["role"] != "Supervisor":
        user["managedBranches"] = normalize_scope_list(
            user.get("managedBranches"),
            empty_default=["ALL"] if user["role"] in GLOBAL_MANAGER_ROLES else [],
        )
        user["managedDepartmentsByBranch"] = {}
    if "permissions" in data and owner_actor:
        user["permissions"] = normalize_user_permissions(data.get("permissions"), user["role"])
    else:
        user["permissions"] = normalize_user_permissions(user.get("permissions"), user["role"])
    try:
        validate_supervisor_configuration(user)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    if "imageFile" in data:
        previous_image = str(user.get("imageFile") or "").strip()
        image_file = data.get("imageFile")
        user["imageFile"] = str(image_file) if image_file else None
        if previous_image.startswith("LOCAL:") and previous_image != user["imageFile"]:
            remove_uploaded_file_if_unused(previous_image.replace("LOCAL:", "", 1).strip())
    if "isActive" in data:
        user["isActive"] = bool(data.get("isActive"))
    save_user_store(users)
    if dict(user) != before_staff:
        record_audit_log(
            auth_user,
            "UPDATE_STAFF",
            staff_audit_target(
                user,
                {
                    "before": {
                        key: before_staff.get(key)
                        for key in ["fullname", "phone", "position", "department", "branch", "role", "isActive"]
                    },
                    "after": {
                        key: user.get(key)
                        for key in ["fullname", "phone", "position", "department", "branch", "role", "isActive"]
                    },
                },
            ),
        )
    if str(before_staff.get("role", "")) != str(user.get("role", "")):
        record_audit_log(
            auth_user,
            "CHANGE_STAFF_ROLE",
            staff_audit_target(user, {"before": before_staff.get("role"), "after": user.get("role")}),
        )
    if str(before_staff.get("branch", "")).strip().upper() != str(user.get("branch", "")).strip().upper():
        record_audit_log(
            auth_user,
            "TRANSFER_STAFF_BRANCH",
            staff_audit_target(user, {"before": before_staff.get("branch"), "after": user.get("branch")}),
        )
    if "isActive" in data and bool(user.get("isActive", False)) != previous_active:
        record_audit_log(
            auth_user,
            "ACTIVATE_STAFF" if bool(user.get("isActive", False)) else "DEACTIVATE_STAFF",
            staff_audit_target(
                user,
                {
                    "before": {"isActive": previous_active},
                    "after": {"isActive": bool(user.get("isActive", False))},
                },
            ),
        )
    current_supervisor_access = {
        "role": str(user.get("role", "")),
        "managedBranches": normalize_scope_list(user.get("managedBranches"), empty_default=[]),
        "managedDepartmentsByBranch": normalize_managed_departments_by_branch(
            user.get("managedDepartmentsByBranch")
        ),
        "permissions": normalize_user_permissions(user.get("permissions"), str(user.get("role", ""))),
    }
    if current_supervisor_access != previous_supervisor_access:
        record_audit_log(
            auth_user,
            "SUPERVISOR_ACCESS_UPDATE",
            {
                "staffId": user["id"],
                "staffName": user["fullname"],
                "before": previous_supervisor_access,
                "after": current_supervisor_access,
            },
        )
    return jsonify({"ok": True, "user": user})


@app.route("/api/staff/<user_id>/archive", methods=["POST", "OPTIONS"])
def archive_staff(user_id: str):
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_owner_admin()
    if error:
        return error
    if user_id == auth_user["id"]:
        return jsonify({"error": "You cannot remove your own account"}), 400
    users = load_user_store()
    user = find_user_by_id(users, user_id)
    if not user:
        return jsonify({"error": "Staff member not found"}), 404
    if not can_view_staff_record(auth_user, user):
        return scoped_access_denial(auth_user)
    if user["role"] in {"OwnerAdmin", "SuperAdmin"}:
        return jsonify({"error": "Cannot archive Owner or Super Admin."}), 400
    user["isArchived"] = True
    user["isActive"] = False
    save_user_store(users)
    revoke_user_sessions(user_id)
    record_audit_log(auth_user, "ARCHIVE_STAFF", staff_audit_target(user))
    notify_active_managers(
        kind="staff",
        title="Staff archived",
        message=f"{auth_user['fullname']} archived {user['fullname']}.",
        link_to="/past-staff",
    )
    return jsonify({"ok": True})


@app.route("/api/staff/<user_id>/restore", methods=["POST", "OPTIONS"])
def restore_staff(user_id: str):
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_owner_admin()
    if error:
        return error
    users = load_user_store()
    user = find_user_by_id(users, user_id)
    if not user:
        return jsonify({"error": "Staff member not found"}), 404
    if not can_view_staff_record(auth_user, user):
        return scoped_access_denial(auth_user)
    user["isArchived"] = False
    user["isActive"] = True
    save_user_store(users)
    record_audit_log(auth_user, "RESTORE_STAFF", staff_audit_target(user))
    notify_active_managers(
        kind="staff",
        title="Staff restored",
        message=f"{auth_user['fullname']} restored {user['fullname']} to the active directory.",
        link_to="/directory",
    )
    return jsonify({"ok": True})


@app.route("/api/staff/<user_id>/delete", methods=["POST", "OPTIONS"])
def delete_staff(user_id: str):
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    if not can_manage_agents_and_customers(auth_user):
        return jsonify({"error": "Only supervisors or owner admin can remove agents."}), 403
    data, error = require_json()
    if error:
        return error
    backup_error = backup_confirmation_error(auth_user, data.get("backupConfirmed"))
    if backup_error:
        return backup_error
    users = load_user_store()
    user = find_user_by_id(users, user_id)
    if not user:
        return jsonify({"error": "Staff member not found"}), 404
    if user_id == auth_user["id"]:
        return jsonify({"error": "You cannot remove your own account"}), 400
    if not can_view_staff_record(auth_user, user):
        return scoped_access_denial(auth_user)
    if user["role"] in {"OwnerAdmin", "SuperAdmin"}:
        return jsonify({"error": "Cannot permanently remove Owner or Super Admin."}), 400
    if not is_global_manager(auth_user) and not is_susu_agent(user):
        return jsonify({"error": "Supervisors can only remove SUSU agents."}), 403
    removed_snapshot = staff_audit_target(user)
    pending = load_pending_verifications()
    pending.pop(user["email"], None)
    passwords = load_password_store()
    passwords.pop(str(user.get("email", "")).strip().lower(), None)
    username = str(user.get("loginUsername", "")).strip().lower()
    if username:
        passwords.pop(f"username:{username}", None)
    presence = load_presence_store()
    presence.pop(user_id, None)
    save_user_store([item for item in users if item.get("id") != user_id])
    save_pending_verifications(pending)
    save_password_store(passwords)
    save_presence_store(presence)
    revoke_user_sessions(user_id)
    record_audit_log(auth_user, "DELETE_STAFF_ACCOUNT", removed_snapshot)
    notify_active_managers(
        kind="staff",
        title="Staff archived",
        message=f"{auth_user['fullname']} removed {user['fullname']}'s login account.",
        link_to="/past-staff",
    )
    return jsonify({"ok": True})


@app.route("/api/agents/create", methods=["POST", "OPTIONS"])
def create_agent_account():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    if not can_manage_agents_and_customers(auth_user):
        return jsonify({"error": "Only supervisors or owner admin can add agents."}), 403
    data, error = require_json()
    if error:
        return error
    try:
        username = normalize_agent_username(data.get("username"))
        temp_password = str(data.get("temporaryPassword") or "").strip()
        phone = normalize_phone(data.get("phone"))
        branch = managed_branch_for_user(auth_user, data.get("branch"))
        fullname = str(data.get("fullname") or username).strip()
        if len(temp_password) < 6:
            return jsonify({"error": "Temporary password must be at least 6 characters."}), 400
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    users = load_user_store()
    if find_user_by_username(users, username):
        return jsonify({"error": "Username already exists."}), 400
    synthetic_email = f"{username}@agents.local"
    if find_user_by_email(users, synthetic_email):
        return jsonify({"error": "Agent already exists."}), 400
    user = normalize_user({
        "id": f"agent-{now_ms()}-{secrets.token_hex(3)}",
        "fullname": fullname,
        "phone": phone,
        "email": synthetic_email,
        "role": "GeneralStaff",
        "position": "SUSU Agent",
        "department": "SUSU",
        "branch": branch,
        "imageFile": None,
        "isActive": True,
        "isVerified": True,
        "lastSeen": 0,
        "registrationTime": now_ms(),
        "isArchived": False,
        "loginUsername": username,
        "createdBySupervisorId": auth_user["id"],
        "createdBySupervisorName": auth_user["fullname"],
        "forcePasswordChange": True,
        "setupComplete": False,
        "setupReason": "first-login",
    })
    users.append(user)
    passwords = load_password_store()
    passwords[agent_password_key(username)] = hash_password_for_storage(temp_password)
    save_user_store(users)
    save_password_store(passwords)
    record_audit_log(auth_user, "CREATE_AGENT_ACCOUNT", staff_audit_target(user, {"username": username}))
    return jsonify({"ok": True, "user": user})


@app.route("/api/agents/<user_id>/reset-password", methods=["POST", "OPTIONS"])
def reset_agent_password(user_id: str):
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    if not can_manage_agents_and_customers(auth_user):
        return jsonify({"error": "Only supervisors or owner admin can reset agent passwords."}), 403
    data, error = require_json()
    if error:
        return error
    requested_username = str(data.get("temporaryUsername") or "").strip()
    try:
        next_username = normalize_agent_username(requested_username) if requested_username else ""
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    temp_password = str(data.get("temporaryPassword") or "").strip()
    if len(temp_password) < 6:
        return jsonify({"error": "Temporary password must be at least 6 characters."}), 400
    users = load_user_store()
    user = find_user_by_id(users, user_id)
    if not user or not is_susu_agent(user):
        return jsonify({"error": "Agent not found."}), 404
    if not can_view_staff_record(auth_user, user):
        return scoped_access_denial(auth_user)
    username = str(user.get("loginUsername") or "").strip().lower()
    if not username:
        return jsonify({"error": "This agent does not have a username login."}), 400
    if next_username and next_username != username:
        existing = find_user_by_username_safe(users, next_username)
        if existing and existing.get("id") != user.get("id"):
            return jsonify({"error": "That username is already assigned to another agent."}), 400
    user["forcePasswordChange"] = True
    user["setupComplete"] = False
    user["setupReason"] = "password-reset"
    if next_username:
        user["loginUsername"] = next_username
    passwords = load_password_store()
    if next_username and next_username != username:
        passwords.pop(agent_password_key(username), None)
    active_username = next_username or username
    passwords[agent_password_key(active_username)] = hash_password_for_storage(temp_password)
    save_user_store(users)
    save_password_store(passwords)
    revoke_user_sessions(user_id)
    record_audit_log(
        auth_user,
        "RESET_AGENT_LOGIN",
        staff_audit_target(user, {"previousUsername": username, "temporaryUsername": active_username}),
    )
    return jsonify({"ok": True, "user": user})


@app.route("/api/supervisors/create", methods=["POST", "OPTIONS"])
def create_supervisor_account():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_owner_admin()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    try:
        fullname = str(data.get("fullname") or "").strip()
        email = validate_email(str(data.get("email") or ""))
        phone = normalize_phone(data.get("phone"))
        branch = managed_branch_for_user(auth_user, data.get("branch"))
        temporary_password = str(data.get("temporaryPassword") or "")
        if len(fullname) < 2:
            raise ValueError("Enter the supervisor's full name.")
        if len(temporary_password) < 8:
            raise ValueError("Temporary password must be at least 8 characters.")
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    users = load_user_store()
    if find_user_by_email(users, email):
        return jsonify({"error": "That official email already belongs to a staff account."}), 400

    supervisor = normalize_user({
        "id": f"supervisor-{now_ms()}-{secrets.token_hex(3)}",
        "fullname": fullname,
        "phone": phone,
        "email": email,
        "role": "Supervisor",
        "position": "SUSU Supervisor",
        "department": "SUSU",
        "branch": branch,
        "managedBranches": [branch],
        "managedDepartmentsByBranch": {},
        "permissions": {
            "customers": True,
            "transactions": True,
            "reports": True,
            "agents": True,
            "branches": True,
            "auditLog": False,
            "backupExport": True,
            "userManagement": False,
        },
        "imageFile": None,
        "isActive": True,
        "isVerified": True,
        "lastSeen": 0,
        "registrationTime": now_ms(),
        "isArchived": False,
    })
    users.append(supervisor)
    passwords = load_password_store()
    passwords[email] = hash_password_for_storage(temporary_password)
    save_user_store(users)
    save_password_store(passwords)
    record_audit_log(auth_user, "CREATE_SUPERVISOR_ACCOUNT", staff_audit_target(supervisor))
    return jsonify({"ok": True, "user": supervisor})


@app.route("/api/staff/<user_id>/reset-email-login", methods=["POST", "OPTIONS"])
def reset_staff_email_login(user_id: str):
    preflight = handle_options()
    if preflight:
        return preflight
    session_token, auth_user, error = require_owner_admin()
    if error:
        return error
    reauth_error = recent_reauthentication_error(session_token)
    if reauth_error:
        return reauth_error
    data, error = require_json()
    if error:
        return error
    new_password = str(data.get("newPassword") or "")
    if len(new_password) < 8:
        return jsonify({"error": "New password must be at least 8 characters."}), 400
    users = load_user_store()
    user = find_user_by_id(users, user_id)
    if not user or user.get("isArchived"):
        return jsonify({"error": "Active staff member not found."}), 404
    if is_owner_admin(user):
        return jsonify({"error": "Use Forgot Password to reset the owner account."}), 400
    email = str(user.get("email") or "").strip().lower()
    if not email or email.endswith("@agents.local"):
        return jsonify({"error": "This account uses Agent username login. Reset it from Agent Management."}), 400
    passwords = load_password_store()
    passwords[email] = hash_password_for_storage(new_password)
    save_password_store(passwords)
    revoke_user_sessions(user["id"])
    record_audit_log(
        auth_user,
        "RESET_STAFF_EMAIL_LOGIN",
        staff_audit_target(user, {"email": email}),
    )
    return jsonify({"ok": True, "user": user})


@app.route("/api/auth/register", methods=["POST", "OPTIONS"])
def auth_register():
    preflight = handle_options()
    if preflight:
        return preflight
    data, error = require_json()
    if error:
        return error
    try:
        settings = load_portal_settings_store()
        if not settings.get("publicRegistrationEnabled", False):
            return jsonify({"error": "Public staff sign-up is currently disabled. Ask a supervisor or owner admin to add your account."}), 403
        email = validate_email(str(data.get("email", "")))
        password = str(data.get("passwordHash", ""))
        requested_department = str(data.get("department", "") or "").strip().upper()
        if requested_department in {"SUSU AGENT", "SUSU SUPERVISOR"}:
            return jsonify({"error": "SUSU agent and supervisor accounts must be created by a supervisor or owner admin."}), 403
        department = "SUSU"
        branch = normalize_portal_branch_name(data.get("branch"))
        if not password:
            return jsonify({"error": "Password is required"}), 400
        if len(password) < 8:
            return jsonify({"error": "Password must be at least 8 characters"}), 400
        if not branch:
            return jsonify({"error": "Branch is required"}), 400
        users = load_user_store()
        existing = find_user_by_email(users, email)
        if existing and existing["isVerified"]:
            return jsonify({"error": "Email already registered"}), 400

        pending = load_pending_verifications()
        new_user = normalize_user({
            "id": existing["id"] if existing else f"user-{int(time.time() * 1000)}",
            "fullname": str(data.get("fullname", "")).strip(),
            "phone": str(data.get("phone", "")).strip(),
            "email": email,
            "role": role_from_department(department),
            "position": str(data.get("position", "Staff")).strip() or "Staff",
            "department": department,
            "branch": branch,
            "imageFile": None,
            "isActive": True,
            "isVerified": False,
            "lastSeen": now_ms(),
            "registrationTime": now_ms(),
            "isArchived": False,
        })
        code = generate_verification_code()
        pending[email] = {
            "user": new_user,
            "passwordHash": hash_password_for_storage(password),
            "code": code,
            "expiresAt": int(time.time()) + int(load_portal_settings_store()["verificationMinutes"]) * 60,
        }
        save_pending_verifications(pending)
        send_verification_code_email(email, code)
        record_audit_log(None, "REGISTRATION_STARTED", staff_audit_target(new_user))
        return jsonify({"ok": True, "user": new_user})
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        app.logger.exception("Registration failed")
        return jsonify({"error": f"Registration failed: {exc}"}), 500


@app.route("/api/auth/verify-email", methods=["POST", "OPTIONS"])
def auth_verify_email():
    preflight = handle_options()
    if preflight:
        return preflight
    data, error = require_json()
    if error:
        return error
    try:
        email = validate_email(str(data.get("email", "")))
        code = "".join(ch for ch in str(data.get("code", "")) if ch.isdigit())
        if len(code) != 6:
            return jsonify({"error": "A 6-digit verification code is required"}), 400

        pending = load_pending_verifications()
        entry = pending.get(email)
        if not entry:
            record_audit_log(
                None,
                "VERIFY_EMAIL_FAILED",
                {"email": email, "reason": "no_pending_verification"},
            )
            return jsonify({"error": "No pending verification for this email"}), 404
        if entry["code"] != code:
            record_audit_log(
                None,
                "VERIFY_EMAIL_FAILED",
                {"email": email, "reason": "incorrect_code"},
            )
            return jsonify({"error": "Incorrect verification code"}), 400

        user = entry["user"]
        user["isVerified"] = True

        users = load_user_store()
        existing = find_user_by_email(users, email)
        if existing:
            existing.update(user)
        else:
            users.append(user)

        passwords = load_password_store()
        passwords[email] = entry["passwordHash"]

        pending.pop(email, None)
        save_user_store(users)
        save_password_store(passwords)
        save_pending_verifications(pending)
        record_audit_log(user, "REGISTRATION_VERIFIED", staff_audit_target(user))
        return jsonify({"ok": True, "user": user})
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/auth/resend-verification", methods=["POST", "OPTIONS"])
def auth_resend_verification():
    preflight = handle_options()
    if preflight:
        return preflight
    data, error = require_json()
    if error:
        return error
    try:
        email = validate_email(str(data.get("email", "")))
        pending = load_pending_verifications()
        entry = pending.get(email)
        if not entry:
            return jsonify({"error": "Email not found"}), 404
        entry["code"] = generate_verification_code()
        entry["expiresAt"] = int(time.time()) + int(load_portal_settings_store()["verificationMinutes"]) * 60
        pending[email] = entry
        save_pending_verifications(pending)
        send_verification_code_email(email, entry["code"])
        return jsonify({"ok": True})
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        app.logger.exception("Verification email failed")
        return jsonify({"error": f"Email could not be sent: {exc}"}), 500


@app.route("/api/auth/login", methods=["POST", "OPTIONS"])
def auth_login():
    preflight = handle_options()
    if preflight:
        return preflight
    data, error = require_json()
    if error:
        return error
    try:
        email = validate_email(str(data.get("email", "")))
        password = str(data.get("passwordHash", ""))
        if not password:
            return jsonify({"error": "Password is required"}), 400
        limit_key = rate_limit_key("staff-login", email)
        if auth_rate_limited(limit_key):
            emit_production_alert(
                "LOGIN_RATE_LIMITED",
                "A staff login was blocked after repeated failed attempts.",
                severity="warning",
                context={"reason": "rate_limited"},
                throttle_key="login-rate-limited",
                throttle_seconds=900,
            )
            return jsonify({"error": "Too many login attempts. Please wait 15 minutes and try again."}), 429

        passwords = load_password_store()
        stored_password = passwords.get(email)
        if not stored_password or not verify_password(stored_password, password):
            record_auth_failure(limit_key)
            record_audit_log(
                None,
                "LOGIN_FAILED",
                {"email": email, "reason": "invalid_credentials"},
            )
            emit_production_alert(
                "LOGIN_FAILED",
                "A staff login failed because the credentials were invalid.",
                severity="warning",
                context={"reason": "invalid_credentials"},
                throttle_key="login-failed",
                throttle_seconds=300,
            )
            return jsonify({"error": "Invalid email or password"}), 401

        users = load_user_store()
        user = find_user_by_email(users, email)
        if not user or user["isArchived"] or not user["isActive"]:
            record_auth_failure(limit_key)
            record_audit_log(
                None,
                "LOGIN_FAILED",
                {"email": email, "reason": "inactive_or_missing_account"},
            )
            emit_production_alert(
                "LOGIN_FAILED",
                "A staff login failed for an inactive or unavailable account.",
                severity="warning",
                context={"reason": "inactive_or_missing_account"},
                throttle_key="login-failed",
                throttle_seconds=300,
            )
            return jsonify({"error": "Invalid email or password"}), 401
        if not user["isVerified"]:
            record_auth_failure(limit_key)
            record_audit_log(
                None,
                "LOGIN_FAILED",
                {"email": email, "reason": "email_not_verified"},
            )
            return jsonify({"error": "Email not verified"}), 403

        if not is_secure_password_hash(stored_password):
            passwords[email] = hash_password_for_storage(password)
            save_password_store(passwords)

        if privileged_mfa_required(user):
            clear_auth_failures(limit_key)
            if trusted_device_is_valid(user["id"]):
                user["lastSeen"] = now_ms()
                save_user_store(users)
                session_token = issue_session(user["id"])
                record_audit_log(user, "LOGIN_TRUSTED_DEVICE", staff_audit_target(user))
                return authenticated_response(user, session_token)
            try:
                challenge = issue_privileged_mfa_challenge(user)
            except Exception as exc:
                app.logger.error("Privileged MFA delivery failed for %s: %s", email, exc)
                record_audit_log(user, "PRIVILEGED_MFA_DELIVERY_FAILED", {"email": email})
                return jsonify({"error": "Verification code could not be delivered. Contact the system owner."}), 503
            record_audit_log(user, "PRIVILEGED_MFA_REQUIRED", {"email": email})
            return jsonify({"ok": True, "requiresMfa": True, **challenge})

        user["lastSeen"] = now_ms()
        save_user_store(users)
        session_token = issue_session(user["id"])
        clear_auth_failures(limit_key)
        record_audit_log(user, "LOGIN", staff_audit_target(user))
        return authenticated_response(user, session_token)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/auth/privileged-mfa/verify", methods=["POST", "OPTIONS"])
def auth_privileged_mfa_verify():
    preflight = handle_options()
    if preflight:
        return preflight
    data, error = require_json()
    if error:
        return error
    challenge_id = str(data.get("challengeId", "") or "").strip()
    code = "".join(ch for ch in str(data.get("code", "") or "") if ch.isdigit())
    trust_device = data.get("trustDevice") is True
    limit_key = rate_limit_key("privileged-mfa", challenge_id)
    if auth_rate_limited(limit_key):
        return jsonify({"error": "Too many verification attempts. Sign in again to request a new code."}), 429
    challenges = load_privileged_mfa_challenges()
    challenge = challenges.get(challenge_id)
    if not challenge or str(challenge.get("purpose", "login")) != "login" or len(code) != 6 or not hmac.compare_digest(
        str(challenge.get("codeHash", "")), privileged_mfa_code_hash(challenge_id, code)
    ):
        record_auth_failure(limit_key)
        if challenge:
            challenge["attempts"] = int(challenge.get("attempts", 0) or 0) + 1
            challenges[challenge_id] = challenge
            atomic_write_json(PRIVILEGED_MFA_PATH, challenges)
        record_audit_log(None, "PRIVILEGED_MFA_FAILED", {"challengeId": challenge_id})
        return jsonify({"error": "Invalid or expired verification code."}), 400
    users = load_user_store()
    user = find_user_by_id(users, challenge["userId"])
    if not user or not privileged_mfa_required(user) or user["isArchived"] or not user["isActive"] or not user["isVerified"]:
        return jsonify({"error": "This privileged account is not available."}), 403
    challenges.pop(challenge_id, None)
    atomic_write_json(PRIVILEGED_MFA_PATH, challenges)
    clear_auth_failures(limit_key)
    user["lastSeen"] = now_ms()
    save_user_store(users)
    session_token = issue_session(user["id"])
    record_audit_log(
        user,
        "LOGIN_MFA_VERIFIED",
        staff_audit_target(user, {"trustedDevice": trust_device}),
    )
    response = authenticated_response(user, session_token)
    if trust_device:
        set_trusted_device_cookie(response, issue_trusted_device(user["id"]))
    return response


@app.route("/api/auth/agent-login", methods=["POST", "OPTIONS"])
def auth_agent_login():
    preflight = handle_options()
    if preflight:
        return preflight
    data, error = require_json()
    if error:
        return error
    try:
        username = normalize_agent_username(data.get("username"))
        password = str(data.get("passwordHash", ""))
        if not password:
            return jsonify({"error": "Password is required"}), 400
        limit_key = rate_limit_key("agent-login", username)
        if auth_rate_limited(limit_key):
            return jsonify({"error": "Too many login attempts. Please wait 15 minutes and try again."}), 429
        users = load_user_store()
        user = find_user_by_username(users, username)
        passwords = load_password_store()
        stored_password = passwords.get(agent_password_key(username))
        if not user or not stored_password or not verify_password(stored_password, password):
            record_auth_failure(limit_key)
            record_audit_log(None, "AGENT_LOGIN_FAILED", {"username": username, "reason": "invalid_credentials"})
            return jsonify({"error": "Invalid username or password"}), 401
        if user["isArchived"] or not user["isActive"]:
            record_auth_failure(limit_key)
            record_audit_log(None, "AGENT_LOGIN_FAILED", {"username": username, "reason": "inactive_or_missing_account"})
            return jsonify({"error": "Invalid username or password"}), 401
        if bool(user.get("forcePasswordChange", False)) or not bool(user.get("setupComplete", True)):
            return jsonify({
                "ok": True,
                "requiresSetup": True,
                "username": username,
                "message": "First login requires phone verification and password reset.",
            })
        user["lastSeen"] = now_ms()
        save_user_store(users)
        session_token = issue_session(user["id"])
        clear_auth_failures(limit_key)
        record_audit_log(user, "AGENT_LOGIN", staff_audit_target(user, {"username": username}))
        return authenticated_response(user, session_token)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/auth/agent-verify-phone", methods=["POST", "OPTIONS"])
def auth_agent_verify_phone():
    preflight = handle_options()
    if preflight:
        return preflight
    data, error = require_json()
    if error:
        return error
    try:
        username = normalize_agent_username(data.get("username"))
        temp_password = str(data.get("temporaryPassword") or "")
        phone = normalize_phone(data.get("phone"))
        limit_key = rate_limit_key("agent-setup-phone", username)
        if auth_rate_limited(limit_key):
            return jsonify({"error": "Too many setup attempts. Please wait 15 minutes and try again."}), 429
        users = load_user_store()
        user = find_user_by_username(users, username)
        passwords = load_password_store()
        stored_password = passwords.get(agent_password_key(username))
        if not user or not stored_password or not verify_password(stored_password, temp_password):
            record_auth_failure(limit_key)
            return jsonify({"error": "Invalid username or temporary password."}), 401
        if user["isArchived"] or not user["isActive"]:
            record_auth_failure(limit_key)
            return jsonify({"error": "Invalid username or password"}), 401
        if "".join(ch for ch in str(user.get("phone") or "") if ch.isdigit()) != "".join(ch for ch in phone if ch.isdigit()):
            record_auth_failure(limit_key)
            return jsonify({"error": "Phone number does not match the supervisor record."}), 400
        clear_auth_failures(limit_key)
        setup_token = issue_agent_setup_token(username, phone)
        app_mode = str(load_portal_settings_store().get("appMode", "test")).lower()
        response = {"ok": True, "message": "Verification token sent."}
        if app_mode == "test":
            response["testToken"] = setup_token["code"]
            response["message"] = "Test verification token generated."
        elif not send_sms_token(phone, setup_token["code"]):
            return jsonify({"error": "SMS delivery is not configured. Contact the supervisor or owner admin."}), 500
        return jsonify(response)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/auth/agent-complete-setup", methods=["POST", "OPTIONS"])
def auth_agent_complete_setup():
    preflight = handle_options()
    if preflight:
        return preflight
    data, error = require_json()
    if error:
        return error
    try:
        username = normalize_agent_username(data.get("username"))
        new_username = normalize_agent_username(data.get("newUsername") or username)
        temp_password = str(data.get("temporaryPassword") or "")
        new_password = str(data.get("newPasswordHash") or "")
        phone = normalize_phone(data.get("phone"))
        token_code = "".join(ch for ch in str(data.get("token") or "") if ch.isdigit())
        limit_key = rate_limit_key("agent-setup-complete", username)
        if auth_rate_limited(limit_key):
            return jsonify({"error": "Too many setup attempts. Please wait 15 minutes and try again."}), 429
        if len(new_password) < 8:
            return jsonify({"error": "New password must be at least 8 characters."}), 400
        setup_tokens = load_agent_setup_tokens()
        setup_entry = setup_tokens.get(username)
        if not setup_entry or setup_entry.get("code") != token_code:
            record_auth_failure(limit_key)
            return jsonify({"error": "Invalid or expired verification token."}), 400
        users = load_user_store()
        user = find_user_by_username(users, username)
        passwords = load_password_store()
        stored_password = passwords.get(agent_password_key(username))
        if not user or not stored_password or not verify_password(stored_password, temp_password):
            record_auth_failure(limit_key)
            return jsonify({"error": "Invalid username or temporary password."}), 401
        existing_username = find_user_by_username_safe(users, new_username)
        if existing_username and existing_username.get("id") != user.get("id"):
            return jsonify({"error": "That permanent username is already used by another agent."}), 400
        if "".join(ch for ch in str(user.get("phone") or "") if ch.isdigit()) != "".join(ch for ch in phone if ch.isdigit()):
            record_auth_failure(limit_key)
            return jsonify({"error": "Phone number does not match the supervisor record."}), 400
        if normalize_phone(setup_entry.get("phone")) and "".join(ch for ch in normalize_phone(setup_entry.get("phone")) if ch.isdigit()) != "".join(ch for ch in phone if ch.isdigit()):
            record_auth_failure(limit_key)
            return jsonify({"error": "Verification token does not match this phone number."}), 400
        if new_username != username:
            passwords.pop(agent_password_key(username), None)
            user["loginUsername"] = new_username
        user["forcePasswordChange"] = False
        user["setupComplete"] = True
        user["setupReason"] = ""
        user["lastSeen"] = now_ms()
        passwords[agent_password_key(new_username)] = hash_password_for_storage(new_password)
        save_user_store(users)
        save_password_store(passwords)
        setup_tokens.pop(username, None)
        save_agent_setup_tokens(setup_tokens)
        session_token = issue_session(user["id"])
        clear_auth_failures(limit_key)
        record_audit_log(
            user,
            "AGENT_SETUP_COMPLETED",
            staff_audit_target(user, {"temporaryUsername": username, "permanentUsername": new_username}),
        )
        return authenticated_response(user, session_token)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/auth/agent-verify-token", methods=["POST", "OPTIONS"])
def auth_agent_verify_token():
    preflight = handle_options()
    if preflight:
        return preflight
    data, error = require_json()
    if error:
        return error
    try:
        username = normalize_agent_username(data.get("username"))
        phone = normalize_phone(data.get("phone"))
        token_code = "".join(ch for ch in str(data.get("token") or "") if ch.isdigit())
        limit_key = rate_limit_key("agent-setup-token", username)
        if auth_rate_limited(limit_key):
            return jsonify({"error": "Too many setup attempts. Please wait 15 minutes and try again."}), 429
        setup_entry = load_agent_setup_tokens().get(username)
        if not setup_entry or setup_entry.get("code") != token_code:
            record_auth_failure(limit_key)
            return jsonify({"error": "Invalid or expired verification token."}), 400
        if normalize_phone(setup_entry.get("phone")) and "".join(ch for ch in normalize_phone(setup_entry.get("phone")) if ch.isdigit()) != "".join(ch for ch in phone if ch.isdigit()):
            record_auth_failure(limit_key)
            return jsonify({"error": "Verification token does not match this phone number."}), 400
        clear_auth_failures(limit_key)
        return jsonify({"ok": True})
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


@app.route("/api/auth/reauthenticate", methods=["POST", "OPTIONS"])
def auth_reauthenticate():
    preflight = handle_options()
    if preflight:
        return preflight
    session_token, auth_user, error = require_authenticated_user()
    if error:
        return error
    data, error = require_json()
    if error:
        return error
    password = str(data.get("password", "") or "")
    if not password:
        return jsonify({"error": "Password is required"}), 400
    passwords = load_password_store()
    email = str(auth_user.get("email", "") or "").strip().lower()
    username = str(auth_user.get("loginUsername", "") or "").strip().lower()
    stored_password = passwords.get(email)
    if username:
        stored_password = passwords.get(agent_password_key(username)) or stored_password
    if not stored_password or not verify_password(stored_password, password):
        record_audit_log(auth_user, "SENSITIVE_REAUTH_FAILED", {"userId": auth_user["id"]})
        return jsonify({"error": "Password is incorrect"}), 401
    sessions = load_sessions()
    session_key = session_token_hash(session_token)
    session = sessions.get(session_key) or sessions.get(session_token)
    if not session:
        return jsonify({"error": "Invalid or expired session"}), 401
    session["recentAuthAt"] = now_seconds()
    session["lastActivityAt"] = now_seconds()
    sessions[session_key] = session
    sessions.pop(session_token, None)
    save_sessions(sessions)
    record_audit_log(auth_user, "SENSITIVE_REAUTH_SUCCESS", {"userId": auth_user["id"]})
    return jsonify({"ok": True})


@app.route("/api/auth/logout", methods=["POST", "OPTIONS"])
def auth_logout():
    preflight = handle_options()
    if preflight:
        return preflight
    token, auth_user, error = require_authenticated_user()
    if error:
        return error
    set_user_last_seen(auth_user["id"], now_ms())
    store = prune_presence(load_presence_store())
    store.pop(auth_user["id"], None)
    save_presence_store(store)
    revoke_session(token)
    record_audit_log(auth_user, "LOGOUT", staff_audit_target(auth_user))
    return clear_session_cookie(jsonify({"ok": True}))


@app.route("/api/auth/trusted-device/forget", methods=["POST", "OPTIONS"])
def auth_forget_trusted_device():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    removed = revoke_current_trusted_device()
    record_audit_log(auth_user, "FORGET_TRUSTED_DEVICE", {"removed": removed})
    return clear_trusted_device_cookie(jsonify({"ok": True, "removed": removed}))


@app.route("/api/auth/sessions/revoke-all", methods=["POST", "OPTIONS"])
def auth_revoke_all_sessions():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    record_audit_log(auth_user, "REVOKE_ALL_SESSIONS", staff_audit_target(auth_user))
    revoke_user_sessions(auth_user["id"])
    response = clear_session_cookie(jsonify({"ok": True}))
    return clear_trusted_device_cookie(response)


@app.route("/api/auth/me", methods=["POST", "OPTIONS"])
def auth_me():
    preflight = handle_options()
    if preflight:
        return preflight
    _, auth_user, error = require_authenticated_user()
    if error:
        return error
    return jsonify({"ok": True, "user": auth_user})


@app.route("/api/auth/request-password-reset", methods=["POST", "OPTIONS"])
def auth_request_password_reset():
    preflight = handle_options()
    if preflight:
        return preflight
    data, error = require_json()
    if error:
        return error
    try:
        email = validate_email(str(data.get("email", "")))
        limit_key = rate_limit_key("password-reset", email)
        if auth_rate_limited(limit_key):
            return jsonify({"error": "Too many reset requests. Please wait 15 minutes and try again."}), 429

        users = load_user_store()
        user = find_user_by_email(users, email)
        if not user:
            record_auth_failure(limit_key)
            return jsonify({"ok": True})

        token = secrets.token_urlsafe(32)
        reset_url = build_reset_url("", token)
        tokens = load_reset_tokens()
        tokens[token] = {
            "email": email,
            "expiresAt": int(time.time()) + int(load_portal_settings_store()["passwordResetMinutes"]) * 60,
        }
        save_reset_tokens(tokens)
        send_password_reset_link_email(email, reset_url)
        clear_auth_failures(limit_key)
        record_audit_log(None, "REQUEST_PASSWORD_RESET", staff_audit_target(user))
        return jsonify({"ok": True})
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        app.logger.exception("Password reset email failed")
        return jsonify({"error": f"Email could not be sent: {exc}"}), 500


@app.route("/api/auth/password-reset", methods=["POST", "OPTIONS"])
def auth_password_reset():
    preflight = handle_options()
    if preflight:
        return preflight
    data, error = require_json()
    if error:
        return error
    token = str(data.get("token", "")).strip()
    new_password = str(data.get("newPasswordHash", ""))
    if not token:
        return jsonify({"error": "token is required"}), 400
    if not new_password:
        return jsonify({"error": "Password is required"}), 400
    if len(new_password) < 8:
        return jsonify({"error": "Password must be at least 8 characters"}), 400

    tokens = load_reset_tokens()
    entry = tokens.get(token)
    if not entry:
        return jsonify({"error": "Invalid or expired reset token"}), 400

    email = entry["email"]
    users = load_user_store()
    if not find_user_by_email(users, email):
        return jsonify({"error": "Invalid or expired reset token"}), 400

    passwords = load_password_store()
    passwords[email] = hash_password_for_storage(new_password)
    tokens.pop(token, None)
    save_password_store(passwords)
    save_reset_tokens(tokens)
    user = find_user_by_email(users, email)
    if user:
        revoke_user_sessions(user["id"])
        record_audit_log(None, "COMPLETE_PASSWORD_RESET", staff_audit_target(user))
    return jsonify({"ok": True})


# Legacy announcement, forms, and training endpoints were removed.


@app.route("/", defaults={"path": ""}, methods=["GET"])
@app.route("/<path:path>", methods=["GET"])
def serve_frontend(path: str):
    requested = str(path or "").strip().lstrip("/")
    if not os.path.isdir(FRONTEND_PUBLIC_DIR):
        return jsonify({"error": "Frontend build is not installed on this server."}), 404

    if requested:
        candidate = os.path.join(FRONTEND_PUBLIC_DIR, requested)
        if os.path.isfile(candidate):
            return send_from_directory(FRONTEND_PUBLIC_DIR, requested, conditional=True)

    index_path = os.path.join(FRONTEND_PUBLIC_DIR, "index.html")
    if os.path.isfile(index_path):
        return send_from_directory(FRONTEND_PUBLIC_DIR, "index.html", conditional=True)

    return jsonify({"error": "Frontend entry point not found."}), 404


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "4185")))


